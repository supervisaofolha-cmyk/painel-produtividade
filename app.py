import csv
import base64
import json
import locale
import os
import pathlib
import re
import shutil
import sqlite3
import tempfile
import time
import unicodedata
import uuid
import zipfile
import builtins
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from difflib import SequenceMatcher
from html import escape, unescape
from html.parser import HTMLParser
from io import BytesIO, StringIO
import calendar
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st

try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:
    psycopg = None
    dict_row = None


ARQUIVO_PRODUTIVIDADE = "produtividade.xlsx"
ARQUIVO_PRODUTIVIDADE_BACKUP = "produtividade_backup.xlsx"
ARQUIVO_USUARIOS = "usuarios.xlsx"
ARQUIVO_LISTA_APOIO = "lista_apoio.xlsx"
ARQUIVO_LISTA_APOIO_BACKUP = "lista_apoio_backup.xlsx"
ARQUIVO_LISTA_APOIO_DB = "lista_apoio.db"
ENV_LISTA_APOIO_DATABASE_URL = "LISTA_APOIO_DATABASE_URL"
LISTA_APOIO_DATABASE_URL_PADRAO = "postgresql://postgres.pogiuykdubgvcjjzjihs:TgKyzWZhAXJjczXd@aws-1-us-west-2.pooler.supabase.com:6543/postgres?sslmode=require"
ARQUIVO_LISTA_APOIO_HISTORICO = "lista_apoio_historico.jsonl"
ARQUIVO_META_BACKUP_LISTA_APOIO = "lista_apoio_backup_meta.json"
PASTA_BACKUP_LISTA_APOIO = "backups_lista_apoio"
ARQUIVO_LOG_BACKUP_PAINEL = "painel_backup_db_erro.log"
ABA_PRODUTIVIDADE = "Produtividade"
ABA_ALIASES = "Aliases"
COLUNA_DIAS_META = "Dias Meta"
COLUNA_ABANDONADAS = "Abandonadas"
CABECALHOS_LISTA_APOIO = [
    "Carimbo de data/hora",
    "Nome do Técnico",
    "Selecione o tópico",
    "Descreva o problema/pedido em poucas palavras",
    "Situação",
    "Responsável/Apoio",
]
TOPICOS_LISTA_APOIO = [
    "13° Salário",
    "Integração",
    "Provisão de Férias",
    "Provisão de 13°",
    "Crédito em conta",
    "Fórmula",
    "Configuração de Rubrica",
    "Médias",
    "Afastamentos",
    "Rescisão",
    "Férias",
    "Folha Mensal",
    "Folha Complementar",
    "Alteração Cadastral",
    "Alteração Salarial",
    "DCTFWEB",
    "FGTS Digital",
    "Tranferencia",
    "Informativos",
    "Cadastro",
    "e-Social",
    "Outros",
]
USUARIOS_APOIO = {"subbrenda", "subluma"}
FUSO_HORARIO_APP = ZoneInfo("America/Araguaina")
PLUG_CHATBOX_URL = "https://tr.plugsocial.com.br/#/app/chatbox"
PLUG_PERFIL_DIR = pathlib.Path(".playwright-plug")
PLUG_GRUPO_FOLHA = "Folha de Pagamento"
PLUG_STATUS_FECHADO = "Fechado"

POWERBI_RESOURCE_KEY = "6b54dc9f-c2f8-4ee5-bbd2-e2ca5781ab06"
POWERBI_API_BASE = "https://wabi-brazil-south-b-primary-api.analysis.windows.net"
X2_CONTROLLER_BASE_URL = "http://192.168.1.252/x2-controller"
POWERBI_FILA_FALLBACK = "Folha - FGTS/DCTF/ESOCIAL"

SGD_BASE_URL = "https://sgd.dominiosistemas.com.br"
SGD_LOGIN_URL = f"{SGD_BASE_URL}/login"
SGD_RELATORIO_URL = f"{SGD_BASE_URL}/sgsc/faces/rel-satisfacao.html"
SGD_RELATORIO_TIMEOUT = 300
SGD_RELATORIO_TENTATIVAS = 3
RO_DRIVE_FOLDER_URL = "https://drive.google.com/drive/folders/1O-1uJ6D9al9piHgOv_Ju0fpL-3Xbz3zK"
RO_PLANILHAS_FIXAS = {
    (2026, 6): "https://docs.google.com/spreadsheets/d/1AZKhjWnIS_hys9B9oaeazcSUP0wS6ExX8U9wRdrktRs/edit?resourcekey#gid=148357412",
}
SERVICO_PLANILHA_LOCAL_URL = "http://127.0.0.1:8765"

COR_LARANJA = "#F97316"
COR_CINZA = "#6B7280"
COR_BRANCO = "#FFFFFF"
FUNDO = "#F5F5F5"

CORES_STATUS = {
    "CRÍTICO": "#DC2626",
    "ATENÇÃO": "#F97316",
    "BOM": "#2563EB",
    "EXCELENTE": "#16A34A",
}

MESES_POWERBI = {
    1: "janeiro",
    2: "fevereiro",
    3: "março",
    4: "abril",
    5: "maio",
    6: "junho",
    7: "julho",
    8: "agosto",
    9: "setembro",
    10: "outubro",
    11: "novembro",
    12: "dezembro",
}

MESES_RO = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Marco",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}

MOTIVOS_RO_CONTAM = {
    "Ligação com duração inferior a 2 minutos/Retorno solicitado pela supervisão",
    "Retorno solicitado pela Supervisão",
    "Ligação transferida para outro técnico ou setor (Transferência)",
}

TECNICOS_DESCONSIDERADOS_ESPERADO = {
    "patricia karla sousa araujo",
    "lorena dias araujo",
    "lucas luiz romero",
}

META_ESPERADA_POR_NIVEL = {
    "Técnico III": 35,
    "Técnico II": 28,
    "Técnico I": 23,
    "JR": 20,
    "Estágio": 10,
}

FERIADOS_FEDERAIS_FIXOS = {
    (1, 1),
    (4, 21),
    (5, 1),
    (9, 7),
    (10, 12),
    (11, 2),
    (11, 15),
    (11, 20),
    (12, 25),
}

FERIADOS_ADICIONAIS = {
    date(2026, 6, 4),
}

PROGRAMACAO_FERIAS = [
    ("Davylla Rodrigues Freitas Silva", date(2026, 5, 18), date(2026, 6, 1)),
    ("Damyllis Lorraine de Oliveira Goncalves", date(2026, 5, 18), date(2026, 6, 1)),
    ("Francilene Ferreira de Jesus", date(2026, 5, 18), date(2026, 6, 16)),
    ("Janaina Lima dos Reis Urany", date(2026, 6, 1), date(2026, 6, 30)),
    ("Kerollen Cristielly de Jesus Siqueira", date(2026, 6, 1), date(2026, 6, 30)),
    ("Joao Frutuoso Machado Neto", date(2026, 6, 15), date(2026, 6, 29)),
    ("Cristiane Ramos da Silva Rocha", date(2026, 6, 15), date(2026, 6, 29)),
    ("Brenda de Menezes Silva", date(2026, 6, 15), date(2026, 6, 29)),
    ("Jessica Fe Moura", date(2026, 5, 25), date(2026, 6, 23)),
    ("Alisson de Freitas Silva", date(2026, 7, 1), date(2026, 7, 15)),
    ("Emilly Kamilly Medeiros Miranda", date(2026, 7, 1), date(2026, 7, 15)),
    ("Jaqueline Evangelista da Silva Martins", date(2026, 7, 1), date(2026, 7, 10)),
    ("Davylla Rodrigues Freitas Silva", date(2026, 7, 15), date(2026, 7, 29)),
    ("Ana Paula Da Luz Silva Aleixo", date(2026, 7, 15), date(2026, 7, 29)),
    ("SubLuma", date(2026, 7, 16), date(2026, 7, 30)),
    ("Lizandra Gomes Duarte", date(2026, 8, 3), date(2026, 8, 17)),
    ("Matheus Farias De Souza", date(2026, 8, 3), date(2026, 8, 17)),
    ("Rafael Gomes de Morais", date(2026, 8, 17), date(2026, 8, 31)),
    ("Sarah Steffanie de Lima Borges", date(2026, 8, 17), date(2026, 9, 15)),
    ("Brenda Oliveira", date(2026, 8, 17), date(2026, 8, 30)),
    ("Francielson de Oliveira", date(2026, 9, 1), date(2026, 9, 15)),
    ("Lucas Luiz Romero", date(2026, 9, 1), date(2026, 9, 15)),
    ("Brenda de Menezes Silva", date(2026, 9, 16), date(2026, 9, 30)),
    ("Maysa Victoria Dias Moura", date(2026, 9, 16), date(2026, 10, 16)),
    ("SubLuma", date(2026, 9, 8), date(2026, 9, 17)),
    ("Brenda Oliveira", date(2026, 9, 21), date(2026, 9, 26)),
    ("Leandro Oliveira de Sousa", date(2026, 10, 1), date(2026, 10, 30)),
    ("Ana Karollyne Souza Faria", date(2026, 10, 1), date(2026, 10, 15)),
    ("Damyllis Lorraine de Oliveira Goncalves", date(2026, 10, 19), date(2026, 11, 2)),
    ("Carlos Mateus Cassimiro Nunes", date(2026, 10, 19), date(2026, 11, 17)),
    ("Esther", date(2026, 10, 19), date(2026, 11, 1)),
    ("Joao Frutuoso Machado Neto", date(2026, 11, 3), date(2026, 11, 17)),
    ("Lizandra Gomes Duarte", date(2026, 11, 3), date(2026, 11, 17)),
    ("Esther", date(2026, 12, 23), date(2026, 12, 29)),
    ("SubLuma", date(2026, 12, 21), date(2026, 12, 25)),
    ("Rafael Gomes de Morais", date(2026, 12, 21), date(2027, 1, 4)),
    ("Matheus Farias De Souza", date(2026, 12, 21), date(2027, 1, 4)),
    ("Ana Paula Da Luz Silva Aleixo", date(2026, 12, 21), date(2027, 1, 4)),
    ("Emilly Kamilly Medeiros Miranda", date(2026, 12, 21), date(2027, 1, 4)),
    ("Isabella Alves Queiroz", date(2026, 12, 21), date(2027, 1, 4)),
    ("Alisson de Freitas Silva", date(2026, 12, 21), date(2027, 1, 4)),
    ("Jaqueline Evangelista da Silva Martins", date(2026, 12, 21), date(2027, 1, 4)),
    ("Isabella Borges de Oliveira", date(2026, 12, 21), date(2027, 1, 4)),
]

PROGRAMACAO_LICENCAS = [
    (
        "Jessica Faria da Silveira",
        date(2026, 5, 31),
        date(2026, 9, 27),
        "Licença-maternidade",
    ),
]

CARGA_DIARIA_PADRAO_MINUTOS = 8 * 60 + 30

PROGRAMACAO_AUSENCIAS = [
    ("Jessica Faria da Silveira", date(2026, 5, 29), "Atestado", None),
    ("Karina Gonçalves Martins", date(2026, 5, 29), "Falta/Atraso", None),
    ("Karina Gonçalves Martins", date(2026, 6, 2), "Atestado", None),
    ("Leandro de Souza Silva", date(2026, 6, 1), "Energia", 4 * 60 + 30),
    ("Sarah Steffanie de Lima Borges", date(2026, 6, 1), "Atestado", None),
    ("Milena Sales", date(2026, 6, 2), "Atestado", None),
    ("Sarah Steffanie de Lima Borges", date(2026, 6, 2), "Atestado", None),
    ("Sarah Steffanie de Lima Borges", date(2026, 6, 3), "Atestado", None),
    ("Leandro de Souza Silva", date(2026, 6, 5), "Energia", 4 * 60),
    ("Paulo Ricardo Santos", date(2026, 6, 5), "Falta/Atraso", None),
    ("Carlos Mateus Cassimiro Nunes", date(2026, 6, 5), "Falta/Atraso", None),
    ("Anna Luiza Rezende Tavares", date(2026, 6, 8), "Falta/Atraso", 4 * 60),
    ("Leandro de Souza Silva", date(2026, 6, 8), "Atestado", None),
    ("Milena Sales", date(2026, 6, 9), "Atestado", 4 * 60),
    ("Leandro de Souza Silva", date(2026, 6, 9), "Atestado", None),
    ("Maria Eduarda Sousa Costa", date(2026, 6, 9), "Atestado", None),
    ("Maria Eduarda Sousa Costa", date(2026, 6, 10), "Atestado", None),
    ("Sarah Steffanie de Lima Borges", date(2026, 6, 10), "Atestado", 4 * 60 + 30),
    ("Daniel Gomes da Silva", date(2026, 6, 10), "Falta/Atraso", 3 * 60),
]


class FormularioSGDParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.no_formulario = False
        self.inputs = []
        self.selects = []
        self._select_atual = None
        self._opcoes_select = []

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == "form" and attrs.get("id") == "formFiltroRelatorio":
            self.no_formulario = True

        if not self.no_formulario:
            return

        if tag == "input":
            self.inputs.append(attrs)
        elif tag == "select":
            self._select_atual = attrs
            self._opcoes_select = []
        elif tag == "option" and self._select_atual is not None:
            self._opcoes_select.append(attrs)

    def handle_endtag(self, tag):
        if self.no_formulario and tag == "select":
            opcao = next(
                (item for item in self._opcoes_select if "selected" in item),
                self._opcoes_select[0] if self._opcoes_select else {"value": ""},
            )
            self.selects.append((self._select_atual, opcao))
            self._select_atual = None
            self._opcoes_select = []

        if tag == "form" and self.no_formulario:
            self.no_formulario = False


def carregar_env_local():
    return _carregar_env_local_cache(assinatura_arquivo(".env"))


@st.cache_data(show_spinner=False)
def _carregar_env_local_cache(_assinatura_env):
    valores = {}
    if not os.path.exists(".env"):
        return valores

    with open(".env", "r", encoding="utf-8") as arquivo:
        for linha in arquivo:
            linha = linha.strip()
            if not linha or linha.startswith("#") or "=" not in linha:
                continue
            chave, valor = linha.split("=", 1)
            valores[chave.strip()] = valor.strip().strip('"').strip("'")

    return valores


def assinatura_arquivo(caminho):
    if not os.path.exists(caminho):
        return "ausente"
    return builtins.str(os.path.getmtime(caminho))


def arquivo_excel_integro(caminho):
    if not os.path.exists(caminho):
        return False

    try:
        with zipfile.ZipFile(caminho) as arquivo_zip:
            return arquivo_zip.testzip() is None
    except Exception:
        return False


def garantir_planilha_produtividade_integra():
    principal_ok = arquivo_excel_integro(ARQUIVO_PRODUTIVIDADE)
    backup_ok = arquivo_excel_integro(ARQUIVO_PRODUTIVIDADE_BACKUP)

    if principal_ok:
        if not backup_ok:
            shutil.copyfile(ARQUIVO_PRODUTIVIDADE, ARQUIVO_PRODUTIVIDADE_BACKUP)
        return

    if backup_ok:
        shutil.copyfile(ARQUIVO_PRODUTIVIDADE_BACKUP, ARQUIVO_PRODUTIVIDADE)
        return

    raise zipfile.BadZipFile(
        "Nenhuma cópia íntegra da produtividade.xlsx foi encontrada."
    )


def ler_dataframe_produtividade():
    garantir_planilha_produtividade_integra()
    return _ler_dataframe_produtividade_cache(assinatura_arquivo(ARQUIVO_PRODUTIVIDADE))


@st.cache_data(show_spinner=False)
def _ler_dataframe_produtividade_cache(_assinatura_produtividade):
    return pd.read_excel(ARQUIVO_PRODUTIVIDADE)


def ler_dataframe_usuarios():
    return _ler_dataframe_usuarios_cache(assinatura_arquivo(ARQUIVO_USUARIOS))


@st.cache_data(show_spinner=False)
def _ler_dataframe_usuarios_cache(_assinatura_usuarios):
    return pd.read_excel(ARQUIVO_USUARIOS)


def carregar_workbook_produtividade(**kwargs):
    garantir_planilha_produtividade_integra()
    return openpyxl.load_workbook(ARQUIVO_PRODUTIVIDADE, **kwargs)


def salvar_workbook_produtividade(wb):
    garantir_planilha_produtividade_integra()
    diretorio = os.path.dirname(os.path.abspath(ARQUIVO_PRODUTIVIDADE)) or "."
    descritor, caminho_temporario = tempfile.mkstemp(
        prefix="produtividade_",
        suffix=".xlsx",
        dir=diretorio,
    )
    os.close(descritor)
    try:
        wb.save(caminho_temporario)
        with zipfile.ZipFile(caminho_temporario) as arquivo_zip:
            if arquivo_zip.testzip() is not None:
                raise zipfile.BadZipFile("Arquivo temporário gerado com corrupção.")
        os.replace(caminho_temporario, ARQUIVO_PRODUTIVIDADE)
        shutil.copyfile(ARQUIVO_PRODUTIVIDADE, ARQUIVO_PRODUTIVIDADE_BACKUP)
        tentar_salvar_backup_arquivo_painel_no_banco(ARQUIVO_PRODUTIVIDADE)
    finally:
        if os.path.exists(caminho_temporario):
            os.remove(caminho_temporario)


def garantir_planilha_lista_apoio_integra():
    principal_existe = os.path.exists(ARQUIVO_LISTA_APOIO)
    principal_ok = arquivo_excel_integro(ARQUIVO_LISTA_APOIO)
    backup_ok = arquivo_excel_integro(ARQUIVO_LISTA_APOIO_BACKUP)

    if principal_ok:
        if not backup_ok:
            shutil.copyfile(ARQUIVO_LISTA_APOIO, ARQUIVO_LISTA_APOIO_BACKUP)
        return

    if backup_ok:
        shutil.copyfile(ARQUIVO_LISTA_APOIO_BACKUP, ARQUIVO_LISTA_APOIO)
        return

    if principal_existe:
        raise zipfile.BadZipFile(
            "Nenhuma cópia íntegra da lista_apoio.xlsx foi encontrada."
        )


def chave_registro_lista_apoio(registro):
    coluna_carimbo = CABECALHOS_LISTA_APOIO[0]
    coluna_tecnico = CABECALHOS_LISTA_APOIO[1]
    coluna_topico = CABECALHOS_LISTA_APOIO[2]
    coluna_descricao = CABECALHOS_LISTA_APOIO[3]
    return "||".join(
        [
            texto_lista_apoio(valor_campo_registro_lista_apoio(registro, coluna_carimbo)),
            texto_lista_apoio(valor_campo_registro_lista_apoio(registro, coluna_tecnico)),
            texto_lista_apoio(valor_campo_registro_lista_apoio(registro, coluna_topico)),
            texto_lista_apoio(
                valor_campo_registro_lista_apoio(
                    registro,
                    coluna_descricao,
                )
            ),
        ]
    )


def dataframe_lista_apoio_vazio():
    return pd.DataFrame(columns=CABECALHOS_LISTA_APOIO)


def valor_campo_registro_lista_apoio(registro, coluna_esperada):
    if coluna_esperada in registro:
        return registro.get(coluna_esperada, "")

    chave_esperada = normalizar_cabecalho(coluna_esperada)
    for chave, valor in registro.items():
        if normalizar_cabecalho(chave) == chave_esperada:
            return valor
    return ""


def padronizar_dataframe_lista_apoio(dataframe):
    if dataframe is None or dataframe.empty:
        return dataframe_lista_apoio_vazio()

    dataframe = dataframe.copy()
    for coluna_lista in CABECALHOS_LISTA_APOIO:
        if coluna_lista not in dataframe.columns:
            chave_esperada = normalizar_cabecalho(coluna_lista)
            coluna_equivalente = next(
                (
                    nome_coluna
                    for nome_coluna in dataframe.columns
                    if normalizar_cabecalho(nome_coluna) == chave_esperada
                ),
                None,
            )
            if coluna_equivalente is not None:
                dataframe[coluna_lista] = dataframe[coluna_equivalente]
            else:
                dataframe[coluna_lista] = ""
        dataframe[coluna_lista] = dataframe[coluna_lista].astype("object")

    dataframe = dataframe[CABECALHOS_LISTA_APOIO]
    colunas_obrigatorias = [
        "Carimbo de data/hora",
        "Nome do Técnico",
        "Selecione o tópico",
        "Descreva o problema/pedido em poucas palavras",
    ]
    for coluna_obrigatoria in colunas_obrigatorias:
        dataframe[coluna_obrigatoria] = dataframe[coluna_obrigatoria].map(
            texto_lista_apoio
        )

    filtro_valido = dataframe["Carimbo de data/hora"].ne("")
    filtro_valido &= dataframe["Nome do Técnico"].ne("")
    filtro_valido &= dataframe["Selecione o tópico"].ne("")
    filtro_valido &= dataframe["Descreva o problema/pedido em poucas palavras"].ne("")
    dataframe = dataframe.loc[filtro_valido].reset_index(drop=True)

    return dataframe


def ler_dataframe_lista_apoio_arquivo(caminho):
    if not os.path.exists(caminho) or not arquivo_excel_integro(caminho):
        return dataframe_lista_apoio_vazio()

    try:
        dataframe = pd.read_excel(caminho)
    except Exception:
        return dataframe_lista_apoio_vazio()

    return padronizar_dataframe_lista_apoio(dataframe)


def registrar_historico_lista_apoio(acao, registro):
    os.makedirs(PASTA_BACKUP_LISTA_APOIO, exist_ok=True)
    payload = {
        "timestamp": datetime.now(FUSO_HORARIO_APP).strftime("%Y-%m-%d %H:%M:%S"),
        "acao": acao,
        "registro": {
            coluna: texto_lista_apoio(registro.get(coluna, ""))
            for coluna in CABECALHOS_LISTA_APOIO
        },
    }
    with open(ARQUIVO_LISTA_APOIO_HISTORICO, "a", encoding="utf-8") as arquivo:
        arquivo.write(json.dumps(payload, ensure_ascii=False) + "\n")


def restaurar_lista_apoio_do_historico():
    if not os.path.exists(ARQUIVO_LISTA_APOIO_HISTORICO):
        return dataframe_lista_apoio_vazio()

    registros = {}
    with open(ARQUIVO_LISTA_APOIO_HISTORICO, "r", encoding="utf-8") as arquivo:
        for linha in arquivo:
            linha = linha.strip()
            if not linha:
                continue
            try:
                evento = json.loads(linha)
            except Exception:
                continue
            registro = evento.get("registro") or {}
            chave = chave_registro_lista_apoio(registro)
            if not chave:
                continue
            registros[chave] = {
                coluna: texto_lista_apoio(
                    valor_campo_registro_lista_apoio(registro, coluna)
                )
                for coluna in CABECALHOS_LISTA_APOIO
            }

    if not registros:
        return dataframe_lista_apoio_vazio()

    dataframe = pd.DataFrame(list(registros.values()))
    return padronizar_dataframe_lista_apoio(dataframe)


def mesclar_dataframes_lista_apoio(dataframes):
    registros = {}
    for dataframe in dataframes:
        dataframe = padronizar_dataframe_lista_apoio(dataframe)
        if dataframe.empty:
            continue
        for _, linha in dataframe.iterrows():
            registro = {
                coluna: texto_lista_apoio(linha.get(coluna, ""))
                for coluna in CABECALHOS_LISTA_APOIO
            }
            chave = chave_registro_lista_apoio(registro)
            if not chave:
                continue
            registros[chave] = registro

    if not registros:
        return dataframe_lista_apoio_vazio()

    return padronizar_dataframe_lista_apoio(pd.DataFrame(list(registros.values())))


def database_url_lista_apoio():
    env_local = carregar_env_local()
    segredo = ""
    try:
        segredo = builtins.str(st.secrets.get(ENV_LISTA_APOIO_DATABASE_URL, "") or "").strip()
    except Exception:
        segredo = ""
    url = builtins.str(
        segredo
        or os.getenv(
            ENV_LISTA_APOIO_DATABASE_URL,
            env_local.get(ENV_LISTA_APOIO_DATABASE_URL, ""),
        )
        or LISTA_APOIO_DATABASE_URL_PADRAO
        or ""
    ).strip()
    if "pooler.supabase.com:6543/" in url:
        url = url.replace(
            "pooler.supabase.com:6543/",
            "pooler.supabase.com:5432/",
        )
    if url and "connect_timeout=" not in url:
        separador = "&" if "?" in url else "?"
        url = f"{url}{separador}connect_timeout=15"
    return url


def backend_lista_apoio():
    url = database_url_lista_apoio()
    if not url:
        return "sqlite"

    esquema = urlparse(url).scheme.lower()
    if esquema.startswith("postgres"):
        return "postgres"
    return "sqlite"


def conexao_lista_apoio_db():
    backend = backend_lista_apoio()
    if backend == "postgres":
        if psycopg is None:
            raise RuntimeError(
                "Banco online da Lista de Apoio configurado, mas a dependência psycopg não está instalada."
            )
        return psycopg.connect(
            database_url_lista_apoio(),
            row_factory=dict_row,
            prepare_threshold=None,
        )

    conexao = sqlite3.connect(ARQUIVO_LISTA_APOIO_DB)
    conexao.row_factory = sqlite3.Row
    return conexao


def valor_backup_painel(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    return valor


def texto_chave_backup_painel(valor):
    valor = valor_backup_painel(valor)
    return builtins.str(valor or "").strip()


def garantir_banco_backup_painel():
    with conexao_lista_apoio_db() as conexao:
        conexao.execute(
            """
            CREATE TABLE IF NOT EXISTS painel_dados_backup (
                id TEXT PRIMARY KEY,
                origem TEXT NOT NULL,
                aba TEXT NOT NULL,
                chave TEXT NOT NULL,
                payload TEXT NOT NULL,
                criado_em TEXT NOT NULL,
                atualizado_em TEXT NOT NULL,
                UNIQUE(origem, aba, chave)
            )
            """
        )
        conexao.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_painel_dados_backup_origem_aba
            ON painel_dados_backup (origem, aba)
            """
        )
        conexao.commit()


def registros_backup_painel_do_workbook(wb):
    registros = []
    abas_permitidas = {ABA_PRODUTIVIDADE, ABA_ALIASES}

    for nome_aba in wb.sheetnames:
        if nome_aba not in abas_permitidas:
            continue

        ws = wb[nome_aba]
        cabecalhos_aba = [
            texto_chave_backup_painel(celula.value)
            for celula in ws[1]
        ]

        for row in range(2, ws.max_row + 1):
            valores = [
                valor_backup_painel(ws.cell(row=row, column=coluna).value)
                for coluna in range(1, len(cabecalhos_aba) + 1)
            ]
            if not any(texto_chave_backup_painel(valor) for valor in valores):
                continue

            payload = {
                cabecalho: valor
                for cabecalho, valor in zip(cabecalhos_aba, valores)
                if cabecalho
            }

            if nome_aba == ABA_PRODUTIVIDADE:
                chave = "||".join(
                    [
                        texto_chave_backup_painel(payload.get("Data")),
                        texto_chave_backup_painel(payload.get("Técnico")),
                    ]
                )
            elif nome_aba == ABA_ALIASES:
                chave = texto_chave_backup_painel(
                    payload.get("Técnico Planilha")
                    or payload.get("T??cnico Planilha")
                    or payload.get("TÃ©cnico Planilha")
                )
            else:
                chave = ""

            if not chave or chave == "||":
                chave = f"linha_{row}"

            registros.append(
                {
                    "origem": ARQUIVO_PRODUTIVIDADE,
                    "aba": nome_aba,
                    "chave": chave,
                    "payload": json.dumps(payload, ensure_ascii=False, default=builtins.str),
                }
            )

    return registros


def salvar_backup_painel_no_banco(wb):
    registros = registros_backup_painel_do_workbook(wb)
    if not registros:
        return 0

    garantir_banco_backup_painel()
    backend = backend_lista_apoio()
    agora = datetime.now(FUSO_HORARIO_APP).strftime("%Y-%m-%d %H:%M:%S")
    consulta = """
        INSERT INTO painel_dados_backup (
            id,
            origem,
            aba,
            chave,
            payload,
            criado_em,
            atualizado_em
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT(origem, aba, chave) DO UPDATE SET
            payload = excluded.payload,
            atualizado_em = excluded.atualizado_em
    """
    if backend != "postgres":
        consulta = """
            INSERT INTO painel_dados_backup (
                id,
                origem,
                aba,
                chave,
                payload,
                criado_em,
                atualizado_em
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(origem, aba, chave) DO UPDATE SET
                payload = excluded.payload,
                atualizado_em = excluded.atualizado_em
        """

    with conexao_lista_apoio_db() as conexao:
        for registro in registros:
            identificador = builtins.str(
                uuid.uuid5(
                    uuid.NAMESPACE_URL,
                    f"{registro['origem']}|{registro['aba']}|{registro['chave']}",
                )
            )
            conexao.execute(
                consulta,
                (
                    identificador,
                    registro["origem"],
                    registro["aba"],
                    registro["chave"],
                    registro["payload"],
                    agora,
                    agora,
                ),
            )
        conexao.commit()

    return len(registros)


def registrar_erro_backup_painel(erro):
    try:
        with open(ARQUIVO_LOG_BACKUP_PAINEL, "a", encoding="utf-8") as arquivo:
            timestamp = datetime.now(FUSO_HORARIO_APP).strftime("%Y-%m-%d %H:%M:%S")
            arquivo.write(f"{timestamp} - {erro}\n")
    except Exception:
        pass


def tentar_salvar_backup_painel_no_banco(wb):
    try:
        return salvar_backup_painel_no_banco(wb)
    except Exception as erro:
        registrar_erro_backup_painel(erro)
        return 0


def garantir_banco_arquivo_backup_painel():
    with conexao_lista_apoio_db() as conexao:
        conexao.execute(
            """
            CREATE TABLE IF NOT EXISTS painel_arquivo_backup (
                origem TEXT PRIMARY KEY,
                conteudo_base64 TEXT NOT NULL,
                tamanho_bytes INTEGER NOT NULL,
                atualizado_em TEXT NOT NULL
            )
            """
        )
        conexao.commit()


def salvar_backup_arquivo_painel_no_banco(caminho_arquivo):
    if not os.path.exists(caminho_arquivo):
        return 0

    garantir_banco_arquivo_backup_painel()
    origem = os.path.basename(caminho_arquivo)
    with open(caminho_arquivo, "rb") as arquivo:
        conteudo = arquivo.read()

    conteudo_base64 = base64.b64encode(conteudo).decode("ascii")
    agora = datetime.now(FUSO_HORARIO_APP).strftime("%Y-%m-%d %H:%M:%S")
    backend = backend_lista_apoio()
    consulta = """
        INSERT INTO painel_arquivo_backup (
            origem,
            conteudo_base64,
            tamanho_bytes,
            atualizado_em
        ) VALUES (%s, %s, %s, %s)
        ON CONFLICT(origem) DO UPDATE SET
            conteudo_base64 = excluded.conteudo_base64,
            tamanho_bytes = excluded.tamanho_bytes,
            atualizado_em = excluded.atualizado_em
    """
    if backend != "postgres":
        consulta = """
            INSERT INTO painel_arquivo_backup (
                origem,
                conteudo_base64,
                tamanho_bytes,
                atualizado_em
            ) VALUES (?, ?, ?, ?)
            ON CONFLICT(origem) DO UPDATE SET
                conteudo_base64 = excluded.conteudo_base64,
                tamanho_bytes = excluded.tamanho_bytes,
                atualizado_em = excluded.atualizado_em
        """

    with conexao_lista_apoio_db() as conexao:
        conexao.execute(
            consulta,
            (origem, conteudo_base64, len(conteudo), agora),
        )
        conexao.commit()

    return len(conteudo)


def tentar_salvar_backup_arquivo_painel_no_banco(caminho_arquivo):
    try:
        return salvar_backup_arquivo_painel_no_banco(caminho_arquivo)
    except Exception as erro:
        registrar_erro_backup_painel(erro)
        return 0


def garantir_banco_ro_backup():
    with conexao_lista_apoio_db() as conexao:
        conexao.execute(
            """
            CREATE TABLE IF NOT EXISTS ro_backup (
                id TEXT PRIMARY KEY,
                data_referencia TEXT NOT NULL,
                arquivo TEXT NOT NULL,
                arquivo_id TEXT NOT NULL,
                origem TEXT NOT NULL,
                url TEXT NOT NULL,
                csv_base64 TEXT NOT NULL,
                contagens_json TEXT NOT NULL,
                tecnicos_origem_json TEXT NOT NULL,
                criado_em TEXT NOT NULL,
                atualizado_em TEXT NOT NULL,
                UNIQUE(data_referencia, arquivo_id)
            )
            """
        )
        conexao.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_ro_backup_data
            ON ro_backup (data_referencia)
            """
        )
        conexao.commit()


def salvar_backup_ro_no_banco(data_referencia, arquivo, texto_csv, contagens, tecnicos_origem):
    garantir_banco_ro_backup()
    backend = backend_lista_apoio()
    data_texto = data_referencia.strftime("%Y-%m-%d")
    agora = datetime.now(FUSO_HORARIO_APP).strftime("%Y-%m-%d %H:%M:%S")
    identificador = builtins.str(
        uuid.uuid5(
            uuid.NAMESPACE_URL,
            f"ro|{data_texto}|{arquivo.get('id', '')}",
        )
    )
    csv_base64 = base64.b64encode(texto_csv.encode("utf-8")).decode("ascii")
    consulta = """
        INSERT INTO ro_backup (
            id,
            data_referencia,
            arquivo,
            arquivo_id,
            origem,
            url,
            csv_base64,
            contagens_json,
            tecnicos_origem_json,
            criado_em,
            atualizado_em
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT(data_referencia, arquivo_id) DO UPDATE SET
            arquivo = excluded.arquivo,
            origem = excluded.origem,
            url = excluded.url,
            csv_base64 = excluded.csv_base64,
            contagens_json = excluded.contagens_json,
            tecnicos_origem_json = excluded.tecnicos_origem_json,
            atualizado_em = excluded.atualizado_em
    """
    if backend != "postgres":
        consulta = """
            INSERT INTO ro_backup (
                id,
                data_referencia,
                arquivo,
                arquivo_id,
                origem,
                url,
                csv_base64,
                contagens_json,
                tecnicos_origem_json,
                criado_em,
                atualizado_em
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(data_referencia, arquivo_id) DO UPDATE SET
                arquivo = excluded.arquivo,
                origem = excluded.origem,
                url = excluded.url,
                csv_base64 = excluded.csv_base64,
                contagens_json = excluded.contagens_json,
                tecnicos_origem_json = excluded.tecnicos_origem_json,
                atualizado_em = excluded.atualizado_em
        """

    with conexao_lista_apoio_db() as conexao:
        conexao.execute(
            consulta,
            (
                identificador,
                data_texto,
                arquivo.get("titulo", ""),
                arquivo.get("id", ""),
                arquivo.get("origem", ""),
                arquivo.get("url", ""),
                csv_base64,
                json.dumps(contagens, ensure_ascii=False, sort_keys=True),
                json.dumps(tecnicos_origem, ensure_ascii=False),
                agora,
                agora,
            ),
        )
        conexao.commit()


def tentar_salvar_backup_ro_no_banco(data_referencia, arquivo, texto_csv, contagens, tecnicos_origem):
    try:
        salvar_backup_ro_no_banco(
            data_referencia,
            arquivo,
            texto_csv,
            contagens,
            tecnicos_origem,
        )
        return True
    except Exception as erro:
        registrar_erro_backup_painel(erro)
        return False


def garantir_banco_lista_apoio():
    with conexao_lista_apoio_db() as conexao:
        conexao.execute(
            """
            CREATE TABLE IF NOT EXISTS lista_apoio (
                id TEXT PRIMARY KEY,
                chave TEXT NOT NULL UNIQUE,
                carimbo_data_hora TEXT NOT NULL,
                nome_tecnico TEXT NOT NULL,
                topico TEXT NOT NULL,
                descricao TEXT NOT NULL,
                situacao TEXT NOT NULL,
                responsavel_apoio TEXT NOT NULL,
                criado_em TEXT NOT NULL,
                atualizado_em TEXT NOT NULL
            )
            """
        )
        conexao.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_lista_apoio_carimbo
            ON lista_apoio (carimbo_data_hora)
            """
        )
        conexao.commit()


def dataframe_para_registros_lista_apoio(dataframe):
    dataframe = padronizar_dataframe_lista_apoio(dataframe)
    registros = []
    for _, linha in dataframe.iterrows():
        registro = {
            coluna: texto_lista_apoio(linha.get(coluna, ""))
            for coluna in CABECALHOS_LISTA_APOIO
        }
        if chave_registro_lista_apoio(registro):
            registros.append(registro)
    return registros


def deduplicar_dataframe_lista_apoio(dataframe):
    dataframe = padronizar_dataframe_lista_apoio(dataframe)
    if dataframe.empty:
        return dataframe

    registros_por_chave = {}
    for _, linha in dataframe.iterrows():
        registro = {
            coluna: texto_lista_apoio(linha.get(coluna, ""))
            for coluna in CABECALHOS_LISTA_APOIO
        }
        chave = chave_registro_lista_apoio(registro)
        if not chave:
            continue
        registros_por_chave[chave] = registro

    if not registros_por_chave:
        return dataframe_lista_apoio_vazio()

    return padronizar_dataframe_lista_apoio(
        pd.DataFrame(list(registros_por_chave.values()))
    )


def espelhar_lista_apoio_para_arquivos(dataframe):
    dataframe = padronizar_dataframe_lista_apoio(dataframe)
    salvar_dataframe_lista_apoio(dataframe)


def tentar_espelhar_lista_apoio_para_arquivos(dataframe):
    try:
        espelhar_lista_apoio_para_arquivos(dataframe)
    except Exception:
        return False
    return True


def ler_lista_apoio_do_banco():
    garantir_banco_lista_apoio()
    backend = backend_lista_apoio()
    coluna_carimbo = CABECALHOS_LISTA_APOIO[0]
    coluna_tecnico = CABECALHOS_LISTA_APOIO[1]
    coluna_topico = CABECALHOS_LISTA_APOIO[2]
    coluna_descricao = CABECALHOS_LISTA_APOIO[3]
    coluna_situacao = CABECALHOS_LISTA_APOIO[4]
    coluna_responsavel = CABECALHOS_LISTA_APOIO[5]
    consulta = """
        SELECT
            carimbo_data_hora,
            nome_tecnico,
            topico,
            descricao,
            situacao,
            responsavel_apoio
        FROM lista_apoio
    """
    if backend == "postgres":
        consulta += """
        ORDER BY to_timestamp(carimbo_data_hora, 'DD/MM/YYYY HH24:MI:SS') ASC,
                 criado_em ASC
        """
    else:
        consulta += """
        ORDER BY datetime(substr(carimbo_data_hora, 7, 4) || '-' || substr(carimbo_data_hora, 4, 2) || '-' || substr(carimbo_data_hora, 1, 2) || ' ' || substr(carimbo_data_hora, 12, 8)) ASC,
                 rowid ASC
        """
    with conexao_lista_apoio_db() as conexao:
        linhas = conexao.execute(consulta).fetchall()

    if not linhas:
        return dataframe_lista_apoio_vazio()

    registros = []
    for linha in linhas:
        registros.append(
            {
                coluna_carimbo: texto_lista_apoio(linha["carimbo_data_hora"]),
                coluna_tecnico: texto_lista_apoio(linha["nome_tecnico"]),
                coluna_topico: texto_lista_apoio(linha["topico"]),
                coluna_descricao: texto_lista_apoio(linha["descricao"]),
                coluna_situacao: texto_lista_apoio(linha["situacao"]),
                coluna_responsavel: texto_lista_apoio(linha["responsavel_apoio"]),
            }
        )
    return deduplicar_dataframe_lista_apoio(pd.DataFrame(registros))
def salvar_registros_lista_apoio_no_banco(registros):
    garantir_banco_lista_apoio()
    backend = backend_lista_apoio()
    agora = datetime.now(FUSO_HORARIO_APP).strftime("%Y-%m-%d %H:%M:%S")
    coluna_carimbo = CABECALHOS_LISTA_APOIO[0]
    coluna_tecnico = CABECALHOS_LISTA_APOIO[1]
    coluna_topico = CABECALHOS_LISTA_APOIO[2]
    coluna_descricao = CABECALHOS_LISTA_APOIO[3]
    coluna_situacao = CABECALHOS_LISTA_APOIO[4]
    coluna_responsavel = CABECALHOS_LISTA_APOIO[5]
    consulta_busca = "SELECT id FROM lista_apoio WHERE chave = %s"
    consulta_upsert = """
                INSERT INTO lista_apoio (
                    id,
                    chave,
                    carimbo_data_hora,
                    nome_tecnico,
                    topico,
                    descricao,
                    situacao,
                    responsavel_apoio,
                    criado_em,
                    atualizado_em
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT(chave) DO UPDATE SET
                    carimbo_data_hora = excluded.carimbo_data_hora,
                    nome_tecnico = excluded.nome_tecnico,
                    topico = excluded.topico,
                    descricao = excluded.descricao,
                    situacao = excluded.situacao,
                    responsavel_apoio = excluded.responsavel_apoio,
                    atualizado_em = excluded.atualizado_em
                """
    if backend != "postgres":
        consulta_busca = "SELECT id FROM lista_apoio WHERE chave = ?"
        consulta_upsert = """
                INSERT INTO lista_apoio (
                    id,
                    chave,
                    carimbo_data_hora,
                    nome_tecnico,
                    topico,
                    descricao,
                    situacao,
                    responsavel_apoio,
                    criado_em,
                    atualizado_em
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(chave) DO UPDATE SET
                    carimbo_data_hora = excluded.carimbo_data_hora,
                    nome_tecnico = excluded.nome_tecnico,
                    topico = excluded.topico,
                    descricao = excluded.descricao,
                    situacao = excluded.situacao,
                    responsavel_apoio = excluded.responsavel_apoio,
                    atualizado_em = excluded.atualizado_em
                """
    with conexao_lista_apoio_db() as conexao:
        for registro in registros:
            chave = chave_registro_lista_apoio(registro)
            if not chave:
                continue
            identificador_existente = conexao.execute(
                consulta_busca,
                (chave,),
            ).fetchone()
            identificador = (
                identificador_existente["id"] if identificador_existente else builtins.str(uuid.uuid4())
            )
            conexao.execute(
                consulta_upsert,
                (
                    identificador,
                    chave,
                    texto_lista_apoio(valor_campo_registro_lista_apoio(registro, coluna_carimbo)),
                    texto_lista_apoio(valor_campo_registro_lista_apoio(registro, coluna_tecnico)),
                    texto_lista_apoio(valor_campo_registro_lista_apoio(registro, coluna_topico)),
                    texto_lista_apoio(
                        valor_campo_registro_lista_apoio(
                            registro,
                            coluna_descricao,
                        )
                    ),
                    texto_lista_apoio(valor_campo_registro_lista_apoio(registro, coluna_situacao)) or "Aberto",
                    texto_lista_apoio(valor_campo_registro_lista_apoio(registro, coluna_responsavel)),
                    agora,
                    agora,
                ),
            )
        conexao.commit()
    return


def deduplicar_lista_apoio_no_banco():
    garantir_banco_lista_apoio()
    backend = backend_lista_apoio()
    consulta = """
        SELECT
            id,
            chave,
            carimbo_data_hora,
            nome_tecnico,
            topico,
            descricao,
            situacao,
            responsavel_apoio,
            criado_em,
            atualizado_em
        FROM lista_apoio
    """
    with conexao_lista_apoio_db() as conexao:
        linhas = conexao.execute(consulta).fetchall()
        if not linhas:
            return 0

        grupos = {}
        for linha in linhas:
            registro = {
                CABECALHOS_LISTA_APOIO[0]: texto_lista_apoio(linha["carimbo_data_hora"]),
                CABECALHOS_LISTA_APOIO[1]: texto_lista_apoio(linha["nome_tecnico"]),
                CABECALHOS_LISTA_APOIO[2]: texto_lista_apoio(linha["topico"]),
                CABECALHOS_LISTA_APOIO[3]: texto_lista_apoio(linha["descricao"]),
                CABECALHOS_LISTA_APOIO[4]: texto_lista_apoio(linha["situacao"]),
                CABECALHOS_LISTA_APOIO[5]: texto_lista_apoio(linha["responsavel_apoio"]),
            }
            chave_natural = chave_registro_lista_apoio(registro)
            if not chave_natural:
                continue
            grupos.setdefault(chave_natural, []).append(
                {
                    "id": linha["id"],
                    "chave_atual": texto_lista_apoio(linha["chave"]),
                    "chave_natural": chave_natural,
                    "atualizado_em": texto_lista_apoio(linha["atualizado_em"]),
                    "criado_em": texto_lista_apoio(linha["criado_em"]),
                }
            )

        ids_para_excluir = []
        atualizacoes = []
        for chave_natural, grupo in grupos.items():
            grupo_ordenado = sorted(
                grupo,
                key=lambda item: (
                    item["chave_atual"] == chave_natural,
                    item["atualizado_em"],
                    item["criado_em"],
                    item["id"],
                ),
                reverse=True,
            )
            manter = grupo_ordenado[0]
            for duplicado in grupo_ordenado[1:]:
                ids_para_excluir.append(duplicado["id"])
            if manter["chave_atual"] != chave_natural:
                atualizacoes.append((chave_natural, manter["id"]))

        if ids_para_excluir:
            marcador = "%s" if backend == "postgres" else "?"
            consulta_delete = (
                f"DELETE FROM lista_apoio WHERE id IN ({', '.join([marcador] * len(ids_para_excluir))})"
            )
            conexao.execute(consulta_delete, tuple(ids_para_excluir))

        consulta_update = (
            "UPDATE lista_apoio SET chave = %s WHERE id = %s"
            if backend == "postgres"
            else "UPDATE lista_apoio SET chave = ? WHERE id = ?"
        )
        for chave_natural, identificador in atualizacoes:
            conexao.execute(consulta_update, (chave_natural, identificador))

        conexao.commit()
        return len(ids_para_excluir) + len(atualizacoes)


def migrar_lista_apoio_para_banco_se_necessario():
    garantir_banco_lista_apoio()
    with conexao_lista_apoio_db() as conexao:
        linha_total = conexao.execute("SELECT COUNT(*) AS total FROM lista_apoio").fetchone()
        total_banco = linha_total["total"] if linha_total else 0

    if total_banco > 0:
        return

    garantir_planilha_lista_apoio_integra()
    dataframe_principal = ler_dataframe_lista_apoio_arquivo(ARQUIVO_LISTA_APOIO)
    dataframe_backup = ler_dataframe_lista_apoio_arquivo(ARQUIVO_LISTA_APOIO_BACKUP)
    dataframe_historico = restaurar_lista_apoio_do_historico()
    dataframe_mesclado = mesclar_dataframes_lista_apoio(
        [dataframe_principal, dataframe_backup, dataframe_historico]
    )

    registros = dataframe_para_registros_lista_apoio(dataframe_mesclado)
    if registros:
        salvar_registros_lista_apoio_no_banco(registros)


def salvar_snapshot_lista_apoio(caminho_origem, prefixo="lista_apoio"):
    if not os.path.exists(caminho_origem):
        return
    os.makedirs(PASTA_BACKUP_LISTA_APOIO, exist_ok=True)
    timestamp = datetime.now(FUSO_HORARIO_APP).strftime("%Y%m%d_%H%M%S")
    destino = os.path.join(
        PASTA_BACKUP_LISTA_APOIO,
        f"{prefixo}_{timestamp}.xlsx",
    )
    shutil.copyfile(caminho_origem, destino)


def salvar_snapshot_banco_lista_apoio(prefixo="lista_apoio_db"):
    if backend_lista_apoio() != "sqlite":
        return
    if not os.path.exists(ARQUIVO_LISTA_APOIO_DB):
        return
    os.makedirs(PASTA_BACKUP_LISTA_APOIO, exist_ok=True)
    timestamp = datetime.now(FUSO_HORARIO_APP).strftime("%Y%m%d_%H%M%S")
    destino = os.path.join(
        PASTA_BACKUP_LISTA_APOIO,
        f"{prefixo}_{timestamp}.db",
    )
    shutil.copyfile(ARQUIVO_LISTA_APOIO_DB, destino)


def bytes_dataframe_excel(dataframe):
    buffer = BytesIO()
    dataframe = padronizar_dataframe_lista_apoio(dataframe)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False)
    return buffer.getvalue()


def garantir_backup_lista_apoio_por_versao_app():
    if not os.path.exists(ARQUIVO_LISTA_APOIO) and not os.path.exists(ARQUIVO_LISTA_APOIO_DB):
        return

    os.makedirs(PASTA_BACKUP_LISTA_APOIO, exist_ok=True)
    caminho_app = os.path.abspath("app.py")
    assinatura_atual = builtins.str(int(os.path.getmtime(caminho_app)))
    caminho_meta = os.path.join(PASTA_BACKUP_LISTA_APOIO, ARQUIVO_META_BACKUP_LISTA_APOIO)

    assinatura_registrada = ""
    if os.path.exists(caminho_meta):
        try:
            with open(caminho_meta, "r", encoding="utf-8") as arquivo_meta:
                assinatura_registrada = builtins.str(
                    json.load(arquivo_meta).get("assinatura_app", "")
                ).strip()
        except Exception:
            assinatura_registrada = ""

    if assinatura_atual == assinatura_registrada:
        return

    salvar_snapshot_lista_apoio(ARQUIVO_LISTA_APOIO, prefixo="lista_apoio_pre_alteracao_app")
    salvar_snapshot_banco_lista_apoio(prefixo="lista_apoio_db_pre_alteracao_app")

    with open(caminho_meta, "w", encoding="utf-8") as arquivo_meta:
        json.dump(
            {
                "assinatura_app": assinatura_atual,
                "atualizado_em": datetime.now(FUSO_HORARIO_APP).strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
            },
            arquivo_meta,
            ensure_ascii=False,
            indent=2,
        )


def salvar_workbook_lista_apoio(wb):
    diretorio = os.path.dirname(os.path.abspath(ARQUIVO_LISTA_APOIO)) or "."
    descritor, caminho_temporario = tempfile.mkstemp(
        prefix="lista_apoio_",
        suffix=".xlsx",
        dir=diretorio,
    )
    os.close(descritor)
    try:
        wb.save(caminho_temporario)
        with zipfile.ZipFile(caminho_temporario) as arquivo_zip:
            if arquivo_zip.testzip() is not None:
                raise zipfile.BadZipFile("Arquivo temporário gerado com corrupção.")
        os.replace(caminho_temporario, ARQUIVO_LISTA_APOIO)
        shutil.copyfile(ARQUIVO_LISTA_APOIO, ARQUIVO_LISTA_APOIO_BACKUP)
        salvar_snapshot_lista_apoio(ARQUIVO_LISTA_APOIO)
    finally:
        if os.path.exists(caminho_temporario):
            os.remove(caminho_temporario)


def salvar_dataframe_lista_apoio(dataframe):
    diretorio = os.path.dirname(os.path.abspath(ARQUIVO_LISTA_APOIO)) or "."
    descritor, caminho_temporario = tempfile.mkstemp(
        prefix="lista_apoio_",
        suffix=".xlsx",
        dir=diretorio,
    )
    os.close(descritor)
    try:
        dataframe.to_excel(caminho_temporario, index=False)
        with zipfile.ZipFile(caminho_temporario) as arquivo_zip:
            if arquivo_zip.testzip() is not None:
                raise zipfile.BadZipFile("Arquivo temporário gerado com corrupção.")
        os.replace(caminho_temporario, ARQUIVO_LISTA_APOIO)
        shutil.copyfile(ARQUIVO_LISTA_APOIO, ARQUIVO_LISTA_APOIO_BACKUP)
        salvar_snapshot_lista_apoio(ARQUIVO_LISTA_APOIO)
    finally:
        if os.path.exists(caminho_temporario):
            os.remove(caminho_temporario)


def garantir_lista_apoio():
    garantir_banco_lista_apoio()
    migrar_lista_apoio_para_banco_se_necessario()
    deduplicar_lista_apoio_no_banco()
    return None, None


def registrar_duvida_apoio(tecnico, topico, resumo):
    garantir_lista_apoio()
    coluna_carimbo = CABECALHOS_LISTA_APOIO[0]
    coluna_tecnico = CABECALHOS_LISTA_APOIO[1]
    coluna_topico = CABECALHOS_LISTA_APOIO[2]
    coluna_descricao = CABECALHOS_LISTA_APOIO[3]
    coluna_situacao = CABECALHOS_LISTA_APOIO[4]
    coluna_responsavel = CABECALHOS_LISTA_APOIO[5]
    registro = {
        coluna_carimbo: datetime.now(FUSO_HORARIO_APP).strftime(
            "%d/%m/%Y %H:%M:%S"
        ),
        coluna_tecnico: builtins.str(tecnico or "").title(),
        coluna_topico: builtins.str(topico or "").strip(),
        coluna_descricao: builtins.str(resumo or "").strip(),
        coluna_situacao: "Aberto",
        coluna_responsavel: "",
    }
    salvar_registros_lista_apoio_no_banco([registro])
    registrar_historico_lista_apoio("create", registro)


def ler_lista_apoio():
    garantir_lista_apoio()
    dataframe_banco = ler_lista_apoio_do_banco()
    if not dataframe_banco.empty:
        return dataframe_banco
    dataframe_historico = restaurar_lista_apoio_do_historico()
    if not dataframe_historico.empty:
        salvar_registros_lista_apoio_no_banco(
            dataframe_para_registros_lista_apoio(dataframe_historico)
        )
        return dataframe_historico
    return dataframe_lista_apoio_vazio()


def texto_lista_apoio(valor):
    if pd.isna(valor):
        return ""
    return builtins.str(valor or "").strip()


def normalizar_valor_lista_apoio(coluna, valor, valor_atual=""):
    texto = texto_lista_apoio(valor)
    texto_atual = texto_lista_apoio(valor_atual)

    if coluna == "Situação":
        if texto == "Respondido":
            texto = "Resolvido pelo técnico"
        if texto_atual == "Respondido":
            texto_atual = "Resolvido pelo técnico"

        opcoes = ["Aberto", "Em análise", "Resolvido pelo técnico", "Finalizado"]
        if texto in opcoes:
            return texto
        if texto_atual in opcoes:
            return texto_atual
        return "Aberto"

    if coluna == "Responsável/Apoio":
        opcoes = ["Brenda", "Luma"]
        if texto in opcoes:
            return texto
        return texto_atual

    return texto or texto_atual


def salvar_lista_apoio(dataframe):
    existente = ler_lista_apoio()
    existente_antes = existente.copy()
    dataframe = dataframe.copy()
    for coluna_lista in CABECALHOS_LISTA_APOIO:
        if coluna_lista not in dataframe.columns:
            dataframe[coluna_lista] = ""
        dataframe[coluna_lista] = dataframe[coluna_lista].astype("object")
        existente[coluna_lista] = existente[coluna_lista].astype("object")
    dataframe = dataframe[CABECALHOS_LISTA_APOIO]

    colunas_editaveis = [CABECALHOS_LISTA_APOIO[4], CABECALHOS_LISTA_APOIO[5]]
    colunas_chave = [
        CABECALHOS_LISTA_APOIO[0],
        CABECALHOS_LISTA_APOIO[1],
        CABECALHOS_LISTA_APOIO[2],
        CABECALHOS_LISTA_APOIO[3],
    ]

    for coluna_lista in colunas_editaveis:
        if coluna_lista not in existente.columns:
            existente[coluna_lista] = ""

    for indice, linha in dataframe.iterrows():
        if indice not in existente.index:
            continue
        for coluna in colunas_editaveis:
            existente.at[indice, coluna] = normalizar_valor_lista_apoio(
                coluna,
                linha[coluna],
                existente.at[indice, coluna],
            )

    edits_por_chave = {}
    for _, linha in dataframe.iterrows():
        chave = tuple(texto_lista_apoio(linha[coluna]) for coluna in colunas_chave)
        edits_por_chave[chave] = {
            coluna: linha[coluna]
            for coluna in colunas_editaveis
        }

    for indice, linha in existente.iterrows():
        chave = tuple(texto_lista_apoio(linha[coluna]) for coluna in colunas_chave)
        if chave not in edits_por_chave:
            continue
        for coluna in colunas_editaveis:
            existente.at[indice, coluna] = normalizar_valor_lista_apoio(
                coluna,
                edits_por_chave[chave][coluna],
                existente.at[indice, coluna],
            )

    for indice, linha in existente.iterrows():
        if indice >= len(existente_antes):
            continue
        registro_antes = {
            coluna: texto_lista_apoio(existente_antes.iloc[indice][coluna])
            for coluna in CABECALHOS_LISTA_APOIO
        }
        registro_depois = {
            coluna: texto_lista_apoio(linha[coluna])
            for coluna in CABECALHOS_LISTA_APOIO
        }
        if registro_antes != registro_depois:
            registrar_historico_lista_apoio("update", registro_depois)

    dataframe_final = existente[CABECALHOS_LISTA_APOIO]
    salvar_registros_lista_apoio_no_banco(
        dataframe_para_registros_lista_apoio(dataframe_final)
    )


def bytes_lista_apoio():
    return _bytes_lista_apoio_cache(versao_lista_apoio())


@st.cache_data(show_spinner=False)
def _bytes_lista_apoio_cache(_versao_lista_apoio):
    garantir_lista_apoio()
    dataframe = ler_lista_apoio_do_banco()
    return bytes_dataframe_excel(dataframe)


def versao_lista_apoio():
    if backend_lista_apoio() == "postgres":
        try:
            with conexao_lista_apoio_db() as conexao:
                linha = conexao.execute(
                    "SELECT COALESCE(MAX(atualizado_em), 'sem_arquivo') AS versao, COUNT(*) AS total FROM lista_apoio"
                ).fetchone()
                if linha:
                    return f"{linha['versao']}|{linha['total']}"
        except Exception:
            return "sem_arquivo"
    if os.path.exists(ARQUIVO_LISTA_APOIO_DB):
        return builtins.str(int(os.path.getmtime(ARQUIVO_LISTA_APOIO_DB)))
    return "sem_arquivo"


def mostrar_lista_apoio_gestao():
    st.divider()
    st.subheader("Lista de Apoio")

    chave_filtro_data = "filtro_data_lista_apoio_widget"
    if chave_filtro_data not in st.session_state:
        st.session_state[chave_filtro_data] = st.session_state.get(
            "filtro_data_lista_apoio_persistido"
        )

    col_atualizar_apoio, col_status_apoio = st.columns([1, 3])
    with col_atualizar_apoio:
        if st.button("Atualizar lista de ajuda", key="atualizar_lista_apoio"):
            st.session_state["versao_editor_lista_apoio"] = versao_lista_apoio()
    with col_status_apoio:
        st.caption("Use o botão para atualizar a lista sem derrubar o login.")

    try:
        lista_apoio = ler_lista_apoio()
    except Exception as erro_lista_apoio:
        st.error(f"Não foi possível carregar a lista de apoio: {erro_lista_apoio}")
        lista_apoio = pd.DataFrame(columns=CABECALHOS_LISTA_APOIO)

    if lista_apoio.empty:
        st.info("Ainda não há dúvidas registradas pelos técnicos.")
        return

    col_filtro_data, col_limpar_filtro = st.columns([3, 1])
    with col_filtro_data:
        filtro_data_apoio = st.date_input(
            "Filtrar ajudas por data",
            value=None,
            format="DD/MM/YYYY",
            key=chave_filtro_data,
        )
        st.session_state["filtro_data_lista_apoio_persistido"] = filtro_data_apoio
    with col_limpar_filtro:
        st.write("")
        st.write("")
        if st.button("Limpar filtro", key="limpar_filtro_lista_apoio"):
            st.session_state["filtro_data_lista_apoio_persistido"] = None
            st.session_state["filtro_data_lista_apoio_widget"] = None
            st.rerun()

    lista_exibida = lista_apoio.copy()
    if filtro_data_apoio:
        if isinstance(filtro_data_apoio, datetime):
            filtro_data_apoio = filtro_data_apoio.date()
        datas_apoio = pd.to_datetime(
            lista_exibida["Carimbo de data/hora"].astype(str).str.strip(),
            format="%d/%m/%Y %H:%M:%S",
            dayfirst=True,
            errors="coerce",
        ).dt.date
        lista_exibida = lista_exibida[datas_apoio == filtro_data_apoio]

    st.caption(f"{len(lista_exibida)} registro(s) exibido(s).")

    if lista_exibida.empty:
        st.info("Nenhuma ajuda registrada para a data informada.")
        st.download_button(
            "Baixar lista em Excel",
            data=bytes_lista_apoio(),
            file_name=ARQUIVO_LISTA_APOIO,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="baixar_lista_apoio_vazia",
        )
        return

    if "versao_editor_lista_apoio" not in st.session_state:
        st.session_state["versao_editor_lista_apoio"] = versao_lista_apoio()

    chave_editor_apoio = (
        "editor_lista_apoio_"
        + re.sub(
            r"[^0-9a-zA-Z]+",
            "_",
            filtro_data_apoio.strftime("%d_%m_%Y") if filtro_data_apoio else "todas",
        )
        + "_"
        + st.session_state["versao_editor_lista_apoio"]
    )

    lista_exibida_editor = lista_exibida.copy()
    for coluna_editor in ["Situação", "Responsável/Apoio"]:
        if coluna_editor in lista_exibida_editor.columns:
            lista_exibida_editor[coluna_editor] = (
                lista_exibida_editor[coluna_editor].fillna("").astype("object")
            )

    with st.form(f"form_{chave_editor_apoio}", clear_on_submit=False):
        lista_editada = st.data_editor(
            lista_exibida_editor,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            disabled=[
                "Carimbo de data/hora",
                "Nome do Técnico",
                "Selecione o tópico",
                "Descreva o problema/pedido em poucas palavras",
            ],
            column_config={
                "Situação": st.column_config.SelectboxColumn(
                    "Situação",
                    options=["Aberto", "Em análise", "Resolvido pelo técnico", "Finalizado"],
                ),
                "Responsável/Apoio": st.column_config.SelectboxColumn(
                    "Responsável/Apoio",
                    options=["Brenda", "Luma"],
                )
            },
            key=chave_editor_apoio,
        )
        st.caption(
            "Depois de ajustar a situação ou o responsável, use o botão abaixo para gravar."
        )
        salvar_lista_apoio_submit = st.form_submit_button("Salvar lista de apoio")

    if salvar_lista_apoio_submit:
        try:
            salvar_lista_apoio(lista_editada)
        except PermissionError:
            st.error("Feche a lista_apoio.xlsx e tente salvar novamente.")
        except Exception as erro_lista_apoio:
            st.error(f"Não foi possível salvar a lista: {erro_lista_apoio}")
        else:
            st.success("Lista de apoio atualizada.")
            st.session_state["versao_editor_lista_apoio"] = versao_lista_apoio()
            st.rerun()

    st.download_button(
        "Baixar lista em Excel",
        data=bytes_lista_apoio(),
        file_name=ARQUIVO_LISTA_APOIO,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="baixar_lista_apoio",
    )


CORES_DISPONIBILIDADE = {
    "Férias": "#2563EB",
    "Licença-maternidade": "#7C3AED",
    "Atestado": "#DC2626",
    "Falta/Atraso": "#DC2626",
    "Energia": "#F97316",
}


def nivel_tecnico_no_mes(dataframe, tecnico, ano, mes):
    limite_mes = pd.Timestamp(
        date(ano, mes, calendar.monthrange(ano, mes)[1])
    )
    dados_tecnico_nivel = dataframe[
        (dataframe["Técnico"].map(normalizar_nome) == normalizar_nome(tecnico))
        & (dataframe["Data"] <= limite_mes)
    ].dropna(subset=["Nível", "Data"])
    if dados_tecnico_nivel.empty:
        return "Não informado"
    return builtins.str(
        dados_tecnico_nivel.sort_values("Data").iloc[-1]["Nível"]
    ).strip()


def dataframe_disponibilidade(dataframe):
    registros = []

    for tecnico, inicio, fim in PROGRAMACAO_FERIAS:
        registros.append(
            {
                "Técnico": tecnico,
                "Tipo": "Férias",
                "Início": inicio,
                "Fim": fim,
                "Duração": f"{(fim - inicio).days + 1} dias",
                "Observação": "Férias programadas",
            }
        )

    for tecnico, inicio, fim, motivo in PROGRAMACAO_LICENCAS:
        registros.append(
            {
                "Técnico": tecnico,
                "Tipo": motivo,
                "Início": inicio,
                "Fim": fim,
                "Duração": f"{(fim - inicio).days + 1} dias",
                "Observação": "Afastamento programado",
            }
        )

    for tecnico, data_ausencia, tipo, minutos in PROGRAMACAO_AUSENCIAS:
        duracao = (
            "Dia integral"
            if minutos is None
            else f"{minutos // 60:02d}:{minutos % 60:02d}"
        )
        registros.append(
            {
                "Técnico": tecnico,
                "Tipo": tipo,
                "Início": data_ausencia,
                "Fim": data_ausencia,
                "Duração": duracao,
                "Observação": (
                    "Ausência integral"
                    if minutos is None
                    else "Ausência parcial"
                ),
            }
        )

    disponibilidade = pd.DataFrame(registros)
    if disponibilidade.empty:
        return disponibilidade

    disponibilidade["Nível"] = disponibilidade.apply(
        lambda linha: nivel_tecnico_no_mes(
            dataframe,
            linha["Técnico"],
            linha["Início"].year,
            linha["Início"].month,
        ),
        axis=1,
    )
    return disponibilidade


def nome_mes_ano(ano, mes):
    nomes_meses = [
        "",
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    return f"{nomes_meses[mes]} de {ano}"


def abreviar_nome_disponibilidade(nome):
    partes = builtins.str(nome or "").title().split()
    if len(partes) <= 2:
        return " ".join(partes)
    return f"{partes[0]} {partes[-1]}"


def html_calendario_disponibilidade(ano, mes, eventos):
    calendario = calendar.Calendar(firstweekday=0)
    semanas = calendario.monthdatescalendar(ano, mes)
    hoje = datetime.now(FUSO_HORARIO_APP).date()
    cabecalhos = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    partes_html = [
        '<div class="disp-calendario">',
        '<div class="disp-semana-cabecalho">',
    ]
    partes_html.extend(
        f'<div class="disp-dia-semana">{dia}</div>' for dia in cabecalhos
    )
    partes_html.append("</div>")

    for semana in semanas:
        partes_html.append('<div class="disp-semana">')
        for data_calendario in semana:
            classes = ["disp-dia"]
            if data_calendario.month != mes:
                classes.append("fora-mes")
            if data_calendario.weekday() >= 5:
                classes.append("fim-semana")
            if data_calendario == hoje:
                classes.append("hoje")

            partes_html.append(
                f'<div class="{" ".join(classes)}">'
                f'<div class="disp-numero-dia">{data_calendario.day}</div>'
            )
            eventos_dia = eventos[
                (eventos["Início"] <= data_calendario)
                & (eventos["Fim"] >= data_calendario)
            ]
            for _, evento in eventos_dia.head(4).iterrows():
                cor = CORES_DISPONIBILIDADE.get(evento["Tipo"], "#6B7280")
                rotulo = escape(abreviar_nome_disponibilidade(evento["Técnico"]))
                detalhe = escape(builtins.str(evento["Duração"]))
                partes_html.append(
                    '<div class="disp-evento" '
                    f'style="border-left-color:{cor};">'
                    f'<span>{rotulo}</span><small>{detalhe}</small>'
                    "</div>"
                )
            if len(eventos_dia) > 4:
                partes_html.append(
                    f'<div class="disp-mais">+{len(eventos_dia) - 4} ocorrência(s)</div>'
                )
            partes_html.append("</div>")
        partes_html.append("</div>")

    partes_html.append("</div>")
    return "".join(partes_html)


def mostrar_gestao_disponibilidade(dataframe):
    st.divider()
    st.subheader("Gestão de Disponibilidade")

    disponibilidade = dataframe_disponibilidade(dataframe)
    if disponibilidade.empty:
        st.info("Ainda não há férias, licenças ou ausências cadastradas.")
        return

    menor_data = disponibilidade["Início"].min()
    maior_data = disponibilidade["Fim"].max()
    periodos = []
    cursor_periodo = date(menor_data.year, menor_data.month, 1)
    while cursor_periodo <= date(maior_data.year, maior_data.month, 1):
        periodos.append((cursor_periodo.year, cursor_periodo.month))
        cursor_periodo = (
            date(cursor_periodo.year + 1, 1, 1)
            if cursor_periodo.month == 12
            else date(cursor_periodo.year, cursor_periodo.month + 1, 1)
        )
    periodo_atual = (datetime.now(FUSO_HORARIO_APP).year, datetime.now(FUSO_HORARIO_APP).month)
    periodo_padrao = periodo_atual if periodo_atual in periodos else periodos[0]
    opcoes_periodo = {
        nome_mes_ano(ano, mes): (ano, mes)
        for ano, mes in periodos
    }
    periodo_padrao_label = next(
        nome
        for nome, periodo in opcoes_periodo.items()
        if periodo == periodo_padrao
    )
    if st.session_state.get("disponibilidade_mes") not in opcoes_periodo:
        st.session_state["disponibilidade_mes"] = periodo_padrao_label

    col_mes, col_tecnico, col_nivel, col_tipo = st.columns([1.15, 1.35, 1, 1])
    with col_mes:
        periodo_label = st.selectbox(
            "Mês",
            list(opcoes_periodo),
            index=list(opcoes_periodo.values()).index(periodo_padrao),
            key="disponibilidade_mes",
        )
    ano_filtro, mes_filtro = opcoes_periodo[periodo_label]
    inicio_mes = date(ano_filtro, mes_filtro, 1)
    fim_mes = date(
        ano_filtro,
        mes_filtro,
        calendar.monthrange(ano_filtro, mes_filtro)[1],
    )

    eventos_mes = disponibilidade[
        (disponibilidade["Início"] <= fim_mes)
        & (disponibilidade["Fim"] >= inicio_mes)
    ].copy()
    tecnicos_filtro = sorted(eventos_mes["Técnico"].unique())
    niveis_filtro = sorted(eventos_mes["Nível"].unique())
    tipos_filtro = sorted(eventos_mes["Tipo"].unique())
    for chave_estado, opcoes_validas in {
        "disponibilidade_tecnico": ["Todos"] + tecnicos_filtro,
        "disponibilidade_nivel": ["Todos"] + niveis_filtro,
        "disponibilidade_tipo": ["Todos"] + tipos_filtro,
    }.items():
        if st.session_state.get(chave_estado) not in opcoes_validas:
            st.session_state[chave_estado] = "Todos"

    with col_tecnico:
        tecnico_filtro = st.selectbox(
            "Técnico",
            ["Todos"] + tecnicos_filtro,
            key="disponibilidade_tecnico",
        )
    with col_nivel:
        nivel_filtro = st.selectbox(
            "Nível",
            ["Todos"] + niveis_filtro,
            key="disponibilidade_nivel",
        )
    with col_tipo:
        tipo_filtro = st.selectbox(
            "Tipo",
            ["Todos"] + tipos_filtro,
            key="disponibilidade_tipo",
        )

    eventos_filtrados = eventos_mes.copy()
    if tecnico_filtro != "Todos":
        eventos_filtrados = eventos_filtrados[
            eventos_filtrados["Técnico"] == tecnico_filtro
        ]
    if nivel_filtro != "Todos":
        eventos_filtrados = eventos_filtrados[
            eventos_filtrados["Nível"] == nivel_filtro
        ]
    if tipo_filtro != "Todos":
        eventos_filtrados = eventos_filtrados[
            eventos_filtrados["Tipo"] == tipo_filtro
        ]

    efetivo_mes = dataframe[
        (dataframe["Data"].dt.year == ano_filtro)
        & (dataframe["Data"].dt.month == mes_filtro)
    ]["Técnico"].nunique()
    total_ferias = eventos_mes[eventos_mes["Tipo"] == "Férias"]["Técnico"].nunique()
    total_licencas = eventos_mes[
        eventos_mes["Tipo"].str.contains("Licença", case=False, na=False)
    ]["Técnico"].nunique()
    total_ausencias = len(
        eventos_mes[
            ~eventos_mes["Tipo"].isin(["Férias", "Licença-maternidade"])
        ]
    )

    st.markdown(
        f"""
        <div class="disp-resumo">
            <div><span>Efetivo no mês</span><strong>{efetivo_mes}</strong></div>
            <div><span>Em férias</span><strong>{total_ferias}</strong></div>
            <div><span>Licenças</span><strong>{total_licencas}</strong></div>
            <div><span>Ausências</span><strong>{total_ausencias}</strong></div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    legenda_disponibilidade = [
        ("Férias", CORES_DISPONIBILIDADE["Férias"]),
        ("Licença", CORES_DISPONIBILIDADE["Licença-maternidade"]),
        ("Atestado/Falta", CORES_DISPONIBILIDADE["Atestado"]),
        ("Ausência parcial", CORES_DISPONIBILIDADE["Energia"]),
    ]
    st.markdown(
        '<div class="disp-legenda">'
        + "".join(
            f'<span><i style="background:{cor};"></i>{rotulo}</span>'
            for rotulo, cor in legenda_disponibilidade
        )
        + "</div>",
        unsafe_allow_html=True,
    )

    aba_calendario, aba_lista, aba_nivel = st.tabs(
        ["Calendário", "Lista", "Resumo por nível"]
    )
    with aba_calendario:
        if eventos_filtrados.empty:
            st.info("Nenhuma ocorrência encontrada para os filtros selecionados.")
        else:
            col_calendario, col_proximos = st.columns([4, 1.1])
            with col_calendario:
                st.markdown(
                    html_calendario_disponibilidade(
                        ano_filtro,
                        mes_filtro,
                        eventos_filtrados,
                    ),
                    unsafe_allow_html=True,
                )
            with col_proximos:
                st.markdown("**Próximos afastamentos**")
                referencia_proximos = max(
                    datetime.now(FUSO_HORARIO_APP).date(),
                    inicio_mes,
                )
                proximos = eventos_filtrados[
                    eventos_filtrados["Fim"] >= referencia_proximos
                ].sort_values(["Início", "Técnico"]).head(6)
                if proximos.empty:
                    st.caption("Nenhum afastamento futuro neste período.")
                else:
                    itens_proximos = []
                    for _, evento in proximos.iterrows():
                        cor = CORES_DISPONIBILIDADE.get(
                            evento["Tipo"],
                            "#6B7280",
                        )
                        itens_proximos.append(
                            '<div class="disp-proximo" '
                            f'style="border-left-color:{cor};">'
                            f'<strong>{escape(abreviar_nome_disponibilidade(evento["Técnico"]))}</strong>'
                            f'<span>{escape(evento["Tipo"])}</span>'
                            f'<small>{evento["Início"].strftime("%d/%m")} a '
                            f'{evento["Fim"].strftime("%d/%m")}</small>'
                            "</div>"
                        )
                    st.markdown("".join(itens_proximos), unsafe_allow_html=True)

    with aba_lista:
        lista_disponibilidade = eventos_filtrados.copy()
        lista_disponibilidade["Início"] = lista_disponibilidade["Início"].map(
            lambda valor: valor.strftime("%d/%m/%Y")
        )
        lista_disponibilidade["Fim"] = lista_disponibilidade["Fim"].map(
            lambda valor: valor.strftime("%d/%m/%Y")
        )
        st.dataframe(
            lista_disponibilidade[
                [
                    "Técnico",
                    "Nível",
                    "Tipo",
                    "Início",
                    "Fim",
                    "Duração",
                    "Observação",
                ]
            ],
            hide_index=True,
            use_container_width=True,
        )

    with aba_nivel:
        if eventos_filtrados.empty:
            st.info("Nenhuma ocorrência encontrada para os filtros selecionados.")
        else:
            resumo_nivel = (
                eventos_filtrados.groupby(["Nível", "Tipo"])
                .size()
                .reset_index(name="Ocorrências")
            )
            grafico_nivel = px.bar(
                resumo_nivel,
                x="Nível",
                y="Ocorrências",
                color="Tipo",
                barmode="group",
                color_discrete_map=CORES_DISPONIBILIDADE,
                text="Ocorrências",
            )
            grafico_nivel.update_layout(
                height=360,
                margin=dict(l=20, r=20, t=20, b=30),
                plot_bgcolor="#FFFFFF",
                paper_bgcolor="#FFFFFF",
                legend_title_text="",
                xaxis_title="",
                yaxis_title="Ocorrências",
            )
            grafico_nivel.update_traces(textposition="outside")
            st.plotly_chart(grafico_nivel, use_container_width=True)


def salvar_env_local(atualizacoes):
    valores = carregar_env_local()
    for chave, valor in atualizacoes.items():
        valores[chave] = builtins.str(valor or "").strip()

    linhas = [f"{chave}={valor}" for chave, valor in sorted(valores.items())]
    with open(".env", "w", encoding="utf-8") as arquivo:
        arquivo.write("\n".join(linhas) + "\n")


def normalizar_nome(valor):
    texto = builtins.str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"\([^)]*\)", " ", texto)
    texto = re.sub(r"\b(de|da|do|das|dos|e)\b", " ", texto)
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def periodos_ferias_tecnico(tecnico):
    tecnico_normalizado = normalizar_nome(tecnico)
    return [
        (inicio, fim)
        for nome, inicio, fim in PROGRAMACAO_FERIAS
        if normalizar_nome(nome) == tecnico_normalizado
    ]


def dias_uteis_ferias_no_mes(tecnico, ano, mes):
    inicio_mes_referencia = date(ano, mes, 1)
    fim_mes_referencia = date(ano, mes, calendar.monthrange(ano, mes)[1])
    dias_ferias = set()

    for inicio_ferias, fim_ferias in periodos_ferias_tecnico(tecnico):
        inicio = max(inicio_ferias, inicio_mes_referencia)
        fim = min(fim_ferias, fim_mes_referencia)
        if inicio > fim:
            continue

        for data_atual in datas_no_periodo(inicio, fim):
            if data_atual.weekday() < 5 and not eh_feriado_federal(data_atual):
                dias_ferias.add(data_atual)

    return len(dias_ferias)


def ferias_ativa_tecnico(tecnico, data_referencia=None):
    data_referencia = data_referencia or datetime.now(FUSO_HORARIO_APP).date()
    for inicio, fim in periodos_ferias_tecnico(tecnico):
        if inicio <= data_referencia <= fim:
            return inicio, fim
    return None


def periodos_licenca_tecnico(tecnico):
    tecnico_normalizado = normalizar_nome(tecnico)
    return [
        (inicio, fim, motivo)
        for nome, inicio, fim, motivo in PROGRAMACAO_LICENCAS
        if normalizar_nome(nome) == tecnico_normalizado
    ]


def licenca_ativa_tecnico(tecnico, data_referencia=None):
    data_referencia = data_referencia or datetime.now(FUSO_HORARIO_APP).date()
    for inicio, fim, motivo in periodos_licenca_tecnico(tecnico):
        if inicio <= data_referencia <= fim:
            return inicio, fim, motivo
    return None


def dias_uteis_licenca_no_mes(tecnico, ano, mes):
    inicio_mes_referencia = date(ano, mes, 1)
    fim_mes_referencia = date(ano, mes, calendar.monthrange(ano, mes)[1])
    dias_licenca = set()

    for inicio_licenca, fim_licenca, _ in periodos_licenca_tecnico(tecnico):
        inicio = max(inicio_licenca, inicio_mes_referencia)
        fim = min(fim_licenca, fim_mes_referencia)
        if inicio > fim:
            continue

        for data_atual in datas_no_periodo(inicio, fim):
            if data_atual.weekday() < 5 and not eh_feriado_federal(data_atual):
                dias_licenca.add(data_atual)

    return len(dias_licenca)


def ausencia_programada_tecnico(tecnico, data_referencia):
    tecnico_normalizado = normalizar_nome(tecnico)
    for nome, data_ausencia, tipo, minutos in PROGRAMACAO_AUSENCIAS:
        if (
            normalizar_nome(nome) == tecnico_normalizado
            and data_ausencia == data_referencia
        ):
            return tipo, minutos
    return None


def abatimento_ausencias_no_mes(tecnico, ano, mes):
    tecnico_normalizado = normalizar_nome(tecnico)
    abatimento = 0.0

    for nome, data_ausencia, _, minutos in PROGRAMACAO_AUSENCIAS:
        if normalizar_nome(nome) != tecnico_normalizado:
            continue
        if data_ausencia.year != ano or data_ausencia.month != mes:
            continue
        if data_ausencia.weekday() >= 5 or eh_feriado_federal(data_ausencia):
            continue

        if minutos is None:
            abatimento += 1
        else:
            abatimento += min(
                max(minutos / CARGA_DIARIA_PADRAO_MINUTOS, 0),
                1,
            )

    return round(abatimento, 4)


def nivel_canonico(valor):
    nivel = normalizar_nome(valor)
    niveis = {
        "tecnico iii": "Técnico III",
        "tecnico ii": "Técnico II",
        "tecnico i": "Técnico I",
        "jr": "JR",
        "estagio": "Estágio",
    }
    return niveis.get(nivel, builtins.str(valor or "").strip())


def meta_esperada_nivel(valor):
    return META_ESPERADA_POR_NIVEL.get(nivel_canonico(valor), 0)


def arredondar_esperado(valor):
    if pd.isna(valor):
        return 0
    return int(
        Decimal(builtins.str(float(valor))).quantize(
            Decimal("1"),
            rounding=ROUND_HALF_UP,
        )
    )


def percentuais_absorcao_por_mes(dataframe, ano, mes):
    dados_mes = dataframe[
        (dataframe["Data"].dt.year == ano)
        & (dataframe["Data"].dt.month == mes)
    ].copy()
    if dados_mes.empty:
        return {nivel: 0 for nivel in META_ESPERADA_POR_NIVEL}

    dados_mes = dados_mes[
        ~dados_mes["Técnico"]
        .map(normalizar_nome)
        .isin(TECNICOS_DESCONSIDERADOS_ESPERADO)
    ].copy()
    dados_mes["Nivel Canonico"] = dados_mes["Nível"].map(nivel_canonico)
    dados_mes = dados_mes.sort_values("Data")
    ultimo_nivel_tecnico = dados_mes.drop_duplicates(
        subset=["Técnico"],
        keep="last",
    )
    total_meta_mes = ultimo_nivel_tecnico["Nivel Canonico"].map(
        META_ESPERADA_POR_NIVEL
    ).fillna(0).sum()

    if not total_meta_mes:
        return {nivel: 0 for nivel in META_ESPERADA_POR_NIVEL}

    return {
        nivel: (meta / total_meta_mes) * 100
        for nivel, meta in META_ESPERADA_POR_NIVEL.items()
    }


def percentual_absorcao_tecnico(dataframe, ano, mes, nivel):
    return percentuais_absorcao_por_mes(dataframe, ano, mes).get(
        nivel_canonico(nivel),
        0,
    )


def total_meta_mensal_por_linha(dataframe):
    totais_por_mes = {}
    for periodo, dados_mes in dataframe.groupby(dataframe["Data"].dt.to_period("M")):
        dados_validos = dados_mes[
            ~dados_mes["Técnico"]
            .map(normalizar_nome)
            .isin(TECNICOS_DESCONSIDERADOS_ESPERADO)
        ].copy()
        if dados_validos.empty:
            totais_por_mes[periodo] = 0
            continue

        dados_validos = dados_validos.sort_values("Data")
        ultimo_nivel_tecnico = dados_validos.drop_duplicates(
            subset=["Técnico"],
            keep="last",
        )
        totais_por_mes[periodo] = ultimo_nivel_tecnico["Nível"].map(
            meta_esperada_nivel
        ).fillna(0).sum()

    return dataframe["Data"].dt.to_period("M").map(totais_por_mes).fillna(0)


def cabecalhos(ws):
    return {
        builtins.str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def garantir_coluna(ws, nome_coluna):
    colunas = cabecalhos(ws)
    if nome_coluna in colunas:
        return colunas[nome_coluna]

    indice = ws.max_column + 1
    ws.cell(row=1, column=indice).value = nome_coluna
    return indice


def normalizar_cabecalho(valor):
    texto = builtins.str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"\s+", " ", texto)
    return texto


def coluna(colunas, *nomes):
    colunas_normalizadas = {
        normalizar_cabecalho(nome): indice
        for nome, indice in colunas.items()
    }

    for nome in nomes:
        indice = colunas_normalizadas.get(normalizar_cabecalho(nome))
        if indice:
            return indice

    raise KeyError(nomes[0])


def nome_coluna_dataframe(dataframe, *nomes):
    colunas_normalizadas = {
        normalizar_cabecalho(nome): nome for nome in dataframe.columns
    }

    for nome in nomes:
        coluna_encontrada = colunas_normalizadas.get(normalizar_cabecalho(nome))
        if coluna_encontrada:
            return coluna_encontrada

    raise KeyError(nomes[0])


def status_por_desvio(desvio):
    if desvio < -5:
        return "CRÍTICO"
    if desvio < 0:
        return "ATENÇÃO"
    if desvio <= 5:
        return "BOM"
    return "EXCELENTE"


def melhor_alias(nome_tecnico, candidatos):
    nome_normalizado = normalizar_nome(nome_tecnico)
    melhor = ""
    melhor_score = 0

    for candidato in candidatos:
        score = SequenceMatcher(
            None,
            nome_normalizado,
            normalizar_nome(candidato),
        ).ratio()
        if score > melhor_score:
            melhor = candidato
            melhor_score = score

    return melhor if melhor_score >= 0.72 else ""


def garantir_estrutura_aliases(ws_aliases):
    headers = [
        "T??cnico Planilha",
        "Agente BI",
        "Agente SGD",
        "Agente Ponto",
        "Agente Chat",
        "Agente RO",
    ]
    for indice, header in enumerate(headers, start=1):
        if ws_aliases.cell(row=1, column=indice).value != header:
            ws_aliases.cell(row=1, column=indice).value = header


def tecnicos_da_planilha(ws, colunas):
    return sorted(
        {
            ws.cell(row=row, column=colunas["Técnico"]).value
            for row in range(2, ws.max_row + 1)
            if ws.cell(row=row, column=colunas["Técnico"]).value
        }
    )


def garantir_linhas_da_data(ws, colunas, data_referencia):
    col_data = coluna(colunas, "Data")
    col_tecnico = coluna(colunas, "Técnico")
    col_nivel = coluna(colunas, "Nível")

    linhas_existentes = []
    datas_existentes = set()
    linhas_por_data = {}

    for row in range(2, ws.max_row + 1):
        data_linha = ws.cell(row=row, column=col_data).value
        if not data_linha:
            continue
        data_linha = data_linha.date() if hasattr(data_linha, "date") else data_linha
        datas_existentes.add(data_linha)
        linhas_por_data.setdefault(data_linha, []).append(row)
        if data_linha == data_referencia:
            linhas_existentes.append(row)

    if linhas_existentes:
        return 0

    datas_anteriores = [data for data in datas_existentes if data < data_referencia]
    if not datas_anteriores:
        return 0

    data_base = max(datas_anteriores)
    criadas = 0

    for row_base in linhas_por_data[data_base]:
        tecnico = ws.cell(row=row_base, column=col_tecnico).value
        nivel = ws.cell(row=row_base, column=col_nivel).value
        nova_linha = [None] * ws.max_column
        nova_linha[col_data - 1] = data_referencia
        nova_linha[col_tecnico - 1] = tecnico
        nova_linha[col_nivel - 1] = nivel
        ws.append(nova_linha)
        criadas += 1

    return criadas


def garantir_aba_aliases(wb, tecnicos, candidatos, coluna_alias):
    if ABA_ALIASES not in wb.sheetnames:
        ws_aliases = wb.create_sheet(ABA_ALIASES)
    else:
        ws_aliases = wb[ABA_ALIASES]

    garantir_estrutura_aliases(ws_aliases)

    existentes = {
        normalizar_nome(ws_aliases.cell(row=row, column=1).value): row
        for row in range(2, ws_aliases.max_row + 1)
        if ws_aliases.cell(row=row, column=1).value
    }

    for tecnico in tecnicos:
        tecnico_normalizado = normalizar_nome(tecnico)
        if tecnico_normalizado not in existentes:
            ws_aliases.append([tecnico, "", "", "", "", ""])
            row = ws_aliases.max_row
            existentes[tecnico_normalizado] = row
        else:
            row = existentes[tecnico_normalizado]

        if not ws_aliases.cell(row=row, column=coluna_alias).value:
            ws_aliases.cell(row=row, column=coluna_alias).value = melhor_alias(
                tecnico,
                candidatos,
            )

    return ws_aliases


def ler_aliases(ws_aliases, coluna_alias):
    aliases = {}
    for row in range(2, ws_aliases.max_row + 1):
        tecnico = ws_aliases.cell(row=row, column=1).value
        alias = ws_aliases.cell(row=row, column=coluna_alias).value
        if tecnico and alias:
            aliases[normalizar_nome(tecnico)] = normalizar_nome(alias)
    return aliases


def localizar_funcionario_por_nome(funcionarios, nome_referencia):
    nome_normalizado = normalizar_nome(nome_referencia)
    if not nome_normalizado:
        return None

    tokens_referencia = set(nome_normalizado.split())
    melhor_funcionario = None
    melhor_score = 0.0

    for funcionario in funcionarios:
        nome_funcionario = funcionario.get("Nome", "")
        nome_funcionario_normalizado = normalizar_nome(nome_funcionario)
        if not nome_funcionario_normalizado:
            continue
        if nome_funcionario_normalizado == nome_normalizado:
            return funcionario

        tokens_funcionario = set(nome_funcionario_normalizado.split())
        intersecao = len(tokens_referencia & tokens_funcionario)
        uniao = len(tokens_referencia | tokens_funcionario) or 1
        score_tokens = intersecao / uniao
        score_texto = SequenceMatcher(
            None,
            nome_normalizado,
            nome_funcionario_normalizado,
        ).ratio()
        score = max(score_texto, (score_tokens * 0.7) + (score_texto * 0.3))

        if score > melhor_score:
            melhor_score = score
            melhor_funcionario = funcionario

    return melhor_funcionario if melhor_score >= 0.6 else None


def data_dia_anterior():
    data_referencia = date.today() - timedelta(days=1)
    while data_referencia.weekday() >= 5:
        data_referencia -= timedelta(days=1)
    return data_referencia


def inicio_mes(data_referencia):
    return date(data_referencia.year, data_referencia.month, 1)


def datas_no_periodo(data_inicial, data_final):
    atual = data_inicial
    while atual <= data_final:
        yield atual
        atual += timedelta(days=1)


def dias_uteis_para_meta(ano, mes):
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    total = 0
    for dia in range(1, ultimo_dia + 1):
        data_atual = date(ano, mes, dia)
        if data_atual.weekday() >= 5:
            continue
        if eh_feriado_federal(data_atual):
            continue
        total += 1
    return total


def eh_feriado_federal(data_atual):
    return (
        (data_atual.month, data_atual.day) in FERIADOS_FEDERAIS_FIXOS
        or data_atual in FERIADOS_ADICIONAIS
    )


def hhmm_para_minutos(valor):
    texto = builtins.str(valor or "").strip()
    if not texto or ":" not in texto:
        return 0

    negativo = texto.startswith("-")
    texto = texto.lstrip("+-")
    partes = texto.split(":")
    if len(partes) != 2 or not all(parte.isdigit() for parte in partes):
        return 0

    minutos = int(partes[0]) * 60 + int(partes[1])
    return -minutos if negativo else minutos


def extrair_valor_input(html, nome_campo):
    padrao = re.compile(
        rf'<input[^>]+name="{re.escape(nome_campo)}"[^>]+value="([^"]*)"',
        re.IGNORECASE,
    )
    match = padrao.search(html)
    return unescape(match.group(1)) if match else ""


def login_pontoweb(email, senha):
    sessao = criar_sessao_http()
    resposta_login = sessao.get(PONTOWEB_AUTH_URL, timeout=30)
    resposta_login.raise_for_status()

    html_login = resposta_login.text
    token_verificacao = extrair_valor_input(html_login, "__RequestVerificationToken")
    cliente_id = extrair_valor_input(html_login, "ClienteId") or PONTOWEB_CLIENT_ID
    redirect_uri = extrair_valor_input(html_login, "RedirectUri") or PONTOWEB_REDIRECT_URI

    payload = {
        "Email": email,
        "Senha": senha,
        "ContinuarConectado": "true",
        "ClienteId": cliente_id,
        "RedirectUri": redirect_uri,
        "__RequestVerificationToken": token_verificacao,
        "action:Login": "Login",
    }

    resposta_autorizacao = sessao.post(
        PONTOWEB_AUTH_URL,
        data=payload,
        headers={"Referer": PONTOWEB_AUTH_URL},
        allow_redirects=False,
        timeout=30,
    )

    if resposta_autorizacao.status_code not in {302, 303}:
        mensagem = extrair_erro_login_pontoweb(resposta_autorizacao.text)
        if mensagem:
            raise ValueError(mensagem)
        raise ValueError("Login do PontoWeb inválido.")

    location = resposta_autorizacao.headers.get("Location", "")
    if not location:
        raise ValueError("O PontoWeb não retornou o código de autorização.")

    match_code = re.search(r"[?&]code=([^&]+)", location)
    if not match_code:
        raise ValueError("Não encontrei o código de autorização do PontoWeb.")

    resposta_token = sessao.post(
        PONTOWEB_TOKEN_URL,
        data={
            "client_id": cliente_id,
            "redirect_uri": redirect_uri,
            "grant_type": "authorization_code",
            "code": match_code.group(1),
        },
        headers={"Content-type": "application/x-www-form-urlencoded; charset=UTF-8"},
        timeout=30,
    )
    resposta_token.raise_for_status()

    dados_token = resposta_token.json()
    return {
        "sessao": sessao,
        "access_token": dados_token["access_token"],
        "refresh_token": dados_token.get("refresh_token", ""),
    }


def cabecalhos_pontoweb(access_token, identificador_banco=""):
    headers = {"Authorization": f"Bearer {access_token}"}
    if identificador_banco:
        headers["secullumbancoselecionado"] = identificador_banco
    return headers


def buscar_toolbar_pontoweb(sessao, access_token):
    resposta = sessao.get(
        f"{PONTOWEB_BASE_URL}Toolbar",
        headers=cabecalhos_pontoweb(access_token),
        timeout=30,
    )
    resposta.raise_for_status()
    return resposta.json()


def localizar_lista_bancos(objeto):
    if isinstance(objeto, list):
        if objeto and all(
            isinstance(item, dict)
            and ("BancoId" in item or "bancoId" in item)
            and ("Identificador" in item or "identificador" in item)
            for item in objeto
        ):
            return objeto
        for item in objeto:
            lista = localizar_lista_bancos(item)
            if lista:
                return lista
        return []

    if isinstance(objeto, dict):
        for chave in ("listaBancos", "ListaBancos", "bancos", "Bancos"):
            valor = objeto.get(chave)
            if isinstance(valor, list) and valor:
                return valor
        for valor in objeto.values():
            lista = localizar_lista_bancos(valor)
            if lista:
                return lista

    return []


def selecionar_banco_pontoweb(lista_bancos, banco_id="", identificador=""):
    if not lista_bancos:
        return {}

    banco_id = builtins.str(banco_id or "").strip()
    identificador = builtins.str(identificador or "").strip()

    for banco in lista_bancos:
        if banco_id and builtins.str(
            banco.get("BancoId", banco.get("bancoId", ""))
        ) == banco_id:
            return banco
        if identificador and builtins.str(
            banco.get("Identificador", banco.get("identificador", ""))
        ) == identificador:
            return banco

    return lista_bancos[0]


def buscar_funcionarios_pontoweb(sessao, access_token, identificador_banco):
    resposta = sessao.get(
        f"{PONTOWEB_BASE_URL}Funcionarios",
        headers=cabecalhos_pontoweb(access_token, identificador_banco),
        timeout=60,
    )
    resposta.raise_for_status()
    return resposta.json()


def buscar_calculos_pontoweb(
    sessao,
    access_token,
    identificador_banco,
    funcionario_id,
    data_inicial,
    data_final,
):
    endpoint = (
        f"{PONTOWEB_BASE_URL}Calculos/"
        f"{funcionario_id}/{data_inicial.strftime('%Y-%m-%d')}/"
        f"{data_final.strftime('%Y-%m-%d')}"
    )
    resposta = sessao.get(
        endpoint,
        headers=cabecalhos_pontoweb(access_token, identificador_banco),
        timeout=60,
    )
    resposta.raise_for_status()
    return resposta.json()


def indice_coluna_calculo(colunas, *nomes):
    nomes_normalizados = {normalizar_nome(nome) for nome in nomes}
    for indice, coluna_calculo in enumerate(colunas):
        nome_coluna = coluna_calculo.get("NomeExibicao") or coluna_calculo.get("Nome")
        if normalizar_nome(nome_coluna) in nomes_normalizados:
            return indice
    return -1


def valor_linha_calculo(linha, indice):
    if indice < 0 or indice >= len(linha):
        return ""
    return linha[indice]


def calcular_resumo_meta_pontoweb_contexto(
    sessao,
    access_token,
    identificador_banco,
    funcionarios,
    tecnico,
    ano,
    mes,
):
    alias_pontoweb = obter_alias_pontoweb(tecnico)
    nomes_funcionarios = [
        funcionario.get("Nome", "")
        for funcionario in funcionarios
        if funcionario.get("Nome")
    ]
    funcionario = None

    for nome_referencia in [alias_pontoweb, tecnico, melhor_alias(tecnico, nomes_funcionarios)]:
        if not nome_referencia:
            continue
        funcionario = localizar_funcionario_por_nome(funcionarios, nome_referencia)
        if funcionario:
            break

    if not funcionario:
        raise ValueError(f"Não encontrei o técnico {tecnico} no PontoWeb.")

    data_inicial = date(ano, mes, 1)
    data_final = date(ano, mes, calendar.monthrange(ano, mes)[1])
    dados_calculo = buscar_calculos_pontoweb(
        sessao,
        access_token,
        identificador_banco,
        funcionario.get("Id"),
        data_inicial,
        data_final,
    )

    colunas = dados_calculo.get("Colunas", [])
    linhas = dados_calculo.get("Linhas", [])
    situacoes = dados_calculo.get("SituacaoDias", [])

    idx_data = indice_coluna_calculo(colunas, "Data")
    idx_faltas = indice_coluna_calculo(colunas, "Faltas")
    idx_carga = indice_coluna_calculo(colunas, "Carga")
    idx_bdeb = indice_coluna_calculo(colunas, "BDeb.", "BDeb")

    abatimento_total = 0.0
    dias_base = dias_uteis_para_meta(ano, mes)

    for indice_linha, linha in enumerate(linhas):
        data_texto = builtins.str(valor_linha_calculo(linha, idx_data)).strip()
        match_data = re.search(r"(\d{2}/\d{2}/\d{4})", data_texto)
        if not match_data:
            continue

        data_linha = pd.to_datetime(match_data.group(1), dayfirst=True, errors="coerce")
        if pd.isna(data_linha):
            continue

        data_linha = data_linha.date()
        if data_linha.year != ano or data_linha.month != mes:
            continue
        if data_linha.weekday() >= 5 or eh_feriado_federal(data_linha):
            continue

        valores_linha = [builtins.str(valor or "").strip().upper() for valor in linha]
        faltas_minutos = max(hhmm_para_minutos(valor_linha_calculo(linha, idx_faltas)), 0)
        carga_minutos = max(hhmm_para_minutos(valor_linha_calculo(linha, idx_carga)), 0)
        bdeb_minutos = max(hhmm_para_minutos(valor_linha_calculo(linha, idx_bdeb)), 0)
        situacao = situacoes[indice_linha] if indice_linha < len(situacoes) else None

        if any("FERIAS" in normalizar_nome(valor) for valor in valores_linha):
            abatimento_total += 1
            continue

        if any("ATESTAD" in normalizar_nome(valor) for valor in valores_linha):
            abatimento_total += 1
            continue

        if carga_minutos <= 0:
            continue

        if situacao == 2 and faltas_minutos >= carga_minutos:
            abatimento_total += 1
            continue

        minutos_ausentes = max(faltas_minutos, bdeb_minutos)
        if 0 < minutos_ausentes < carga_minutos:
            abatimento_total += minutos_ausentes / carga_minutos

    dias_ferias_programadas = dias_uteis_ferias_no_mes(tecnico, ano, mes)
    dias_com_abatimento_ponto = max(dias_base - abatimento_total, 0)
    dias_com_ferias_programadas = max(dias_base - dias_ferias_programadas, 0)
    dias_considerados = min(
        dias_com_abatimento_ponto,
        dias_com_ferias_programadas,
    )
    return {
        "dias_base": dias_base,
        "abatimento": round(abatimento_total, 2),
        "dias_ferias_programadas": dias_ferias_programadas,
        "dias_considerados": round(dias_considerados, 2),
    }


def calcular_resumo_meta_pontoweb(
    tecnico,
    ano,
    mes,
    email,
    senha,
    banco_id="",
    banco_identificador="",
):
    login = login_pontoweb(email, senha)
    sessao = login["sessao"]
    access_token = login["access_token"]

    toolbar = buscar_toolbar_pontoweb(sessao, access_token)
    lista_bancos = localizar_lista_bancos(toolbar)
    banco = selecionar_banco_pontoweb(lista_bancos, banco_id, banco_identificador)
    identificador_banco = builtins.str(
        banco.get("Identificador", banco.get("identificador", ""))
    )
    funcionarios = buscar_funcionarios_pontoweb(sessao, access_token, identificador_banco)
    return calcular_resumo_meta_pontoweb_contexto(
        sessao,
        access_token,
        identificador_banco,
        funcionarios,
        tecnico,
        ano,
        mes,
    )


@st.cache_data(ttl=1800, show_spinner=False)
def obter_resumo_meta_pontoweb(
    tecnico,
    ano,
    mes,
    email,
    senha,
    banco_id="",
    banco_identificador="",
):
    return calcular_resumo_meta_pontoweb(
        tecnico,
        ano,
        mes,
        email,
        senha,
        banco_id,
        banco_identificador,
    )


def criar_sessao_http():
    sessao = requests.Session()
    sessao.trust_env = False
    return sessao


def atualizar_via_servico_local(fonte, data_referencia):
    sessao = criar_sessao_http()
    resposta = sessao.post(
        f"{SERVICO_PLANILHA_LOCAL_URL}/atualizar",
        json={
            "fonte": fonte,
            "data": data_referencia.strftime("%d/%m/%Y"),
        },
        timeout=600,
    )
    resposta.raise_for_status()
    payload = resposta.json()
    if not payload.get("ok"):
        raise RuntimeError(payload.get("erro", "Falha na atualização local."))
    return payload.get("resultados", {})


def servico_local_disponivel():
    try:
        sessao = criar_sessao_http()
        resposta = sessao.get(
            f"{SERVICO_PLANILHA_LOCAL_URL}/health",
            timeout=5,
        )
        resposta.raise_for_status()
        payload = resposta.json()
        return bool(payload.get("ok"))
    except Exception:
        return False


@st.cache_data(ttl=300, show_spinner=False)
def listar_arquivos_ro():
    sessao = criar_sessao_http()
    resposta = sessao.get(RO_DRIVE_FOLDER_URL, timeout=30)
    resposta.raise_for_status()

    arquivos = []
    vistos = set()
    for arquivo_id, titulo in re.findall(
        r'data-id="([^"]+)"[^>]*data-tooltip="([^"]+)"',
        resposta.text,
    ):
        if arquivo_id in vistos:
            continue
        vistos.add(arquivo_id)
        arquivos.append(
            {
                "id": arquivo_id,
                "titulo": unescape(titulo),
            }
        )

    return arquivos


def planilha_ro_do_mes(data_referencia):
    chave_mes = (data_referencia.year, data_referencia.month)
    link_fixo = RO_PLANILHAS_FIXAS.get(chave_mes)
    if link_fixo:
        correspondencia = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", link_fixo)
        if correspondencia:
            return {
                "id": correspondencia.group(1),
                "titulo": f"RO {MESES_RO[data_referencia.month]} {data_referencia.year} (respostas)",
                "origem": "link_fixo",
                "url": link_fixo,
            }

    esperado = normalizar_cabecalho(
        f"RO {MESES_RO[data_referencia.month]} {data_referencia.year} (respostas)"
    )

    for arquivo in listar_arquivos_ro():
        if esperado in normalizar_cabecalho(arquivo["titulo"]):
            return {
                **arquivo,
                "origem": "drive",
                "url": f"https://docs.google.com/spreadsheets/d/{arquivo['id']}/edit",
            }

    raise ValueError("Não encontrei a planilha de respostas do RO para esse mês.")


def buscar_ro_forms(data_referencia):
    arquivo = planilha_ro_do_mes(data_referencia)
    sessao = criar_sessao_http()
    resposta = sessao.get(
        f"https://docs.google.com/spreadsheets/d/{arquivo['id']}/export?format=csv",
        timeout=60,
    )
    resposta.raise_for_status()

    texto_csv = resposta.content.decode("utf-8-sig", errors="replace")
    leitor = csv.DictReader(StringIO(texto_csv))
    contagens = {}
    tecnicos_origem = []
    motivos_validos = {
        normalizar_cabecalho(motivo) for motivo in MOTIVOS_RO_CONTAM
    }

    for linha in leitor:
        linha_normalizada = {
            normalizar_cabecalho(chave): (valor or "").strip()
            for chave, valor in linha.items()
            if chave
        }

        tecnico = linha_normalizada.get("nome tecnico", "")
        data_texto = linha_normalizada.get("data", "")
        motivo = linha_normalizada.get("ligacao inferior a 2 min", "")

        if (
            not tecnico
            or not data_texto
            or normalizar_cabecalho(motivo) not in motivos_validos
        ):
            continue

        data_linha = pd.to_datetime(data_texto, dayfirst=True, errors="coerce")
        if pd.isna(data_linha) or data_linha.date() != data_referencia:
            continue

        chave = normalizar_nome(tecnico)
        contagens[chave] = contagens.get(chave, 0) + 1
        tecnicos_origem.append(tecnico)

    tentar_salvar_backup_ro_no_banco(
        data_referencia,
        arquivo,
        texto_csv,
        contagens,
        tecnicos_origem,
    )

    return {
        "arquivo": arquivo["titulo"],
        "contagens": contagens,
        "tecnicos_origem": tecnicos_origem,
    }


def headers_powerbi():
    return {
        "Accept": "application/json",
        "ActivityId": builtins.str(uuid.uuid4()),
        "RequestId": builtins.str(uuid.uuid4()),
        "X-PowerBI-ResourceKey": POWERBI_RESOURCE_KEY,
        "User-Agent": "Mozilla/5.0",
        "Origin": "https://app.powerbi.com",
        "Referer": "https://app.powerbi.com/",
        "Content-Type": "application/json",
    }


@st.cache_data(ttl=300, show_spinner=False)
def carregar_metadados_powerbi():
    url = (
        f"{POWERBI_API_BASE}/public/reports/{POWERBI_RESOURCE_KEY}"
        "/modelsAndExploration?preferReadOnlySession=true"
    )
    sessao = criar_sessao_http()
    resposta = sessao.get(url, headers=headers_powerbi(), timeout=30)
    resposta.raise_for_status()
    return resposta.json()


def trimestre_powerbi(data_referencia):
    return f"Trim {((data_referencia.month - 1) // 3) + 1}"


@st.cache_data(ttl=300, show_spinner=False)
def carregar_filas_folha_bi():
    try:
        sessao = criar_sessao_http()
        resposta = sessao.get(
            f"{X2_CONTROLLER_BASE_URL}/api/ami/queues",
            timeout=30,
        )
        resposta.raise_for_status()
        dados = resposta.json()
        filas = []

        for fila in dados.get("queues", []):
            nome = builtins.str(fila.get("name") or "").strip()
            if nome and "folha" in nome.lower() and nome not in filas:
                filas.append(nome)

        if filas:
            return filas
    except Exception:
        pass

    return [POWERBI_FILA_FALLBACK]


def aplicar_filtro_fila_powerbi(query):
    fontes = query.setdefault("From", [])
    if not any(
        fonte.get("Entity") == "Filas" and fonte.get("Name") == "f"
        for fonte in fontes
    ):
        fontes.append({"Name": "f", "Entity": "Filas", "Type": 0})

    valores_fila = [
        [{"Literal": {"Value": f"'{fila}'"}}]
        for fila in carregar_filas_folha_bi()
    ]

    query.setdefault("Where", []).append(
        {
            "Condition": {
                "In": {
                    "Expressions": [
                        {
                            "Column": {
                                "Expression": {"SourceRef": {"Source": "f"}},
                                "Property": "Fila",
                            }
                        }
                    ],
                    "Values": valores_fila,
                }
            }
        }
    )


def aplicar_filtro_data_powerbi(query, data_referencia):
    fonte_data = next(
        (
            fonte
            for fonte in query.setdefault("From", [])
            if builtins.str(fonte.get("Entity", "")).startswith("LocalDateTable_")
        ),
        None,
    )
    if fonte_data is None:
        fonte_data = {
            "Name": "l",
            "Entity": "LocalDateTable_5b42e521-bbb7-4d39-8e7e-735e21b48675",
            "Type": 0,
        }
        query["From"].append(fonte_data)

    nome_fonte = fonte_data["Name"]
    query.setdefault("Where", []).insert(
        0,
        {
            "Condition": {
                "In": {
                    "Expressions": [
                        {
                            "Column": {
                                "Expression": {"SourceRef": {"Source": nome_fonte}},
                                "Property": "Ano",
                            }
                        },
                        {
                            "Column": {
                                "Expression": {"SourceRef": {"Source": nome_fonte}},
                                "Property": "Trimestre",
                            }
                        },
                        {
                            "Column": {
                                "Expression": {"SourceRef": {"Source": nome_fonte}},
                                "Property": "Mês",
                            }
                        },
                        {
                            "Column": {
                                "Expression": {"SourceRef": {"Source": nome_fonte}},
                                "Property": "Dia",
                            }
                        },
                    ],
                    "Values": [
                        [
                            {"Literal": {"Value": f"{data_referencia.year}L"}},
                            {
                                "Literal": {
                                    "Value": f"'{trimestre_powerbi(data_referencia)}'"
                                }
                            },
                            {
                                "Literal": {
                                    "Value": f"'{MESES_POWERBI[data_referencia.month]}'"
                                }
                            },
                            {"Literal": {"Value": f"{data_referencia.day}L"}},
                        ]
                    ],
                }
            }
        },
    )


def consulta_visual_powerbi(visual):
    if visual.get("query"):
        return json.loads(visual["query"])

    configuracao = json.loads(visual["config"])
    visual_config = configuracao["singleVisual"]
    query = json.loads(
        json.dumps(
            visual_config["prototypeQuery"],
            ensure_ascii=False,
        )
    )
    tipo_visual = visual_config.get("visualType")
    total_selecoes = len(query.get("Select", []))

    if tipo_visual == "tableEx":
        binding = {
            "Primary": {
                "Groupings": [
                    {
                        "Projections": list(range(total_selecoes)),
                        "Subtotal": 1,
                    }
                ]
            },
            "DataReduction": {
                "DataVolume": 3,
                "Primary": {"Window": {"Count": 500}},
            },
            "Version": 1,
        }
    else:
        binding = {
            "Primary": {
                "Groupings": [
                    {
                        "Projections": list(range(total_selecoes)),
                    }
                ]
            },
            "DataReduction": {
                "DataVolume": 3,
                "Primary": {"Top": {}},
            },
            "Version": 1,
        }

    return {
        "Commands": [
            {
                "SemanticQueryDataShapeCommand": {
                    "Query": query,
                    "Binding": binding,
                    "ExecutionMetricsKind": 1,
                }
            }
        ]
    }


def montar_consulta_powerbi(data_referencia):
    metadados = carregar_metadados_powerbi()
    secao_mapa = next(
        secao
        for secao in metadados["exploration"]["sections"]
        if secao.get("displayName") == "Mapa"
    )
    visual = secao_mapa["visualContainers"][0]
    consulta = consulta_visual_powerbi(visual)
    comando = consulta["Commands"][0]["SemanticQueryDataShapeCommand"]
    comando["Query"]["Where"] = []
    aplicar_filtro_data_powerbi(comando["Query"], data_referencia)
    aplicar_filtro_fila_powerbi(comando["Query"])

    return {
        "version": "1.0.0",
        "queries": [
            {
                "Query": consulta,
                "ApplicationContext": {
                    "DatasetId": builtins.str(metadados["models"][0]["id"]),
                    "Sources": [
                        {
                            "ReportId": metadados["exploration"]["reportId"],
                            "VisualId": json.loads(visual["config"])["name"],
                        }
                    ],
                },
            }
        ],
        "cancelQueries": [],
        "modelId": metadados["models"][0]["id"],
    }


def montar_consulta_abandonadas_powerbi(data_referencia):
    metadados = carregar_metadados_powerbi()
    secao_abandonos = next(
        secao
        for secao in metadados["exploration"]["sections"]
        if secao.get("displayName") == "Abandonos"
    )
    visual = next(
        visual
        for visual in secao_abandonos["visualContainers"]
        if "Abandonadas.MAB" in visual.get("query", "")
        or "Abandonadas.MAB" in visual.get("config", "")
    )
    consulta = consulta_visual_powerbi(visual)
    comando = consulta["Commands"][0]["SemanticQueryDataShapeCommand"]
    query = comando["Query"]
    query["Where"] = []
    aplicar_filtro_data_powerbi(query, data_referencia)
    query["Select"] = [
        {
            "Aggregation": {
                "Expression": {
                    "Column": {
                        "Expression": {"SourceRef": {"Source": "a"}},
                        "Property": "Evento",
                    }
                },
                "Function": 5,
            },
            "Name": "CountNonNull(Abandonadas.Evento)",
            "NativeReferenceName": "Abandonadas",
        }
    ]
    aplicar_filtro_fila_powerbi(query)

    return {
        "version": "1.0.0",
        "queries": [
            {
                "Query": consulta,
                "ApplicationContext": {
                    "DatasetId": builtins.str(metadados["models"][0]["id"]),
                    "Sources": [
                        {
                            "ReportId": metadados["exploration"]["reportId"],
                            "VisualId": json.loads(visual["config"])["name"],
                        }
                    ],
                },
            }
        ],
        "cancelQueries": [],
        "modelId": metadados["models"][0]["id"],
    }


def valor_powerbi(linha, indice_coluna, valores, indice_valor, anterior):
    repetidos = int(linha.get("R", 0))
    nulos = int(linha.get("Ø", 0))

    if nulos & (1 << indice_coluna):
        return None, indice_valor
    if repetidos & (1 << indice_coluna):
        return anterior[indice_coluna], indice_valor
    return valores[indice_valor], indice_valor + 1


def segundos_para_tma_planilha(valor):
    segundos = int(round(float(valor or 0)))
    minutos, resto = divmod(segundos, 60)
    return round(minutos + (resto / 100), 2)


def decodificar_linhas_powerbi(linhas):
    registros = []
    anterior = [None] * 8

    for linha in linhas:
        valores = linha.get("C", [])
        atual = []
        indice_valor = 0

        for indice_coluna in range(8):
            valor, indice_valor = valor_powerbi(
                linha,
                indice_coluna,
                valores,
                indice_valor,
                anterior,
            )
            atual.append(valor)

        anterior = atual

        if atual[0]:
            registros.append(
                {
                    "agente": atual[0],
                    "atendidas": int(atual[1] or 0),
                    "maior_2min": int(atual[2] or 0),
                    "tma": segundos_para_tma_planilha(atual[7] or 0),
                }
            )

    return registros


def buscar_produtividade_powerbi(data_referencia):
    sessao = criar_sessao_http()
    resposta = sessao.post(
        f"{POWERBI_API_BASE}/public/reports/querydata?synchronous=true",
        headers=headers_powerbi(),
        json=montar_consulta_powerbi(data_referencia),
        timeout=45,
    )
    resposta.raise_for_status()
    dados = resposta.json()
    linhas = (
        dados["results"][0]["result"]["data"]["dsr"]["DS"][0]["PH"][1]["DM1"]
    )
    return decodificar_linhas_powerbi(linhas)


def buscar_abandonadas_powerbi(data_referencia):
    sessao = criar_sessao_http()
    resposta = sessao.post(
        f"{POWERBI_API_BASE}/public/reports/querydata?synchronous=true",
        headers=headers_powerbi(),
        json=montar_consulta_abandonadas_powerbi(data_referencia),
        timeout=45,
    )
    resposta.raise_for_status()
    dados = resposta.json()
    data_shape = dados["results"][0]["result"]["data"]["dsr"]["DS"][0]["PH"][0]
    linhas = data_shape.get("DM0") or data_shape.get("DM1") or []
    if not linhas:
        return 0
    return int(round(float(linhas[0].get("M0") or 0)))


def atualizar_planilha_com_bi(data_referencia):
    registros_bi = buscar_produtividade_powerbi(data_referencia)
    abandonadas = buscar_abandonadas_powerbi(data_referencia)
    registros_por_agente = {
        normalizar_nome(registro["agente"]): registro
        for registro in registros_bi
    }

    wb = carregar_workbook_produtividade()
    ws = wb[ABA_PRODUTIVIDADE]
    colunas = cabecalhos(ws)
    col_data = coluna(colunas, "Data")
    col_tecnico = coluna(colunas, "Técnico")
    col_atendidas = coluna(colunas, "Atendidas")
    col_2min = coluna(colunas, " > 2min", "> 2min", ">2min")
    col_tma = coluna(colunas, "TMA")
    col_abandonadas = garantir_coluna(ws, COLUNA_ABANDONADAS)
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_abandonadas)].hidden = True
    linhas_criadas = garantir_linhas_da_data(ws, colunas, data_referencia)
    tecnicos = tecnicos_da_planilha(ws, colunas)
    ws_aliases = garantir_aba_aliases(
        wb,
        tecnicos,
        [registro["agente"] for registro in registros_bi],
        coluna_alias=2,
    )
    aliases = ler_aliases(ws_aliases, coluna_alias=2)

    atualizados = []
    sem_alias = []

    for row in range(2, ws.max_row + 1):
        data_linha = ws.cell(row=row, column=col_data).value
        if not data_linha:
            continue
        data_linha = data_linha.date() if hasattr(data_linha, "date") else data_linha
        if data_linha != data_referencia:
            continue

        tecnico = ws.cell(row=row, column=col_tecnico).value
        ws.cell(row=row, column=col_atendidas).value = 0
        ws.cell(row=row, column=col_2min).value = 0
        ws.cell(row=row, column=col_tma).value = 0
        ws.cell(row=row, column=col_abandonadas).value = abandonadas
        chave_agente = aliases.get(normalizar_nome(tecnico), normalizar_nome(tecnico))
        registro = registros_por_agente.get(chave_agente)

        if not registro:
            sem_alias.append(tecnico)
            continue

        ws.cell(row=row, column=col_atendidas).value = registro["atendidas"]
        ws.cell(row=row, column=col_2min).value = registro["maior_2min"]
        ws.cell(row=row, column=col_tma).value = registro["tma"]
        atualizados.append(tecnico)

    zerar_colunas_sem_movimento_ws(ws)
    salvar_workbook_produtividade(wb)

    return {
        "data": data_referencia.strftime("%d/%m/%Y"),
        "fonte": len(registros_bi),
        "abandonadas": abandonadas,
        "linhas_criadas": linhas_criadas,
        "atualizados": len(atualizados),
        "sem_alias": sorted(set(sem_alias)),
    }


def login_sgd(usuario, senha):
    sessao = criar_sessao_http()
    resposta = sessao.post(
        SGD_LOGIN_URL,
        data={
            "p_chat_plug": "",
            "j_username": usuario,
            "j_password": senha,
            "loginBtn": "Entrar",
        },
        timeout=30,
    )
    resposta.raise_for_status()
    if "login.html" in resposta.url or "invalido.html" in resposta.url:
        raise ValueError("Login do SGD inválido.")
    return sessao


def montar_campos_sgd(html, data_inicial, data_final, extensao="EXCEL_XLSX"):
    parser = FormularioSGDParser()
    parser.feed(html)
    dados = []
    submit_name = "formFiltroRelatorio:j_id_jsp_1813582124_61"

    for attrs in parser.inputs:
        nome = attrs.get("name")
        tipo = (attrs.get("type") or "").lower()

        if tipo == "submit" and attrs.get("value", "").startswith("Gerar"):
            submit_name = nome

        if not nome or tipo in {"submit", "button", "image", "reset"}:
            continue

        if tipo == "radio":
            if "checked" in attrs:
                dados.append((nome, attrs.get("value", "")))
        elif tipo == "checkbox":
            if nome not in {
                "formFiltroRelatorio:desconsiderarSscAnexadaSaNe",
                "formFiltroRelatorio:considerarUnidadesAtendidas",
            }:
                dados.append((nome, attrs.get("value", "on")))
        else:
            dados.append((nome, attrs.get("value", "")))

    for select, opcao in parser.selects:
        nome = select.get("name")
        if nome:
            dados.append((nome, opcao.get("value", "")))

    def definir_unico(nome, valor):
        nonlocal dados
        dados = [(chave, item) for chave, item in dados if chave != nome]
        dados.append((nome, valor))

    definir_unico("formFiltroRelatorio:dataInicial", data_inicial.strftime("%d/%m/%y"))
    definir_unico("formFiltroRelatorio:dataFinal", data_final.strftime("%d/%m/%y"))
    definir_unico("formFiltroRelatorio:extensao", extensao)
    dados.append((submit_name, "Gerar Relatório"))
    return dados


def gerar_relatorio_sgd(sessao, data_inicial, data_final):
    resposta = sessao.get(SGD_RELATORIO_URL, timeout=30)
    resposta.raise_for_status()
    html = resposta.text

    link_download = None
    for _ in range(10):
        dados = montar_campos_sgd(html, data_inicial, data_final)
        ultima_excecao = None
        for tentativa in range(1, SGD_RELATORIO_TENTATIVAS + 1):
            try:
                resposta = sessao.post(
                    SGD_RELATORIO_URL,
                    data=dados,
                    timeout=SGD_RELATORIO_TIMEOUT,
                )
                resposta.raise_for_status()
                break
            except requests.exceptions.ReadTimeout as exc:
                ultima_excecao = exc
                if tentativa == SGD_RELATORIO_TENTATIVAS:
                    periodo = (
                        f"{data_inicial.strftime('%d/%m/%Y')} a "
                        f"{data_final.strftime('%d/%m/%Y')}"
                    )
                    raise TimeoutError(
                        "O SGD demorou mais do que o esperado para gerar o "
                        f"relatÃ³rio do perÃ­odo {periodo}. Tente novamente em "
                        "alguns minutos."
                    ) from exc
                time.sleep(3 * tentativa)
        else:
            if ultima_excecao:
                raise ultima_excecao

        tipo = resposta.headers.get("content-type", "")
        if "spreadsheet" in tipo:
            return resposta.content

        html = resposta.text
        match = re.search(r'href="([^"]+\.xlsx)"', html)
        if match:
            link_download = match.group(1)
            break

    if not link_download:
        raise ValueError("O SGD não gerou o arquivo Excel do relatório.")

    resposta = sessao.get(f"{SGD_BASE_URL}{link_download}", timeout=60)
    resposta.raise_for_status()
    return resposta.content


def extrair_registros_sgd(conteudo_xlsx):
    wb = openpyxl.load_workbook(BytesIO(conteudo_xlsx), data_only=True)
    ws = wb.active
    registros = []

    for row in range(10, ws.max_row + 1):
        tecnico = ws.cell(row=row, column=2).value
        if not tecnico or builtins.str(tecnico).strip().lower().startswith("total"):
            continue

        registros.append(
            {
                "tecnico": tecnico,
                "satisfacao": float(ws.cell(row=row, column=5).value or 0),
                "votacao": float(ws.cell(row=row, column=11).value or 0),
                "ssc": int(ws.cell(row=row, column=10).value or 0),
            }
        )

    return registros


def buscar_satisfacao_sgd(data_inicial, data_final, usuario, senha):
    sessao = login_sgd(usuario, senha)
    conteudo = gerar_relatorio_sgd(sessao, data_inicial, data_final)
    return extrair_registros_sgd(conteudo)


def buscar_satisfacao_sgd_diaria(data_referencia, usuario, senha):
    sessao = login_sgd(usuario, senha)
    registros = []

    if data_referencia.weekday() >= 5:
        return registros

    try:
        conteudo = gerar_relatorio_sgd(sessao, data_referencia, data_referencia)
    except ValueError:
        return registros

    for registro in extrair_registros_sgd(conteudo):
        registro["data"] = data_referencia
        registros.append(registro)

    return registros


def atualizar_planilha_com_sgd(data_referencia, usuario, senha):
    data_inicial = inicio_mes(data_referencia)
    registros_periodo = buscar_satisfacao_sgd(
        data_inicial,
        data_referencia,
        usuario,
        senha,
    )
    registros_sgd = buscar_satisfacao_sgd_diaria(
        data_referencia,
        usuario,
        senha,
    )
    registros_periodo_por_tecnico = {
        normalizar_nome(registro["tecnico"]): registro
        for registro in registros_periodo
    }
    registros_por_chave = {
        (normalizar_nome(registro["tecnico"]), registro["data"]): registro
        for registro in registros_sgd
    }

    wb = carregar_workbook_produtividade()
    ws = wb[ABA_PRODUTIVIDADE]
    colunas = cabecalhos(ws)
    col_data = coluna(colunas, "Data")
    col_tecnico = coluna(colunas, "Técnico")
    col_ssc = coluna(colunas, "SSC")
    col_satisfacao = coluna(colunas, "Satisfação")
    col_votacao = coluna(colunas, "Votação")
    linhas_criadas = garantir_linhas_da_data(ws, colunas, data_referencia)
    tecnicos = tecnicos_da_planilha(ws, colunas)
    ws_aliases = garantir_aba_aliases(
        wb,
        tecnicos,
        [registro["tecnico"] for registro in registros_sgd],
        coluna_alias=3,
    )
    aliases = ler_aliases(ws_aliases, coluna_alias=3)

    atualizados = []
    sem_alias = []

    for row in range(2, ws.max_row + 1):
        data_linha = ws.cell(row=row, column=col_data).value
        if not data_linha:
            continue
        data_linha = data_linha.date() if hasattr(data_linha, "date") else data_linha
        if data_linha != data_referencia:
            continue

        tecnico = ws.cell(row=row, column=col_tecnico).value
        chave_sgd = aliases.get(normalizar_nome(tecnico), normalizar_nome(tecnico))
        registro_periodo = registros_periodo_por_tecnico.get(chave_sgd)
        registro = registros_por_chave.get((chave_sgd, data_linha))

        if not registro_periodo and not registro:
            ws.cell(row=row, column=col_ssc).value = 0
            ws.cell(row=row, column=col_satisfacao).value = 0
            ws.cell(row=row, column=col_votacao).value = 0
            sem_alias.append(tecnico)
            continue

        ws.cell(row=row, column=col_ssc).value = 0 if not registro else registro["ssc"]
        ws.cell(row=row, column=col_satisfacao).value = (
            0 if not registro_periodo else registro_periodo["satisfacao"]
        )
        ws.cell(row=row, column=col_votacao).value = (
            0 if not registro_periodo else registro_periodo["votacao"]
        )
        atualizados.append(f"{tecnico} - {data_linha}")

    zerar_colunas_sem_movimento_ws(ws)
    salvar_workbook_produtividade(wb)

    return {
        "data": data_referencia.strftime("%d/%m/%Y"),
        "periodo": f"{data_inicial.strftime('%d/%m/%Y')} a {data_referencia.strftime('%d/%m/%Y')}",
        "fonte": len(registros_sgd),
        "dias_processados": 1,
        "linhas_criadas": linhas_criadas,
        "atualizados": len(atualizados),
        "sem_alias": sorted(set(sem_alias)),
    }


def atualizar_planilha_com_ro(data_referencia):
    dados_ro = buscar_ro_forms(data_referencia)
    contagens = dados_ro["contagens"]

    wb = carregar_workbook_produtividade()
    ws = wb[ABA_PRODUTIVIDADE]
    colunas = cabecalhos(ws)
    col_data = coluna(colunas, "Data")
    col_tecnico = coluna(colunas, "Tecnico", "T?cnico", "T??cnico")
    col_ro = coluna(colunas, "RO")
    linhas_criadas = garantir_linhas_da_data(ws, colunas, data_referencia)
    tecnicos = tecnicos_da_planilha(ws, colunas)
    ws_aliases = garantir_aba_aliases(
        wb,
        tecnicos,
        dados_ro.get("tecnicos_origem", list(contagens.keys())),
        coluna_alias=6,
    )
    aliases = ler_aliases(ws_aliases, coluna_alias=6)

    atualizados = 0
    sem_alias = []

    for row in range(2, ws.max_row + 1):
        data_linha = ws.cell(row=row, column=col_data).value
        if not data_linha:
            continue
        data_linha = data_linha.date() if hasattr(data_linha, "date") else data_linha
        if data_linha != data_referencia:
            continue

        tecnico = ws.cell(row=row, column=col_tecnico).value
        chave_ro = aliases.get(normalizar_nome(tecnico), normalizar_nome(tecnico))
        valor_ro = contagens.get(chave_ro, 0)
        ws.cell(row=row, column=col_ro).value = valor_ro
        if not valor_ro and chave_ro not in contagens:
            sem_alias.append(tecnico)
        atualizados += 1

    zerar_colunas_sem_movimento_ws(ws)
    salvar_workbook_produtividade(wb)

    return {
        "data": data_referencia.strftime("%d/%m/%Y"),
        "arquivo": dados_ro["arquivo"],
        "fonte": sum(contagens.values()),
        "linhas_criadas": linhas_criadas,
        "atualizados": atualizados,
        "sem_alias": sorted(set(sem_alias)),
    }

def localizar_navegador_plug():
    candidatos = [
        os.getenv("PLUG_BROWSER_PATH", ""),
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    ]
    for caminho in candidatos:
        if caminho and os.path.exists(caminho):
            return caminho
    return None


def periodo_chat_plug(data_referencia):
    inicio = datetime.combine(data_referencia, datetime.min.time())
    fim = datetime.combine(data_referencia, datetime.max.time()).replace(
        hour=23,
        minute=59,
        second=0,
        microsecond=0,
    )
    return (
        inicio.strftime("%d/%m/%Y %H:%M"),
        fim.strftime("%d/%m/%Y %H:%M"),
        inicio.strftime("%d/%m/%Y %H:%M")
        + " - "
        + fim.strftime("%d/%m/%Y %H:%M"),
    )


def abrir_sessao_plug():
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError(
            "A biblioteca playwright não está instalada nesta máquina."
        ) from exc

    playwright = sync_playwright().start()
    contexto = None
    try:
        kwargs = {
            "user_data_dir": builtins.str(PLUG_PERFIL_DIR.resolve()),
            "headless": False,
        }
        executable_path = localizar_navegador_plug()
        if executable_path:
            kwargs["executable_path"] = executable_path
        contexto = playwright.chromium.launch_persistent_context(**kwargs)
        pagina = contexto.pages[0] if contexto.pages else contexto.new_page()
        pagina.goto(PLUG_CHATBOX_URL, wait_until="domcontentloaded", timeout=60000)

        limite = time.time() + 180
        while time.time() < limite:
            url_atual = pagina.url
            if "#/app/" in url_atual and "login" not in url_atual.lower():
                pagina.goto(
                    PLUG_CHATBOX_URL,
                    wait_until="domcontentloaded",
                    timeout=60000,
                )
                return playwright, contexto, pagina
            pagina.wait_for_timeout(1000)

        raise TimeoutError(
            "Faça login no PLUG na janela aberta e tente novamente."
        )
    except Exception:
        if contexto:
            contexto.close()
        playwright.stop()
        raise


def localizar_mes_no_calendario_plug(pagina, texto_mes_ano):
    locator = pagina.locator(".daterangepicker .month")
    textos = [texto.strip() for texto in locator.all_text_contents()]
    if texto_mes_ano in textos:
        return textos.index(texto_mes_ano)
    return -1


def ajustar_periodo_plug(pagina, data_referencia):
    _, _, periodo_texto = periodo_chat_plug(data_referencia)
    pagina.locator('input[name="datetimesCreated"]').click()
    pagina.wait_for_selector(".daterangepicker", timeout=15000)

    meses_pt = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Março",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro",
    }
    texto_mes_ano = f"{meses_pt[data_referencia.month]} {data_referencia.year}"

    for _ in range(24):
        indice_mes = localizar_mes_no_calendario_plug(pagina, texto_mes_ano)
        if indice_mes >= 0:
            break
        pagina.locator(".daterangepicker .prev.available").click()
        pagina.wait_for_timeout(200)
    else:
        raise RuntimeError("Não foi possível localizar o mês no calendário do PLUG.")

    calendario_alvo = ".drp-calendar.left" if indice_mes == 0 else ".drp-calendar.right"
    dia = builtins.str(data_referencia.day)
    celulas = pagina.locator(
        f"{calendario_alvo} td.available, {calendario_alvo} td.today.available"
    )

    encontrou = False
    for indice in range(celulas.count()):
        texto_celula = builtins.str(celulas.nth(indice).inner_text()).strip()
        if texto_celula == dia:
            celulas.nth(indice).click()
            pagina.wait_for_timeout(150)
            encontrou = True
            break
    if not encontrou:
        raise RuntimeError("Não foi possível selecionar o dia no calendário do PLUG.")

    pagina.locator(".drp-calendar.left .hourselect").select_option("0")
    pagina.locator(".drp-calendar.left .minuteselect").select_option("0")
    pagina.locator(".drp-calendar.right .hourselect").select_option("23")
    pagina.locator(".drp-calendar.right .minuteselect").select_option("59")
    pagina.locator(".daterangepicker .applyBtn").click()
    pagina.wait_for_timeout(1200)

    valor_atual = pagina.locator('input[name="datetimesCreated"]').input_value().strip()
    if valor_atual != periodo_texto:
        raise RuntimeError(
            f"Período do PLUG não foi aplicado corretamente: {valor_atual}"
        )


def selecionar_status_plug(pagina, status):
    pagina.locator("ng-select .ng-select-container").click()
    pagina.wait_for_selector('ng-dropdown-panel[role="listbox"]', timeout=10000)
    opcao = pagina.locator('ng-dropdown-panel [role="option"]', has_text=status)
    if opcao.count() != 1:
        raise RuntimeError(f"Status '{status}' não encontrado no PLUG.")
    opcao.click()
    pagina.wait_for_timeout(1200)


def obter_atendentes_plug(pagina):
    return pagina.evaluate(
        """() => Array.from(document.querySelectorAll('select'))[2]
            ? Array.from(document.querySelectorAll('select'))[2].options
                .filter((opt) => opt.value)
                .map((opt) => ({
                    value: opt.value,
                    text: opt.textContent.trim()
                }))
            : []"""
    )


def total_chat_plug_atendente(pagina, value_atendente):
    selects = pagina.locator("select")
    if selects.count() < 3:
        raise RuntimeError("Filtro de atendente não encontrado no PLUG.")
    selects.nth(2).select_option(value_atendente)
    pagina.wait_for_timeout(1200)
    total_texto = pagina.locator("text=Total:").first.inner_text(timeout=10000)
    match = re.search(r"Total:\s*(\d+)", total_texto)
    if not match:
        raise RuntimeError("Não foi possível ler o total de chats no PLUG.")
    return int(match.group(1))


def buscar_chat_plug(data_referencia):
    playwright, contexto, pagina = abrir_sessao_plug()
    try:
        pagina.goto(PLUG_CHATBOX_URL, wait_until="domcontentloaded", timeout=60000)
        pagina.wait_for_selector('button[title="Filtros"]', timeout=30000)
        pagina.locator('button[title="Filtros"]').click()
        pagina.wait_for_selector('input[name="datetimesCreated"]', timeout=15000)
        ajustar_periodo_plug(pagina, data_referencia)

        selects = pagina.locator("select")
        if selects.count() < 5:
            raise RuntimeError("Filtros principais do PLUG não foram carregados.")
        selects.nth(4).select_option(label=PLUG_GRUPO_FOLHA)
        pagina.wait_for_timeout(1200)
        selecionar_status_plug(pagina, PLUG_STATUS_FECHADO)

        atendentes = obter_atendentes_plug(pagina)
        contagens = {}
        for atendente in atendentes:
            contagens[normalizar_nome(atendente["text"])] = total_chat_plug_atendente(
                pagina,
                atendente["value"],
            )

        return {
            "contagens": contagens,
            "atendentes": [item["text"] for item in atendentes],
        }
    finally:
        contexto.close()
        playwright.stop()


def atualizar_planilha_com_chat(data_referencia):
    dados_chat = buscar_chat_plug(data_referencia)
    contagens = dados_chat["contagens"]

    wb = carregar_workbook_produtividade()
    ws = wb[ABA_PRODUTIVIDADE]
    colunas = cabecalhos(ws)
    col_data = coluna(colunas, "Data")
    col_tecnico = coluna(colunas, "Técnico")
    col_chat = coluna(colunas, "CHAT")
    linhas_criadas = garantir_linhas_da_data(ws, colunas, data_referencia)
    tecnicos = tecnicos_da_planilha(ws, colunas)
    ws_aliases = garantir_aba_aliases(
        wb,
        tecnicos,
        dados_chat["atendentes"],
        coluna_alias=5,
    )
    aliases = ler_aliases(ws_aliases, coluna_alias=5)

    atualizados = 0
    sem_alias = []

    for row in range(2, ws.max_row + 1):
        data_linha = ws.cell(row=row, column=col_data).value
        if not data_linha:
            continue
        data_linha = data_linha.date() if hasattr(data_linha, "date") else data_linha
        if data_linha != data_referencia:
            continue

        tecnico = ws.cell(row=row, column=col_tecnico).value
        chave_chat = aliases.get(normalizar_nome(tecnico), normalizar_nome(tecnico))
        if chave_chat not in contagens:
            ws.cell(row=row, column=col_chat).value = 0
            sem_alias.append(tecnico)
            continue

        ws.cell(row=row, column=col_chat).value = contagens[chave_chat]
        atualizados += 1

    zerar_colunas_sem_movimento_ws(ws)
    salvar_workbook_produtividade(wb)

    return {
        "data": data_referencia.strftime("%d/%m/%Y"),
        "fonte": len(contagens),
        "linhas_criadas": linhas_criadas,
        "atualizados": atualizados,
        "sem_alias": sorted(set(sem_alias)),
    }


def atualizar_planilha_com_pontoweb_mes(
    ano,
    mes,
    email,
    senha,
    banco_id="",
    banco_identificador="",
):
    login = login_pontoweb(email, senha)
    sessao = login["sessao"]
    access_token = login["access_token"]
    toolbar = buscar_toolbar_pontoweb(sessao, access_token)
    lista_bancos = localizar_lista_bancos(toolbar)
    banco = selecionar_banco_pontoweb(lista_bancos, banco_id, banco_identificador)
    identificador_banco = builtins.str(
        banco.get("Identificador", banco.get("identificador", ""))
    )
    funcionarios = buscar_funcionarios_pontoweb(sessao, access_token, identificador_banco)
    nomes_funcionarios = [
        funcionario.get("Nome", "")
        for funcionario in funcionarios
        if funcionario.get("Nome")
    ]

    wb = carregar_workbook_produtividade()
    ws = wb[ABA_PRODUTIVIDADE]
    colunas = cabecalhos(ws)
    col_data = coluna(colunas, "Data")
    col_tecnico = coluna(colunas, "Técnico")
    col_dias_meta = garantir_coluna(ws, COLUNA_DIAS_META)
    colunas = cabecalhos(ws)

    tecnicos_mes = sorted(
        {
            builtins.str(ws.cell(row=row, column=col_tecnico).value).strip()
            for row in range(2, ws.max_row + 1)
            if ws.cell(row=row, column=col_tecnico).value
            and (
                (
                    ws.cell(row=row, column=col_data).value.date()
                    if hasattr(ws.cell(row=row, column=col_data).value, "date")
                    else ws.cell(row=row, column=col_data).value
                ).year
                == ano
            )
            and (
                (
                    ws.cell(row=row, column=col_data).value.date()
                    if hasattr(ws.cell(row=row, column=col_data).value, "date")
                    else ws.cell(row=row, column=col_data).value
                ).month
                == mes
            )
        }
    )

    ws_aliases = garantir_aba_aliases(
        wb,
        tecnicos_mes,
        nomes_funcionarios,
        coluna_alias=4,
    )
    aliases_pontoweb = ler_aliases(ws_aliases, coluna_alias=4)

    dias_por_tecnico = {}
    erros = {}
    for tecnico in tecnicos_mes:
        try:
            nome_referencia = aliases_pontoweb.get(normalizar_nome(tecnico), tecnico)
            resumo = calcular_resumo_meta_pontoweb_contexto(
                sessao,
                access_token,
                identificador_banco,
                funcionarios,
                nome_referencia,
                ano,
                mes,
            )
            dias_por_tecnico[normalizar_nome(tecnico)] = resumo["dias_considerados"]
        except Exception as erro:
            erros[tecnico] = builtins.str(erro)

    atualizados = 0
    for row in range(2, ws.max_row + 1):
        data_linha = ws.cell(row=row, column=col_data).value
        if not data_linha:
            continue
        data_linha = data_linha.date() if hasattr(data_linha, "date") else data_linha
        if data_linha.year != ano or data_linha.month != mes:
            continue

        tecnico = ws.cell(row=row, column=col_tecnico).value
        chave = normalizar_nome(tecnico)
        if chave in dias_por_tecnico:
            ws.cell(row=row, column=col_dias_meta).value = dias_por_tecnico[chave]
            atualizados += 1

    salvar_workbook_produtividade(wb)
    return {
        "tecnicos_processados": len(dias_por_tecnico),
        "linhas_atualizadas": atualizados,
        "erros": erros,
    }


def zerar_colunas_sem_movimento_ws(ws):
    colunas = cabecalhos(ws)
    col_atendidas = coluna(colunas, "Atendidas")
    col_chat = coluna(colunas, "CHAT")
    col_ro = coluna(colunas, "RO")
    colunas_zeradas = [
        col_atendidas,
        coluna(colunas, " > 2min", "> 2min", ">2min"),
        coluna(colunas, "TMA"),
        col_ro,
        col_chat,
        coluna(colunas, "Realizado"),
        coluna(colunas, "Esperado"),
        coluna(colunas, "Desvio"),
    ]

    for row in range(2, ws.max_row + 1):
        atendidas = ws.cell(row=row, column=col_atendidas).value or 0
        chat = ws.cell(row=row, column=col_chat).value or 0
        ro = ws.cell(row=row, column=col_ro).value or 0
        try:
            atendidas = float(atendidas)
        except Exception:
            atendidas = 0
        try:
            chat = float(chat)
        except Exception:
            chat = 0
        try:
            ro = float(ro)
        except Exception:
            ro = 0

        if atendidas == 0 and chat == 0 and ro == 0:
            for indice_coluna in colunas_zeradas:
                ws.cell(row=row, column=indice_coluna).value = 0


def recalcular_colunas_derivadas(df):
    coluna_2min = " > 2min" if " > 2min" in df.columns else "> 2min"
    for coluna in [coluna_2min, "RO", "CHAT"]:
        if coluna in df.columns:
            df[coluna] = pd.to_numeric(df[coluna], errors="coerce").fillna(0)

    if "Atendidas" in df.columns:
        df["Atendidas"] = pd.to_numeric(df["Atendidas"], errors="coerce").fillna(0)
    if COLUNA_ABANDONADAS not in df.columns:
        df[COLUNA_ABANDONADAS] = 0
    df[COLUNA_ABANDONADAS] = pd.to_numeric(
        df[COLUNA_ABANDONADAS],
        errors="coerce",
    ).fillna(0)
    if "TMA" in df.columns:
        df["TMA"] = pd.to_numeric(df["TMA"], errors="coerce").fillna(0)

    linhas_sem_movimento = (df["Atendidas"] == 0) & (df["CHAT"] == 0) & (df["RO"] == 0)
    colunas_para_zerar = [coluna_2min, "RO", "CHAT"]
    if "TMA" in df.columns:
        colunas_para_zerar.append("TMA")
    for nome_coluna in colunas_para_zerar:
        if nome_coluna in df.columns:
            df.loc[linhas_sem_movimento, nome_coluna] = 0

    df["Realizado"] = df[coluna_2min] + df["RO"] + df["CHAT"]
    meta_por_linha = df["Nível"].map(meta_esperada_nivel).fillna(0)
    tecnicos_normalizados = df["Técnico"].apply(normalizar_nome)
    atendidas_validas = df["Atendidas"].where(
        ~tecnicos_normalizados.isin(TECNICOS_DESCONSIDERADOS_ESPERADO),
        0,
    )
    total_meta_mes = total_meta_mensal_por_linha(df)
    total_atendidas_dia = df["Atendidas"].groupby(df["Data"]).transform("sum")
    total_abandonadas_dia = df[COLUNA_ABANDONADAS].groupby(df["Data"]).transform("max")
    total_base_fila_dia = total_atendidas_dia + total_abandonadas_dia
    proporcional = total_base_fila_dia * (
        meta_por_linha / total_meta_mes.replace(0, pd.NA)
    )
    proporcional = proporcional.fillna(0)
    df["Esperado"] = proporcional.clip(upper=meta_por_linha).map(
        arredondar_esperado
    )
    df.loc[linhas_sem_movimento, "Esperado"] = 0
    linhas_em_afastamento = df.apply(
        lambda linha: (
            not pd.isna(linha["Data"])
            and (
                ferias_ativa_tecnico(
                    linha["Técnico"],
                    pd.Timestamp(linha["Data"]).date(),
                )
                is not None
                or licenca_ativa_tecnico(
                    linha["Técnico"],
                    pd.Timestamp(linha["Data"]).date(),
                )
                is not None
            )
        ),
        axis=1,
    )
    df.loc[linhas_em_afastamento, "Esperado"] = 0
    for indice, linha in df.iterrows():
        if pd.isna(linha["Data"]):
            continue

        ausencia = ausencia_programada_tecnico(
            linha["Técnico"],
            pd.Timestamp(linha["Data"]).date(),
        )
        if ausencia is None:
            continue

        _, minutos_ausentes = ausencia
        if minutos_ausentes is None:
            df.at[indice, "Esperado"] = 0
            continue

        proporcao_trabalhada = max(
            1 - (minutos_ausentes / CARGA_DIARIA_PADRAO_MINUTOS),
            0,
        )
        df.at[indice, "Esperado"] = arredondar_esperado(
            df.at[indice, "Esperado"] * proporcao_trabalhada
        )
    df["Desvio"] = df["Realizado"] - df["Esperado"]
    df["Classificação"] = df["Desvio"].apply(status_por_desvio)
    return df


def ultima_data_com_valor(dados, coluna, permitir_zero=False):
    if coluna not in dados.columns or dados.empty:
        return pd.NaT

    serie = pd.to_numeric(dados[coluna], errors="coerce")
    if permitir_zero:
        filtro = serie.notna()
    else:
        filtro = serie.fillna(0) > 0

    dados_validos = dados.loc[filtro]
    if dados_validos.empty:
        return pd.NaT
    return dados_validos["Data"].max()


def precisa_atualizar_colunas(dataframe, data_referencia, colunas_verificacao):
    dados_dia = dataframe[dataframe["Data"] == pd.Timestamp(data_referencia)]
    if dados_dia.empty:
        return True

    for nome_coluna in colunas_verificacao:
        if nome_coluna not in dados_dia.columns:
            return True
        if dados_dia[nome_coluna].isna().any():
            return True

    return False


def atualizar_dados_automaticamente():
    if st.session_state.get("auto_refresh_executado"):
        return

    st.session_state["auto_refresh_executado"] = True
    data_referencia = data_dia_anterior()
    env_local = carregar_env_local()
    usuario_sgd = env_local.get("SGD_USUARIO", os.getenv("SGD_USUARIO", ""))
    senha_sgd = env_local.get("SGD_SENHA", os.getenv("SGD_SENHA", ""))

    try:
        df_atual = ler_dataframe_produtividade()
        df_atual.columns = df_atual.columns.str.strip()
        df_atual["Data"] = pd.to_datetime(df_atual["Data"], dayfirst=True, errors="coerce")
        df_atual = df_atual.dropna(subset=["Data"])
    except Exception:
        return

    try:
        if precisa_atualizar_colunas(
            df_atual,
            data_referencia,
            ["Atendidas", "> 2min" if "> 2min" in df_atual.columns else " > 2min", "TMA"],
        ):
            atualizar_planilha_com_bi(data_referencia)
    except Exception:
        pass

    try:
        if precisa_atualizar_colunas(df_atual, data_referencia, ["RO"]):
            atualizar_planilha_com_ro(data_referencia)
    except Exception:
        pass

    try:
        if usuario_sgd and senha_sgd and precisa_atualizar_colunas(
            df_atual,
            data_referencia,
            ["SSC", "Satisfação", "Votação"],
        ):
            atualizar_planilha_com_sgd(data_referencia, usuario_sgd, senha_sgd)
    except Exception:
        pass


try:
    locale.setlocale(locale.LC_TIME, "pt_BR.UTF-8")
except Exception:
    pass

st.set_page_config(
    page_title="Painel de Produtividade",
    layout="wide",
    initial_sidebar_state="expanded",
)
atualizar_dados_automaticamente()

st.markdown(
    f"""
<style>
.main {{
    background-color: {FUNDO};
}}
[data-testid="stMetric"] {{
    background-color: {COR_BRANCO};
    border: 1px solid #E5E7EB;
    padding: 15px;
    border-radius: 12px;
    box-shadow: 0px 2px 5px rgba(0,0,0,0.05);
}}
[data-testid="stRadio"] [role="radiogroup"] {{
    display:flex;
    gap:4px;
    border-bottom:1px solid #E5E7EB;
    margin-bottom:18px;
}}
[data-testid="stRadio"] [role="radiogroup"] label {{
    padding:9px 12px;
    margin:0;
    border-bottom:2px solid transparent;
}}
[data-testid="stRadio"] [role="radiogroup"] label > div:first-child {{
    display:none;
}}
[data-testid="stRadio"] [role="radiogroup"] label:has(input:checked) {{
    border-bottom-color:#F97316;
    color:#111827;
    font-weight:700;
}}
.gestao-kpi-grid {{
    display:grid;
    grid-template-columns:repeat(4,minmax(0,1fr));
    gap:16px;
    margin-bottom:18px;
}}
.gestao-kpi-card {{
    background:#FFFFFF;
    border:1px solid #D9E1EC;
    border-radius:12px;
    min-height:112px;
    padding:16px 18px;
    box-shadow:0 8px 24px rgba(15, 23, 42, 0.04);
    display:flex;
    justify-content:space-between;
    align-items:flex-end;
    gap:12px;
}}
.gestao-kpi-card small {{
    display:block;
    color:#374151;
    font-size:12px;
    margin-bottom:8px;
}}
.gestao-kpi-card strong {{
    color:#111827;
    font-size:26px;
    line-height:1;
    font-weight:700;
}}
.gestao-kpi-card.negative strong,
.gestao-kpi-card.negative .gestao-kpi-extra {{
    color:#DC2626;
}}
.gestao-kpi-extra {{
    color:#2563EB;
    font-size:12px;
    font-weight:600;
    text-align:right;
}}
.gestao-kpi-spark {{
    color:#16A34A;
    font-size:36px;
    line-height:1;
    letter-spacing:0;
}}
.gestao-kpi-spark.blue {{
    color:#2563EB;
}}
.gestao-kpi-spark.slate {{
    color:#475569;
}}
.gestao-card {{
    background:#FFFFFF;
    border:1px solid #D9E1EC;
    border-radius:14px;
    padding:18px;
    box-shadow:0 8px 24px rgba(15, 23, 42, 0.04);
}}
.gestao-card h4 {{
    color:#111827;
    font-size:16px;
    font-weight:700;
    margin:0 0 14px 0;
}}
.gestao-grid-duplo {{
    display:grid;
    grid-template-columns:1fr 1fr;
    gap:16px;
    margin-bottom:18px;
}}
.gestao-status-grid {{
    display:grid;
    grid-template-columns:repeat(4,minmax(0,1fr));
    gap:12px;
}}
.gestao-status-resumo {{
    background:#FFFFFF;
    border:1px solid #D9E1EC;
    border-top:4px solid #6B7280;
    border-radius:10px;
    padding:14px 14px 16px;
}}
.gestao-status-resumo span,
.gestao-status-resumo small,
.gestao-status-resumo strong {{
    display:block;
}}
.gestao-status-resumo span {{
    color:#111827;
    font-size:12px;
    font-weight:700;
    margin-bottom:10px;
}}
.gestao-status-resumo strong {{
    color:#111827;
    font-size:22px;
    line-height:1;
    margin-bottom:10px;
}}
.gestao-status-resumo em {{
    color:#374151;
    display:block;
    font-size:11px;
    font-style:normal;
    margin-bottom:10px;
}}
.gestao-status-resumo small {{
    font-size:10px;
    line-height:1.35;
}}
.gestao-absorcao {{
    display:flex;
    flex-direction:column;
    gap:14px;
    padding-top:4px;
}}
.gestao-absorcao-linha {{
    display:grid;
    grid-template-columns:86px minmax(0,1fr) 48px;
    align-items:center;
    gap:10px;
}}
.gestao-absorcao-linha span {{
    color:#111827;
    font-size:12px;
}}
.gestao-absorcao-linha strong {{
    color:#111827;
    font-size:12px;
    text-align:right;
}}
.gestao-absorcao-trilho {{
    height:16px;
    background:
        linear-gradient(to right, transparent 0, transparent calc(33.33% - 1px), #E5E7EB calc(33.33% - 1px), #E5E7EB calc(33.33% + 1px), transparent calc(33.33% + 1px), transparent calc(66.66% - 1px), #E5E7EB calc(66.66% - 1px), #E5E7EB calc(66.66% + 1px), transparent calc(66.66% + 1px)),
        #F8FAFC;
    border-radius:999px;
    overflow:hidden;
    border:1px solid #E5E7EB;
}}
.gestao-absorcao-trilho i {{
    display:block;
    height:100%;
    background:linear-gradient(90deg, #16A34A 0%, #16A34A 100%);
    border-radius:999px;
}}
.gestao-linha-inferior {{
    display:grid;
    grid-template-columns:minmax(0,2fr) minmax(320px,1fr);
    gap:16px;
    margin-bottom:18px;
}}
.gestao-alertas {{
    display:flex;
    flex-direction:column;
    gap:10px;
}}
.gestao-alerta {{
    display:flex;
    align-items:center;
    justify-content:space-between;
    gap:10px;
    background:#FFFFFF;
    border:1px solid #E5E7EB;
    border-radius:10px;
    padding:12px 14px;
    color:#111827;
    font-size:12px;
}}
.gestao-alerta-label {{
    display:flex;
    align-items:center;
    gap:10px;
}}
.gestao-alerta-icone {{
    width:24px;
    height:24px;
    border-radius:999px;
    display:flex;
    align-items:center;
    justify-content:center;
    color:#FFFFFF;
    font-size:13px;
    font-weight:700;
}}
.gestao-alerta-seta {{
    color:#6B7280;
    font-size:16px;
}}
.gestao-resumo-faixa {{
    display:grid;
    grid-template-columns:1fr 1fr;
    gap:16px;
    margin-bottom:18px;
}}
.gestao-resumo-card {{
    background:#FFFFFF;
    border:1px solid #D9E1EC;
    border-radius:14px;
    padding:16px 18px;
    box-shadow:0 8px 24px rgba(15, 23, 42, 0.04);
}}
.gestao-resumo-card h4 {{
    color:#111827;
    font-size:16px;
    font-weight:700;
    margin:0 0 14px 0;
}}
.gestao-resumo-grid {{
    display:grid;
    grid-template-columns:repeat(4,minmax(0,1fr));
    gap:14px;
}}
.gestao-resumo-item {{
    display:flex;
    align-items:flex-start;
    gap:10px;
}}
.gestao-resumo-bola {{
    width:22px;
    height:22px;
    border-radius:999px;
    display:flex;
    align-items:center;
    justify-content:center;
    font-size:12px;
    font-weight:700;
    flex-shrink:0;
}}
.gestao-resumo-texto strong {{
    color:#111827;
    display:block;
    font-size:14px;
    line-height:1.1;
}}
.gestao-resumo-texto span {{
    color:#374151;
    display:block;
    font-size:11px;
    margin-top:3px;
}}
.card-tecnico {{
    padding:12px;
    border-radius:10px;
    margin-bottom:10px;
    color:white;
    font-weight:bold;
    font-size:16px;
}}
.resultado-cabecalho {{
    display:flex;
    align-items:center;
    justify-content:space-between;
    gap:16px;
    margin:4px 0 14px;
}}
.resultado-identidade {{
    min-width:0;
}}
.resultado-nome {{
    color:#111827;
    font-size:24px;
    font-weight:700;
    line-height:1.2;
}}
.resultado-nivel {{
    color:#6B7280;
    font-size:13px;
    margin-top:4px;
}}
.resultado-status {{
    color:#FFFFFF;
    border-radius:6px;
    padding:6px 10px;
    font-size:12px;
    font-weight:700;
    white-space:nowrap;
}}
.resultado-grid-principal {{
    display:grid;
    grid-template-columns:repeat(3,minmax(0,1fr));
    gap:12px;
}}
.resultado-grid-secundario {{
    display:grid;
    grid-template-columns:repeat(5,minmax(0,1fr));
    gap:10px;
    margin-top:12px;
}}
.resultado-card {{
    background:#FFFFFF;
    border:1px solid #E5E7EB;
    border-radius:8px;
    padding:14px 16px;
    min-width:0;
}}
.resultado-card.destaque {{
    min-height:92px;
}}
.resultado-label {{
    color:#6B7280;
    font-size:12px;
    line-height:1.3;
}}
.resultado-valor {{
    color:#111827;
    font-size:28px;
    font-weight:700;
    line-height:1.15;
    margin-top:8px;
}}
.resultado-card.secundario .resultado-valor {{
    font-size:21px;
    margin-top:6px;
}}
.resultado-progresso-bloco {{
    margin:14px 0 2px;
}}
.resultado-progresso-texto {{
    display:flex;
    justify-content:space-between;
    color:#4B5563;
    font-size:12px;
    margin-bottom:6px;
}}
.resultado-progresso-trilho {{
    width:100%;
    height:8px;
    background:#E5E7EB;
    border-radius:4px;
    overflow:hidden;
}}
.resultado-progresso-barra {{
    height:100%;
    border-radius:4px;
}}
.resultado-faixa {{
    display:grid;
    grid-template-columns:repeat(2,minmax(0,1fr));
    gap:10px;
    margin-top:12px;
}}
.produtividade-resumo {{
    display:flex;
    align-items:center;
    gap:24px;
    flex-wrap:wrap;
    margin:4px 0 8px;
    color:#4B5563;
    font-size:13px;
}}
.produtividade-resumo-item {{
    display:flex;
    align-items:center;
    gap:7px;
}}
.produtividade-resumo-ponto {{
    width:9px;
    height:9px;
    border-radius:50%;
    flex:0 0 9px;
}}
.disp-resumo {{
    display:grid;
    grid-template-columns:repeat(4,minmax(0,1fr));
    gap:10px;
    margin:14px 0 18px;
}}
.disp-resumo > div {{
    background:#FFFFFF;
    border:1px solid #E5E7EB;
    border-radius:7px;
    padding:12px 14px;
}}
.disp-resumo span {{
    display:block;
    color:#4B5563;
    font-size:12px;
}}
.disp-resumo strong {{
    display:block;
    color:#111827;
    font-size:24px;
    margin-top:4px;
}}
.disp-legenda {{
    display:flex;
    align-items:center;
    flex-wrap:wrap;
    gap:16px;
    color:#4B5563;
    font-size:11px;
    margin:-6px 0 14px;
}}
.disp-legenda span {{
    display:inline-flex;
    align-items:center;
    gap:6px;
}}
.disp-legenda i {{
    width:9px;
    height:9px;
    border-radius:2px;
    display:inline-block;
}}
.disp-calendario {{
    border:1px solid #E5E7EB;
    border-radius:7px;
    overflow-x:auto;
    overflow-y:hidden;
    background:#FFFFFF;
}}
.disp-semana-cabecalho,
.disp-semana {{
    display:grid;
    grid-template-columns:repeat(7,minmax(0,1fr));
}}
.disp-dia-semana {{
    background:#F9FAFB;
    border-bottom:1px solid #E5E7EB;
    color:#4B5563;
    font-size:11px;
    font-weight:700;
    padding:8px;
    text-align:center;
}}
.disp-dia {{
    min-height:116px;
    border-right:1px solid #E5E7EB;
    border-bottom:1px solid #E5E7EB;
    padding:6px;
    min-width:0;
}}
.disp-semana .disp-dia:last-child {{
    border-right:0;
}}
.disp-dia.fora-mes {{
    background:#FAFAFA;
    color:#9CA3AF;
}}
.disp-dia.fim-semana {{
    background:#F9FAFB;
}}
.disp-dia.hoje {{
    box-shadow:inset 0 0 0 2px #F97316;
}}
.disp-numero-dia {{
    color:#374151;
    font-size:11px;
    font-weight:700;
    margin-bottom:5px;
}}
.disp-evento {{
    background:#F9FAFB;
    border-left:3px solid #6B7280;
    border-radius:4px;
    margin-bottom:4px;
    padding:4px 5px;
    overflow:hidden;
}}
.disp-evento span,
.disp-evento small {{
    display:block;
    overflow:hidden;
    text-overflow:ellipsis;
    white-space:nowrap;
}}
.disp-evento span {{
    color:#111827;
    font-size:10px;
    font-weight:700;
}}
.disp-evento small {{
    color:#6B7280;
    font-size:9px;
    margin-top:1px;
}}
.disp-mais {{
    color:#6B7280;
    font-size:9px;
    padding:2px;
}}
.disp-proximo {{
    background:#FFFFFF;
    border:1px solid #E5E7EB;
    border-left:4px solid #6B7280;
    border-radius:6px;
    margin-bottom:7px;
    padding:8px 9px;
}}
.disp-proximo strong,
.disp-proximo span,
.disp-proximo small {{
    display:block;
}}
.disp-proximo strong {{
    color:#111827;
    font-size:11px;
}}
.disp-proximo span {{
    color:#4B5563;
    font-size:10px;
    margin-top:2px;
}}
.disp-proximo small {{
    color:#6B7280;
    font-size:9px;
    margin-top:3px;
}}
.login-marca {{
    position:fixed;
    top:0;
    left:0;
    width:10px;
    height:100vh;
    background:#166534;
    z-index:9999;
}}
.login-marca::after {{
    content:"";
    position:absolute;
    top:0;
    left:10px;
    width:4px;
    height:120px;
    background:#F97316;
}}
.login-topo {{
    color:#374151;
    font-size:12px;
    text-align:right;
    margin:4px 2px 54px 0;
}}
.login-cabecalho {{
    text-align:center;
    margin-bottom:22px;
}}
.login-simbolo {{
    display:inline-grid;
    grid-template-columns:repeat(3,7px);
    align-items:end;
    gap:3px;
    height:28px;
    margin-bottom:14px;
}}
.login-simbolo span {{
    display:block;
    border-radius:2px 2px 0 0;
}}
.login-titulo {{
    color:#111827;
    font-size:30px;
    font-weight:700;
    line-height:1.2;
}}
.login-subtitulo {{
    color:#374151;
    font-size:14px;
    margin-top:8px;
}}
.login-indicadores {{
    display:grid;
    grid-template-columns:repeat(3,1fr);
    margin-top:24px;
    padding-top:18px;
    border-top:1px solid #E5E7EB;
}}
.login-indicador {{
    color:#374151;
    font-size:12px;
    text-align:center;
    padding:0 8px;
}}
.login-indicador + .login-indicador {{
    border-left:1px solid #E5E7EB;
}}
.login-indicador span {{
    display:block;
    width:8px;
    height:8px;
    border-radius:50%;
    margin:0 auto 7px;
}}
.login-atualizacao {{
    color:#374151;
    font-size:11px;
    text-align:center;
    margin-top:22px;
}}
@media (max-width: 900px) {{
    .gestao-kpi-grid,
    .gestao-grid-duplo,
    .gestao-linha-inferior,
    .gestao-resumo-faixa {{
        grid-template-columns:1fr;
    }}
    .gestao-status-grid,
    .gestao-resumo-grid {{
        grid-template-columns:repeat(2,minmax(0,1fr));
    }}
    .gestao-absorcao-linha {{
        grid-template-columns:76px minmax(0,1fr) 42px;
    }}
    .resultado-grid-principal {{
        grid-template-columns:1fr;
    }}
    .resultado-grid-secundario {{
        grid-template-columns:repeat(2,minmax(0,1fr));
    }}
    .resultado-faixa {{
        grid-template-columns:1fr;
    }}
    .disp-resumo {{
        grid-template-columns:repeat(2,minmax(0,1fr));
    }}
    .disp-dia {{
        min-height:82px;
        padding:3px;
    }}
    .disp-semana-cabecalho,
    .disp-semana {{
        min-width:760px;
    }}
    .disp-evento small {{
        display:none;
    }}
    .resultado-nome {{
        font-size:20px;
    }}
}}
@media (max-width: 640px) {{
    .gestao-kpi-grid,
    .gestao-status-grid,
    .gestao-resumo-grid,
    .gestao-resumo-faixa {{
        grid-template-columns:1fr;
    }}
}}
</style>
""",
    unsafe_allow_html=True,
)

df = ler_dataframe_produtividade()
usuarios = ler_dataframe_usuarios()

df.columns = df.columns.str.strip()
usuarios.columns = usuarios.columns.str.strip().str.lower()

df["Técnico"] = df["Técnico"].astype(str).str.lower().str.strip()
usuarios["usuario"] = usuarios["usuario"].astype(str).str.lower().str.strip()
usuarios["senha"] = (
    usuarios["senha"]
    .astype(str)
    .str.replace(".0", "", regex=False)
    .str.strip()
)
usuarios["tecnico"] = usuarios["tecnico"].astype(str).str.lower().str.strip()

df["Data"] = pd.to_datetime(df["Data"], dayfirst=True, errors="coerce")
df = df.dropna(subset=["Data"])
df["Data Formatada"] = df["Data"].dt.strftime("%d/%m/%Y")
for coluna_numerica in ["SSC", "Satisfação", "Votação"]:
    if coluna_numerica in df.columns:
        df[coluna_numerica] = pd.to_numeric(
            df[coluna_numerica], errors="coerce"
        ).fillna(0)
df = recalcular_colunas_derivadas(df)
env_local = carregar_env_local()
for chave, valor_padrao in {
    "auth_ok": False,
    "auth_modo_gestao": False,
    "auth_usuario": "",
    "auth_tecnico": "",
}.items():
    if chave not in st.session_state:
        st.session_state[chave] = valor_padrao

if not st.session_state["auth_ok"]:
    st.markdown(
        """
        <style>
        [data-testid="stSidebar"],
        [data-testid="collapsedControl"] {
            display:none;
        }
        [data-testid="stAppViewContainer"] > .main {
            background:#FFFFFF;
        }
        [data-testid="stForm"] {
            border:0;
            padding:0;
        }
        [data-testid="stFormSubmitButton"] button {
            width:100%;
            background:#F97316;
            color:#FFFFFF;
            border:1px solid #F97316;
            border-radius:7px;
            font-weight:700;
            min-height:42px;
        }
        [data-testid="stFormSubmitButton"] button:hover {
            background:#EA580C;
            border-color:#EA580C;
            color:#FFFFFF;
        }
        .login-autoria {
            margin-top:10px;
            color:#374151;
            font-size:12px;
            font-weight:700;
            text-align:center;
        }
        </style>
        <div class="login-marca"></div>
        <div class="login-topo">Soft News · Folha</div>
        """,
        unsafe_allow_html=True,
    )

    data_atualizacao_login = df["Data"].max()
    data_atualizacao_texto = (
        data_atualizacao_login.strftime("%d/%m/%Y")
        if not pd.isna(data_atualizacao_login)
        else "não disponível"
    )

    espaco_esquerdo, coluna_login, espaco_direito = st.columns([1, 1.05, 1])
    with coluna_login:
        st.markdown(
            """
            <div class="login-cabecalho">
                <div class="login-simbolo" aria-hidden="true">
                    <span style="height:13px;background:#2563EB;"></span>
                    <span style="height:21px;background:#F97316;"></span>
                    <span style="height:28px;background:#16A34A;"></span>
                </div>
                <div class="login-titulo">Painel de Produtividade</div>
                <div class="login-subtitulo">
                    Acompanhe seus resultados e indicadores
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        with st.form("form_login"):
            usuario_input = st.text_input(
                "Usuário",
                key="login_usuario_input",
                placeholder="Informe seu usuário",
            )
            senha_input = st.text_input(
                "Senha",
                type="password",
                key="login_senha_input",
                placeholder="Informe sua senha",
            )
            entrar = st.form_submit_button("Entrar")

        st.markdown(
            f"""
            <div class="login-indicadores">
                <div class="login-indicador">
                    <span style="background:#16A34A;"></span>
                    Produtividade
                </div>
                <div class="login-indicador">
                    <span style="background:#F97316;"></span>
                    Qualidade
                </div>
                <div class="login-indicador">
                    <span style="background:#2563EB;"></span>
                    Evolução
                </div>
            </div>
            <div class="login-atualizacao">
                Dados atualizados em {data_atualizacao_texto}
            </div>
            <div class="login-autoria">
                Criado por Esther Alves Queiroz
            </div>
            """,
            unsafe_allow_html=True,
        )

    if entrar:
        usuario_digitado = builtins.str(usuario_input).lower().strip()
        senha_digitada = builtins.str(senha_input).replace(".0", "").strip()
        login_usuario = usuarios[
            (usuarios["usuario"] == usuario_digitado)
            & (usuarios["senha"] == senha_digitada)
        ]
        usuario_apoio_valido = (
            usuario_digitado in USUARIOS_APOIO and not login_usuario.empty
        )

        if usuario_digitado == "gestao" and senha_digitada == "30071997":
            st.session_state["auth_ok"] = True
            st.session_state["auth_modo_gestao"] = True
            st.session_state["auth_usuario"] = "gestao"
            st.session_state["auth_tecnico"] = ""
            st.rerun()
        elif usuario_apoio_valido:
            st.session_state["auth_ok"] = True
            st.session_state["auth_modo_gestao"] = True
            st.session_state["auth_usuario"] = usuario_digitado
            st.session_state["auth_tecnico"] = ""
            st.rerun()
        elif not login_usuario.empty:
            st.session_state["auth_ok"] = True
            st.session_state["auth_modo_gestao"] = False
            st.session_state["auth_usuario"] = usuario_digitado
            st.session_state["auth_tecnico"] = builtins.str(
                login_usuario.iloc[0]["tecnico"]
            ).lower().strip()
            st.rerun()
        else:
            st.error("Usuário ou senha inválidos.")

    st.stop()

modo_gestao = st.session_state["auth_modo_gestao"]
usuario_digitado = st.session_state["auth_usuario"]
tecnico = st.session_state["auth_tecnico"]

st.title("Painel de Produtividade")
st.sidebar.title("Painel")

visao_gestao = None
if modo_gestao:
    visao_gestao = st.radio(
        "Navegação da gestão",
        [
            "Visão Geral",
            "Resultados Individuais",
            "Disponibilidade",
            "Lista de Apoio",
            "Atualizações",
        ],
        index=0,
        horizontal=True,
        label_visibility="collapsed",
        key="visao_gestao",
    )

if st.sidebar.button("Sair", key="logout_painel"):
    for chave, valor_padrao in {
        "auth_ok": False,
        "auth_modo_gestao": False,
        "auth_usuario": "",
        "auth_tecnico": "",
        "login_usuario_input": "",
        "login_senha_input": "",
    }.items():
        st.session_state[chave] = valor_padrao
    st.rerun()

if modo_gestao:
    if usuario_digitado in USUARIOS_APOIO:
        st.sidebar.success(f"Bem-vinda {usuario_digitado.title()}")
    else:
        st.sidebar.success("Bem-vinda Gestão")

    st.sidebar.caption(
        f"Dados atualizados em {df['Data'].max().strftime('%d/%m/%Y')}"
    )

if modo_gestao and visao_gestao == "Atualizações":
    st.subheader("Atualizações de Dados")
    st.caption(
        "Atualize as fontes do painel sem misturar os controles com os resultados."
    )
    data_referencia = data_dia_anterior()
    env_local = carregar_env_local()
    usuario_sgd_padrao = env_local.get("SGD_USUARIO", os.getenv("SGD_USUARIO", ""))
    senha_sgd_padrao = env_local.get("SGD_SENHA", os.getenv("SGD_SENHA", ""))

    st.divider()
    st.caption(f"Data de referência: {data_referencia.strftime('%d/%m/%Y')}")
    chat_local_disponivel = servico_local_disponivel()

    col_bi, col_ro, col_chat = st.columns(3)
    with col_bi:
        atualizar_bi = st.button(
            "Atualizar BI do dia anterior",
            use_container_width=True,
        )
    with col_ro:
        atualizar_ro = st.button(
            "Atualizar RO do dia anterior",
            use_container_width=True,
        )
    with col_chat:
        atualizar_chat = st.button(
            "Atualizar CHAT do dia anterior",
            disabled=not chat_local_disponivel,
            use_container_width=True,
        )

    if atualizar_bi:
        with st.spinner("Buscando dados no PowerBI e atualizando a planilha..."):
            try:
                origem_atualizacao = "painel"
                try:
                    resultado = atualizar_via_servico_local("bi", data_referencia)["bi"]
                    origem_atualizacao = "planilha local"
                except Exception:
                    resultado = atualizar_planilha_com_bi(data_referencia)
            except PermissionError:
                st.error("Feche a produtividade.xlsx no Excel e tente novamente.")
            except Exception as erro:
                st.error(f"Não foi possível atualizar o BI: {erro}")
            else:
                st.success(
                    f"{resultado['atualizados']} técnicos atualizados em {resultado['data']}."
                )
                st.caption(f"Origem da gravação: {origem_atualizacao}.")
                if resultado["linhas_criadas"]:
                    st.info(
                        f"{resultado['linhas_criadas']} linhas foram criadas para essa data."
                    )
                if resultado["sem_alias"]:
                    st.warning(
                        "Revise a coluna Agente BI da aba Aliases para: "
                        + ", ".join(resultado["sem_alias"][:8])
                    )
                st.cache_data.clear()
                st.rerun()

    if atualizar_ro:
        with st.spinner("Buscando dados do RO e atualizando a planilha..."):
            try:
                origem_atualizacao = "painel"
                try:
                    resultado = atualizar_via_servico_local("ro", data_referencia)["ro"]
                    origem_atualizacao = "planilha local"
                except Exception:
                    resultado = atualizar_planilha_com_ro(data_referencia)
            except PermissionError:
                st.error("Feche a produtividade.xlsx no Excel e tente novamente.")
            except Exception as erro:
                st.error(f"Não foi possível atualizar o RO: {erro}")
            else:
                st.success(
                    f"{resultado['atualizados']} técnicos atualizados em {resultado['data']}."
                )
                st.caption(f"Origem da gravação: {origem_atualizacao}.")
                st.info(
                    f"Origem: {resultado['arquivo']}. Total de RO válidos: {resultado['fonte']}."
                )
                if resultado["linhas_criadas"]:
                    st.info(
                        f"{resultado['linhas_criadas']} linhas foram criadas para essa data."
                    )
                st.cache_data.clear()
                st.rerun()

    if not chat_local_disponivel:
        st.info(
            "O CHAT só pode ser atualizado no painel local "
            "(http://localhost:8501) com o serviço local ativo, "
            "porque ele depende do navegador desta máquina logado no PLUG."
        )

    if atualizar_chat:
        with st.spinner("Buscando chats fechados no PLUG e atualizando a planilha..."):
            try:
                resultado = atualizar_via_servico_local("chat", data_referencia)["chat"]
                origem_atualizacao = "planilha local"
            except PermissionError:
                st.error("Feche a produtividade.xlsx no Excel e tente novamente.")
            except Exception as erro:
                st.error(f"Não foi possível atualizar o CHAT: {erro}")
            else:
                st.success(
                    f"{resultado['atualizados']} técnicos atualizados em {resultado['data']}."
                )
                st.caption(f"Origem da gravação: {origem_atualizacao}.")
                if resultado["linhas_criadas"]:
                    st.info(
                        f"{resultado['linhas_criadas']} linhas foram criadas para essa data."
                    )
                if resultado["sem_alias"]:
                    st.warning(
                        "Revise a coluna Agente Chat da aba Aliases para: "
                        + ", ".join(resultado["sem_alias"][:8])
                    )
                st.cache_data.clear()
                st.rerun()

    st.divider()
    st.markdown("#### SGD")
    col_usuario_sgd, col_senha_sgd, col_botao_sgd = st.columns([1, 1, 0.8])
    with col_usuario_sgd:
        usuario_sgd = st.text_input("Usuário SGD", value=usuario_sgd_padrao)
    with col_senha_sgd:
        senha_sgd = st.text_input(
            "Senha SGD",
            value=senha_sgd_padrao,
            type="password",
        )
    with col_botao_sgd:
        st.write("")
        atualizar_sgd = st.button(
            "Atualizar SGD do mês",
            use_container_width=True,
        )

    if atualizar_sgd:
        if not usuario_sgd or not senha_sgd:
            st.error("Informe usuário e senha do SGD.")
        else:
            with st.spinner("Gerando relatório no SGD e atualizando a planilha..."):
                try:
                    origem_atualizacao = "painel"
                    try:
                        resultado = atualizar_via_servico_local("sgd", data_referencia)["sgd"]
                        origem_atualizacao = "planilha local"
                    except Exception:
                        resultado = atualizar_planilha_com_sgd(
                            data_referencia,
                            usuario_sgd,
                            senha_sgd,
                        )
                except PermissionError:
                    st.error("Feche a produtividade.xlsx no Excel e tente novamente.")
                except Exception as erro:
                    st.error(f"Não foi possível atualizar o SGD: {erro}")
                else:
                    st.success(
                        f"{resultado['atualizados']} registros atualizados em "
                        f"{resultado['dias_processados']} dias. "
                        f"Período: {resultado['periodo']}."
                    )
                    st.caption(f"Origem da gravação: {origem_atualizacao}.")
                    if resultado["linhas_criadas"]:
                        st.info(
                            f"{resultado['linhas_criadas']} linhas foram criadas para essa data."
                        )
                    if resultado["sem_alias"]:
                        st.warning(
                            "Revise a coluna Agente SGD da aba Aliases para: "
                            + ", ".join(resultado["sem_alias"][:8])
                        )
                    st.cache_data.clear()
                    st.rerun()

if not modo_gestao:
    st.sidebar.success(f"Bem-vindo(a), {tecnico.title()}")

if modo_gestao and visao_gestao == "Atualizações":
    st.stop()

if modo_gestao:
    tecnicos_disponiveis = sorted(df["Técnico"].unique())
    tecnico = st.session_state.get("tecnico_gestao", tecnicos_disponiveis[0])
    if tecnico not in tecnicos_disponiveis:
        tecnico = tecnicos_disponiveis[0]

dados_tecnico = df[df["Técnico"] == tecnico]

if dados_tecnico.empty:
    st.error("Nenhum dado encontrado.")
    st.stop()

mes_atual = df["Data"].dt.month.max()
ano_atual = df["Data"].dt.year.max()

dados_mes_atual = dados_tecnico[
    (dados_tecnico["Data"].dt.month == mes_atual)
    & (dados_tecnico["Data"].dt.year == ano_atual)
]

ultima_data_mes = dados_mes_atual["Data"].max()
ultima_data_produtividade = ultima_data_com_valor(dados_mes_atual, "Realizado")
if pd.isna(ultima_data_produtividade):
    ultima_data_produtividade = ultima_data_mes
dados_ultimo_dia = dados_mes_atual[dados_mes_atual["Data"] == ultima_data_produtividade]

ultima_data_satisfacao = ultima_data_com_valor(dados_mes_atual, "Satisfação")
if pd.isna(ultima_data_satisfacao):
    ultima_data_satisfacao = ultima_data_mes
dados_ultimo_dia_satisfacao = dados_mes_atual[
    dados_mes_atual["Data"] == ultima_data_satisfacao
]

ultima_data_votacao = ultima_data_com_valor(dados_mes_atual, "Votação")
if pd.isna(ultima_data_votacao):
    ultima_data_votacao = ultima_data_mes
dados_ultimo_dia_votacao = dados_mes_atual[
    dados_mes_atual["Data"] == ultima_data_votacao
]


def montar_grafico_ranking(ranking, titulo, eixo_x, meta_referencia=None):
    def abreviar_nome_ranking(nome):
        partes = builtins.str(nome or "").title().split()
        if len(partes) <= 2:
            return " ".join(partes)
        return f"{partes[0]} {partes[-1]}"

    ranking = (
        ranking.copy()
        .sort_values(by="Realizado", ascending=False)
        .reset_index(drop=True)
    )
    ranking["Posição"] = ranking.index + 1
    ranking["Técnico Completo"] = ranking["Técnico"].str.title()
    ranking["Técnico Exibição"] = ranking["Técnico"].map(abreviar_nome_ranking)
    ranking["Cor"] = ranking["Técnico"].apply(
        lambda nome: "Selecionado" if nome == tecnico else "Mesmo nível"
    )
    ranking["Rótulo"] = ranking.apply(
        lambda linha: (
            f"{int(linha['Posição'])}º · {int(linha['Realizado'])}"
            if linha["Posição"] <= 3
            else f"{int(linha['Realizado'])}"
        ),
        axis=1,
    )

    grafico = px.bar(
        ranking,
        x="Técnico Exibição",
        y="Realizado",
        color="Cor",
        text="Rótulo",
        custom_data=["Técnico Completo", "Posição"],
        labels={
            "Realizado": eixo_x,
            "Técnico Exibição": "Técnico",
            "Cor": "",
        },
        title=titulo,
        color_discrete_map={
            "Selecionado": COR_LARANJA,
            "Mesmo nível": "#9CA3AF",
        },
        category_orders={
            "Técnico Exibição": ranking["Técnico Exibição"].tolist()
        },
    )

    grafico.update_traces(
        textposition="outside",
        textfont_color="#374151",
        textfont_size=12,
        width=0.54,
        cliponaxis=False,
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "Posição: %{customdata[1]}º<br>"
            f"{eixo_x}: %{{y}}<extra></extra>"
        ),
    )
    if meta_referencia is not None and meta_referencia > 0:
        grafico.add_hline(
            y=meta_referencia,
            line_color="#EF4444",
            line_dash="dash",
            line_width=1.4,
            annotation_text=(
                f"Meta do nível: {arredondar_esperado(meta_referencia)}"
            ),
            annotation_position="top left",
            annotation_font_color="#B91C1C",
            annotation_font_size=11,
            annotation_bgcolor="#FFFFFF",
        )
    grafico.update_layout(
        plot_bgcolor=COR_BRANCO,
        paper_bgcolor=COR_BRANCO,
        font_color="#4B5563",
        showlegend=False,
        bargap=0.42,
        height=430,
        margin=dict(l=45, r=20, t=62, b=95),
        xaxis_title="Técnico",
        yaxis_title=eixo_x,
        xaxis=dict(
            type="category",
            showgrid=False,
            tickangle=-25,
            tickfont=dict(size=11, color="#4B5563"),
            linecolor="#E5E7EB",
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor="#F3F4F6",
            gridwidth=1,
            zeroline=False,
            rangemode="tozero",
            tickfont=dict(size=11, color="#6B7280"),
        ),
        title=dict(
            font=dict(size=15, color="#111827"),
            x=0,
        ),
    )
    return grafico

if modo_gestao and visao_gestao == "Lista de Apoio":
    mostrar_lista_apoio_gestao()
    st.stop()

if modo_gestao and visao_gestao == "Disponibilidade":
    mostrar_gestao_disponibilidade(df)
    st.stop()

if modo_gestao and visao_gestao == "Visão Geral":
    def formatar_inteiro_gestao(valor):
        return f"{int(valor):,}".replace(",", ".")

    def formatar_percentual_gestao(valor, casas=2):
        return f"{valor:.{casas}f}%".replace(".", ",")

    periodos_gestao = sorted(
        {
            (data.year, data.month)
            for data in df["Data"].dropna()
        },
        reverse=True,
    )
    opcoes_periodo_gestao = {
        nome_mes_ano(ano_periodo, mes_periodo): (ano_periodo, mes_periodo)
        for ano_periodo, mes_periodo in periodos_gestao
    }
    periodo_padrao_gestao = nome_mes_ano(ano_atual, mes_atual)
    if periodo_padrao_gestao not in opcoes_periodo_gestao:
        periodo_padrao_gestao = next(iter(opcoes_periodo_gestao))

    col_periodo_gestao, _ = st.columns([0.22, 0.78])
    with col_periodo_gestao:
        periodo_gestao_label = st.selectbox(
            "Mês da visão geral",
            list(opcoes_periodo_gestao.keys()),
            index=list(opcoes_periodo_gestao.keys()).index(periodo_padrao_gestao),
            key="gestao_visao_geral_mes",
            label_visibility="collapsed",
        )

    ano_visao_gestao, mes_visao_gestao = opcoes_periodo_gestao[periodo_gestao_label]
    dados_gestao_mes = df[
        (df["Data"].dt.month == mes_visao_gestao)
        & (df["Data"].dt.year == ano_visao_gestao)
    ].copy()

    realizado_gestao = int(dados_gestao_mes["Realizado"].sum())
    esperado_gestao = int(dados_gestao_mes["Esperado"].sum())
    desvio_gestao = realizado_gestao - esperado_gestao
    desvio_percentual_gestao = (
        (desvio_gestao / esperado_gestao) * 100 if esperado_gestao else 0
    )

    data_status = dados_gestao_mes["Data"].max()
    status_atual = (
        dados_gestao_mes[dados_gestao_mes["Data"] == data_status][
            ["Técnico", "Nível", "Classificação"]
        ]
        .dropna(subset=["Técnico", "Classificação"])
        .sort_values(by=["Classificação", "Técnico"])
    )
    tecnicos_ativos_gestao = int(status_atual["Técnico"].nunique())
    percentuais_absorcao_mes = percentuais_absorcao_por_mes(
        df,
        ano_visao_gestao,
        mes_visao_gestao,
    )

    st.markdown(
        f"""
        <div class="gestao-kpi-grid">
            <div class="gestao-kpi-card">
                <div>
                    <small>Realizado total</small>
                    <strong>{formatar_inteiro_gestao(realizado_gestao)}</strong>
                </div>
                <div class="gestao-kpi-spark">▁▃▂▅▄▆▅▇</div>
            </div>
            <div class="gestao-kpi-card">
                <div>
                    <small>Esperado total</small>
                    <strong>{formatar_inteiro_gestao(esperado_gestao)}</strong>
                </div>
                <div class="gestao-kpi-spark blue">▂▃▄▃▅▄▆▇</div>
            </div>
            <div class="gestao-kpi-card {'negative' if desvio_gestao < 0 else ''}">
                <div>
                    <small>Desvio</small>
                    <strong>{desvio_gestao:+d}</strong>
                </div>
                <div class="gestao-kpi-extra">{formatar_percentual_gestao(desvio_percentual_gestao)}<br>vs esperado</div>
            </div>
            <div class="gestao-kpi-card">
                <div>
                    <small>Técnicos ativos</small>
                    <strong>{formatar_inteiro_gestao(tecnicos_ativos_gestao)}</strong>
                </div>
                <div class="gestao-kpi-spark slate">◔</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    descricoes_status = {
        "CRÍTICO": ("Técnicos", "≤ 21% da meta"),
        "ATENÇÃO": ("Técnicos", "> 21% e ≤ 98%"),
        "BOM": ("Técnicos", "> 98% e ≤ 120%"),
        "EXCELENTE": ("Técnicos", "> 120% da meta"),
    }
    cores_texto_status = {
        "CRÍTICO": "#DC2626",
        "ATENÇÃO": "#F97316",
        "BOM": "#2563EB",
        "EXCELENTE": "#15803D",
    }
    blocos_status = []
    for status in ["CRÍTICO", "ATENÇÃO", "BOM", "EXCELENTE"]:
        quantidade = int((status_atual["Classificação"] == status).sum())
        legenda, referencia = descricoes_status[status]
        blocos_status.append(
            '<div class="gestao-status-resumo" '
            f'style="border-top-color:{CORES_STATUS[status]};">'
            f"<span>{status}</span>"
            f"<strong>{quantidade}</strong>"
            f"<em>{legenda}</em>"
            f'<small style="color:{cores_texto_status[status]};">{referencia}</small>'
            "</div>"
        )

    maior_percentual = max(percentuais_absorcao_mes.values(), default=0)
    largura_referencia = maior_percentual if maior_percentual > 0 else 1
    barras_absorcao = []
    for nivel, percentual in percentuais_absorcao_mes.items():
        largura = min((percentual / largura_referencia) * 100, 100)
        barras_absorcao.append(
            '<div class="gestao-absorcao-linha">'
            f"<span>{escape(nivel)}</span>"
            f'<div class="gestao-absorcao-trilho"><i style="width:{largura:.2f}%;"></i></div>'
            f"<strong>{formatar_percentual_gestao(percentual)}</strong>"
            "</div>"
        )

    st.markdown(
        '<div class="gestao-grid-duplo">'
        '<div class="gestao-card"><h4>Distribuição por classificação</h4>'
        '<div class="gestao-status-grid">'
        + "".join(blocos_status)
        + "</div></div>"
        '<div class="gestao-card"><h4>Percentual de Absorção por Nível</h4>'
        '<div class="gestao-absorcao">'
        + "".join(barras_absorcao)
        + "</div></div></div>",
        unsafe_allow_html=True,
    )

    niveis_disponiveis = sorted(
        [
            nivel
            for nivel in dados_gestao_mes["Nível"].dropna().unique()
            if builtins.str(nivel).strip()
        ]
    )
    nivel_ranking_geral = st.selectbox(
        "Selecione o nível do ranking geral",
        ["Todos os níveis"] + niveis_disponiveis,
        key="gestao_ranking_nivel",
    )

    base_ranking_geral = dados_gestao_mes.copy()
    if nivel_ranking_geral != "Todos os níveis":
        base_ranking_geral = base_ranking_geral[
            base_ranking_geral["Nível"] == nivel_ranking_geral
        ]

    ranking_geral_mensal = (
        base_ranking_geral.groupby("Técnico", as_index=False)[["Realizado", "Esperado"]]
        .sum()
        .reset_index(drop=True)
    )
    ranking_geral_mensal["Produtividade"] = ranking_geral_mensal.apply(
        lambda linha: (
            (linha["Realizado"] / linha["Esperado"]) * 100
            if linha["Esperado"] > 0
            else 0
        ),
        axis=1,
    )
    ranking_geral_mensal = ranking_geral_mensal.sort_values(
        by="Produtividade",
        ascending=False,
    ).head(10)

    def abreviar_nome_gestao(nome):
        partes = builtins.str(nome or "").title().split()
        if len(partes) <= 2:
            return "<br>".join(partes)
        return f"{partes[0]}<br>{partes[-1]}"

    ranking_geral_mensal["Técnico Exibição"] = ranking_geral_mensal["Técnico"].map(
        abreviar_nome_gestao
    )
    ranking_geral_mensal["Cor"] = [
        COR_LARANJA if indice == 0 else "#4B5563"
        for indice in range(len(ranking_geral_mensal))
    ]
    ranking_geral_mensal["Rótulo"] = ranking_geral_mensal["Produtividade"].map(
        lambda valor: formatar_percentual_gestao(valor, 0)
    )

    grafico_geral = go.Figure()
    grafico_geral.add_bar(
        x=ranking_geral_mensal["Técnico Exibição"],
        y=ranking_geral_mensal["Produtividade"],
        marker_color=ranking_geral_mensal["Cor"],
        text=ranking_geral_mensal["Rótulo"],
        textposition="outside",
        hovertemplate="<b>%{x}</b><br>Produtividade: %{y:.2f}%<extra></extra>",
    )
    grafico_geral.add_hline(
        y=100,
        line_dash="dash",
        line_color="#9CA3AF",
        annotation_text="Meta (100%)",
        annotation_position="top right",
    )
    grafico_geral.update_layout(
        title="Ranking Geral",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=10, r=10, t=52, b=10),
        height=320,
        showlegend=False,
        xaxis=dict(title="", tickfont=dict(size=11)),
        yaxis=dict(
            title="Produtividade (%)",
            gridcolor="#E5E7EB",
            zeroline=False,
        ),
    )

    tecnicos_abaixo_meta = int(
        status_atual["Classificação"].isin(["CRÍTICO", "ATENÇÃO"]).sum()
    )
    tecnicos_em_atencao = int((status_atual["Classificação"] == "ATENÇÃO").sum())
    hoje_gestao = datetime.now(FUSO_HORARIO_APP).date()
    ferias_hoje_set = {
        normalizar_nome(nome)
        for nome, inicio, fim in PROGRAMACAO_FERIAS
        if inicio <= hoje_gestao <= fim
    }
    licencas_hoje_set = {
        normalizar_nome(nome)
        for nome, inicio, fim, _ in PROGRAMACAO_LICENCAS
        if inicio <= hoje_gestao <= fim
    }
    ausencias_integral_hoje = {
        normalizar_nome(nome)
        for nome, data_ausencia, _, minutos in PROGRAMACAO_AUSENCIAS
        if data_ausencia == hoje_gestao and minutos is None
    }
    ausencias_parciais_hoje = {
        normalizar_nome(nome)
        for nome, data_ausencia, _, minutos in PROGRAMACAO_AUSENCIAS
        if data_ausencia == hoje_gestao and minutos is not None
    }
    proximas_ferias = sum(
        1
        for _, inicio, _ in PROGRAMACAO_FERIAS
        if hoje_gestao <= inicio <= hoje_gestao + timedelta(days=15)
    )
    tecnicos_base_hoje = {
        normalizar_nome(nome)
        for nome in status_atual["Técnico"].dropna().tolist()
    }
    afastados_integral_hoje = licencas_hoje_set | ausencias_integral_hoje
    trabalhando_hoje = max(
        len(tecnicos_base_hoje - ferias_hoje_set - afastados_integral_hoje),
        0,
    )

    try:
        lista_apoio_resumo = ler_lista_apoio().copy()
        lista_apoio_resumo["Carimbo_dt"] = pd.to_datetime(
            lista_apoio_resumo["Carimbo de data/hora"],
            format="%d/%m/%Y %H:%M:%S",
            errors="coerce",
        )
        lista_aberta = lista_apoio_resumo[
            ~lista_apoio_resumo["Situação"].isin(
                ["Finalizado", "Resolvido pelo técnico"]
            )
        ].copy()
        agora_apoio = datetime.now(FUSO_HORARIO_APP).replace(tzinfo=None)
        lista_aberta["Horas em aberto"] = (
            agora_apoio - lista_aberta["Carimbo_dt"]
        ).dt.total_seconds().div(3600).fillna(0)
        ajudas_criticas = int((lista_aberta["Horas em aberto"] > 48).sum())
        ajudas_em_analise = int(
            (
                (lista_aberta["Situação"] == "Em análise")
                & (lista_aberta["Horas em aberto"] <= 48)
            ).sum()
        )
        ajudas_aguardando = int(
            (
                (lista_aberta["Situação"] == "Aberto")
                & (lista_aberta["Horas em aberto"] <= 48)
            ).sum()
        )
        ajudas_em_aberto = int(len(lista_aberta))
    except Exception:
        ajudas_criticas = 0
        ajudas_em_analise = 0
        ajudas_aguardando = 0
        ajudas_em_aberto = 0

    alertas = [
        ("#DC2626", "!", f"{tecnicos_abaixo_meta} técnicos abaixo da meta"),
        ("#F97316", "!", f"{tecnicos_em_atencao} técnicos entre atenção e meta"),
        ("#2563EB", "□", f"{proximas_ferias} técnicos com férias nos próximos 15 dias"),
        ("#7C3AED", "◌", f"{ajudas_em_aberto} ajudas em aberto"),
    ]
    html_alertas = "".join(
        '<div class="gestao-alerta">'
        '<div class="gestao-alerta-label">'
        f'<div class="gestao-alerta-icone" style="background:{cor};">{icone}</div>'
        f"<span>{escape(texto)}</span>"
        '</div><div class="gestao-alerta-seta">›</div></div>'
        for cor, icone, texto in alertas
    )

    col_ranking, col_alertas = st.columns([1.85, 1.25])
    with col_ranking:
        st.plotly_chart(
            grafico_geral,
            use_container_width=True,
            config={"displayModeBar": False},
        )
    with col_alertas:
        st.markdown(
            '<div class="gestao-card"><h4>Alertas do mês</h4>'
            f'{html_alertas}</div>',
            unsafe_allow_html=True,
        )

    st.markdown(
        f"""
        <div class="gestao-resumo-faixa">
            <div class="gestao-resumo-card">
                <h4>Disponibilidade hoje ({hoje_gestao.strftime('%d/%m/%Y')})</h4>
                <div class="gestao-resumo-grid">
                    <div class="gestao-resumo-item">
                        <div class="gestao-resumo-bola" style="color:#16A34A;border:1px solid #16A34A;">○</div>
                        <div class="gestao-resumo-texto"><strong>{trabalhando_hoje}</strong><span>Trabalhando</span></div>
                    </div>
                    <div class="gestao-resumo-item">
                        <div class="gestao-resumo-bola" style="color:#F97316;border:1px solid #F97316;">◔</div>
                        <div class="gestao-resumo-texto"><strong>{len(ferias_hoje_set)}</strong><span>Férias</span></div>
                    </div>
                    <div class="gestao-resumo-item">
                        <div class="gestao-resumo-bola" style="color:#2563EB;border:1px solid #2563EB;">＋</div>
                        <div class="gestao-resumo-texto"><strong>{len(afastados_integral_hoje)}</strong><span>Afastados</span></div>
                    </div>
                    <div class="gestao-resumo-item">
                        <div class="gestao-resumo-bola" style="color:#7C3AED;border:1px solid #7C3AED;">◷</div>
                        <div class="gestao-resumo-texto"><strong>{len(ausencias_parciais_hoje)}</strong><span>Saída antecipada</span></div>
                    </div>
                </div>
            </div>
            <div class="gestao-resumo-card">
                <h4>Ajudas em aberto</h4>
                <div class="gestao-resumo-grid">
                    <div class="gestao-resumo-item">
                        <div class="gestao-resumo-bola" style="color:#DC2626;border:1px solid #DC2626;">!</div>
                        <div class="gestao-resumo-texto"><strong>{ajudas_criticas}</strong><span>Críticas (&gt; 48h)</span></div>
                    </div>
                    <div class="gestao-resumo-item">
                        <div class="gestao-resumo-bola" style="color:#F97316;border:1px solid #F97316;">⌕</div>
                        <div class="gestao-resumo-texto"><strong>{ajudas_em_analise}</strong><span>Em análise</span></div>
                    </div>
                    <div class="gestao-resumo-item">
                        <div class="gestao-resumo-bola" style="color:#2563EB;border:1px solid #2563EB;">◌</div>
                        <div class="gestao-resumo-texto"><strong>{ajudas_aguardando}</strong><span>Aguardando apoio</span></div>
                    </div>
                    <div class="gestao-resumo-item">
                        <div class="gestao-resumo-bola" style="color:#111827;border:1px solid #CBD5E1;">#</div>
                        <div class="gestao-resumo-texto"><strong>{ajudas_em_aberto}</strong><span>Total</span></div>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.stop()

if modo_gestao and visao_gestao == "Resultados Individuais":
    tecnico = st.selectbox(
        "Selecione o Técnico",
        tecnicos_disponiveis,
        index=tecnicos_disponiveis.index(tecnico),
        key="tecnico_gestao",
    )

st.divider()
st.subheader("Resultados Individuais")

nivel_mode = dados_mes_atual["Nível"].dropna().mode()
nivel_tecnico = None if nivel_mode.empty else nivel_mode.iloc[0]

coluna_votacao = nome_coluna_dataframe(
    dados_ultimo_dia_votacao,
    "Votação",
    "VotaÃ§Ã£o",
    "Vota??o",
)
votacao_ultimo_dia = round(dados_ultimo_dia_votacao[coluna_votacao].mean(), 2)

coluna_satisfacao = nome_coluna_dataframe(
    dados_ultimo_dia_satisfacao,
    "Satisfação",
    "SatisfaÃ§Ã£o",
    "Satisfa??o",
)
satisfacao_ultimo_dia = round(
    dados_ultimo_dia_satisfacao[coluna_satisfacao].mean(),
    2,
)

coluna_classificacao = nome_coluna_dataframe(
    dados_mes_atual,
    "Classificação",
    "ClassificaÃ§Ã£o",
    "Classifica??o",
)
classificacao_mode = dados_mes_atual[coluna_classificacao].dropna().mode()
classificacao = (
    "Sem classificação"
    if classificacao_mode.empty
    else classificacao_mode.iloc[0]
)

atendidas_total = int(dados_mes_atual["Atendidas"].sum())
realizado_total = int(dados_mes_atual["Realizado"].sum())
esperado_total = int(dados_mes_atual["Esperado"].sum())
desvio_total = realizado_total - esperado_total
ssc_total = int(dados_mes_atual["SSC"].sum())
ro_total = int(dados_mes_atual["RO"].sum())
progresso_meta = (
    (realizado_total / esperado_total) * 100
    if esperado_total > 0
    else 0
)
largura_progresso = min(max(progresso_meta, 0), 100)
cor_desvio = "#16A34A" if desvio_total >= 0 else "#DC2626"
cor_progresso = "#16A34A" if progresso_meta >= 100 else COR_LARANJA
cor_classificacao = CORES_STATUS.get(classificacao, "#6B7280")
nome_exibicao = escape(tecnico.title())
nivel_exibicao = escape(builtins.str(nivel_tecnico or "Nível não informado"))
classificacao_exibicao = escape(builtins.str(classificacao))
sinal_desvio = "+" if desvio_total > 0 else ""
cor_votacao = "#DC2626" if votacao_ultimo_dia < 21 else "#111827"
cor_satisfacao = "#DC2626" if satisfacao_ultimo_dia < 98 else "#111827"

st.markdown(
    f"""
    <div class="resultado-cabecalho">
        <div class="resultado-identidade">
            <div class="resultado-nome">{nome_exibicao}</div>
            <div class="resultado-nivel">{nivel_exibicao}</div>
        </div>
        <div class="resultado-status" style="background:{cor_classificacao};">
            {classificacao_exibicao}
        </div>
    </div>
    <div class="resultado-grid-principal">
        <div class="resultado-card destaque">
            <div class="resultado-label">Realizado</div>
            <div class="resultado-valor">{realizado_total}</div>
        </div>
        <div class="resultado-card destaque">
            <div class="resultado-label">Esperado</div>
            <div class="resultado-valor">{esperado_total}</div>
        </div>
        <div class="resultado-card destaque">
            <div class="resultado-label">Desvio</div>
            <div class="resultado-valor" style="color:{cor_desvio};">
                {sinal_desvio}{desvio_total}
            </div>
        </div>
    </div>
    <div class="resultado-progresso-bloco">
        <div class="resultado-progresso-texto">
            <span>Progresso da meta mensal</span>
            <strong>{progresso_meta:.0f}% da meta</strong>
        </div>
        <div class="resultado-progresso-trilho">
            <div class="resultado-progresso-barra"
                 style="width:{largura_progresso:.2f}%;background:{cor_progresso};">
            </div>
        </div>
    </div>
    <div class="resultado-grid-secundario">
        <div class="resultado-card secundario">
            <div class="resultado-label">Atendidas</div>
            <div class="resultado-valor">{atendidas_total}</div>
        </div>
        <div class="resultado-card secundario">
            <div class="resultado-label">SSC</div>
            <div class="resultado-valor">{ssc_total}</div>
        </div>
        <div class="resultado-card secundario">
            <div class="resultado-label">RO</div>
            <div class="resultado-valor">{ro_total}</div>
        </div>
        <div class="resultado-card secundario">
            <div class="resultado-label">Votação Média</div>
            <div class="resultado-valor" style="color:{cor_votacao};">
                {votacao_ultimo_dia:.2f}%
            </div>
        </div>
        <div class="resultado-card secundario">
            <div class="resultado-label">Satisfação</div>
            <div class="resultado-valor" style="color:{cor_satisfacao};">
                {satisfacao_ultimo_dia:.2f}%
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

if not modo_gestao:
    percentual_absorcao = percentual_absorcao_tecnico(
        df,
        ano_atual,
        mes_atual,
        nivel_tecnico,
    )
    dias_base_calendario = float(dias_uteis_para_meta(ano_atual, mes_atual))
    dias_meta_valor = dias_base_calendario
    dias_ferias_programadas = dias_uteis_ferias_no_mes(
        tecnico,
        ano_atual,
        mes_atual,
    )
    dias_licenca_programados = dias_uteis_licenca_no_mes(
        tecnico,
        ano_atual,
        mes_atual,
    )
    abatimento_ausencias_programadas = abatimento_ausencias_no_mes(
        tecnico,
        ano_atual,
        mes_atual,
    )
    dias_meta_programados = max(
        (
            dias_meta_valor
            - dias_ferias_programadas
            - dias_licenca_programados
            - abatimento_ausencias_programadas
        ),
        0,
    )
    dias_meta_valor = dias_meta_programados
    fonte_dias_meta = "Calendário do mês"
    dias_segunda_sexta = sum(
        1
        for dia in range(
            1,
            calendar.monthrange(ano_atual, mes_atual)[1] + 1,
        )
        if date(ano_atual, mes_atual, dia).weekday() < 5
    )
    feriados_em_dias_uteis = int(dias_segunda_sexta - dias_base_calendario)
    detalhe_dias_meta = (
        f"{dias_segunda_sexta} dias de segunda a sexta"
        f" - {feriados_em_dias_uteis} feriado(s)"
        f" - {dias_ferias_programadas} dia(s) de férias"
        f" - {dias_licenca_programados} dia(s) de licença"
        f" - {abatimento_ausencias_programadas:g} dia(s) de ausência"
        f" = {dias_meta_valor:g} dias."
    )
    if COLUNA_DIAS_META in dados_tecnico.columns:
        dias_meta_salvos = (
            dados_tecnico[
                (dados_tecnico["Data"].dt.month == mes_atual)
                & (dados_tecnico["Data"].dt.year == ano_atual)
            ][COLUNA_DIAS_META]
            .dropna()
        )
        if not dias_meta_salvos.empty:
            dias_meta_valor = min(
                float(dias_meta_salvos.iloc[-1]),
                dias_meta_programados,
            )
            fonte_dias_meta = "Planilha"
            detalhe_dias_meta = "Valor atualizado pela gestão a partir do PontoWeb."
    if dias_meta_valor.is_integer():
        dias_meta_exibicao = builtins.str(int(dias_meta_valor))
    else:
        dias_meta_exibicao = f"{dias_meta_valor:.2f}".replace(".", ",")

    st.markdown(
        f"""
        <div class="resultado-faixa">
            <div class="resultado-card secundario">
                <div class="resultado-label">Percentual de Absorção</div>
                <div class="resultado-valor">{percentual_absorcao:.2f}%</div>
            </div>
            <div class="resultado-card secundario">
                <div class="resultado-label">Dias Úteis para Meta</div>
                <div class="resultado-valor">{dias_meta_exibicao}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.caption(f"Fonte: {fonte_dias_meta}. {detalhe_dias_meta}")

    ferias_ativa = ferias_ativa_tecnico(tecnico)
    if ferias_ativa:
        inicio_ferias, fim_ferias = ferias_ativa
        st.info(
            "Em férias: "
            f"{inicio_ferias.strftime('%d/%m/%Y')} a "
            f"{fim_ferias.strftime('%d/%m/%Y')}"
        )

    licenca_ativa = licenca_ativa_tecnico(tecnico)
    if licenca_ativa:
        inicio_licenca, fim_licenca, motivo_licenca = licenca_ativa
        st.info(
            f"{motivo_licenca}: "
            f"{inicio_licenca.strftime('%d/%m/%Y')} a "
            f"{fim_licenca.strftime('%d/%m/%Y')}"
        )

    st.caption("Referência da classificação")
    col_leg_1, col_leg_2, col_leg_3, col_leg_4 = st.columns(4)
    for status, descricao, coluna in zip(
        ["CRÍTICO", "ATENÇÃO", "BOM", "EXCELENTE"],
        [
            "Desvio menor que -5",
            "Desvio entre -5 e menor que 0",
            "Desvio entre 0 e 5",
            "Desvio acima de 5",
        ],
        [col_leg_1, col_leg_2, col_leg_3, col_leg_4],
    ):
        with coluna:
            st.markdown(
                f"""
                <div style="
                    border:1px solid #E5E7EB;
                    border-left:6px solid {CORES_STATUS[status]};
                    border-radius:10px;
                    padding:10px 12px;
                    background-color:#FFFFFF;
                    min-height:74px;
                ">
                    <div style="font-weight:700;color:#111827;font-size:13px;">
                        {status}
                    </div>
                    <div style="color:#6B7280;font-size:12px;margin-top:4px;">
                        {descricao}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.divider()
    st.subheader("Lista de Apoio")
    col_atualizar_apoio_tecnico, col_status_apoio_tecnico = st.columns([1, 3])
    with col_atualizar_apoio_tecnico:
        if st.button("Atualizar minhas ajudas", key="atualizar_lista_apoio_tecnico"):
            st.rerun()
    with col_status_apoio_tecnico:
        st.caption("Use o botão para atualizar suas ajudas sem sair do painel.")

    with st.form("form_lista_apoio", clear_on_submit=True):
        topico_apoio = st.selectbox("Tópico do problema", TOPICOS_LISTA_APOIO)
        resumo_apoio = st.text_area(
            "Descreva a dúvida de forma resumida",
            height=110,
        )
        enviar_apoio = st.form_submit_button("Registrar dúvida")

    if enviar_apoio:
        if not resumo_apoio.strip():
            st.warning("Descreva a dúvida antes de registrar.")
        else:
            try:
                registrar_duvida_apoio(
                    tecnico,
                    topico_apoio,
                    resumo_apoio,
                )
            except PermissionError:
                st.error("Feche a lista_apoio.xlsx e tente registrar novamente.")
            except Exception as erro_lista_apoio:
                st.error(f"Não foi possível registrar a dúvida: {erro_lista_apoio}")
            else:
                st.success("Dúvida registrada para apoio.")

    try:
        lista_apoio_tecnico = ler_lista_apoio()
        lista_apoio_tecnico = lista_apoio_tecnico[
            lista_apoio_tecnico["Nome do Técnico"].map(normalizar_nome) == normalizar_nome(tecnico)
        ].copy()
    except Exception as erro_lista_apoio:
        st.caption(f"Não foi possível carregar suas ajudas registradas: {erro_lista_apoio}")
        lista_apoio_tecnico = pd.DataFrame(columns=CABECALHOS_LISTA_APOIO)

    if lista_apoio_tecnico.empty:
        st.caption("Você ainda não possui ajudas registradas.")
    else:
        lista_apoio_tecnico["Data"] = pd.to_datetime(
            lista_apoio_tecnico["Carimbo de data/hora"],
            dayfirst=True,
            errors="coerce",
        )
        lista_apoio_tecnico = lista_apoio_tecnico.sort_values(
            by="Data",
            ascending=False,
            na_position="last",
        )
        lista_apoio_tecnico["Carimbo de data/hora"] = lista_apoio_tecnico[
            "Carimbo de data/hora"
        ].fillna("")

        st.caption("Acompanhe abaixo a situação das ajudas que você registrou.")
        st.dataframe(
            lista_apoio_tecnico[
                [
                    "Carimbo de data/hora",
                    "Selecione o tópico",
                    "Descreva o problema/pedido em poucas palavras",
                    "Situação",
                    "Responsável/Apoio",
                ]
            ],
            use_container_width=True,
            hide_index=True,
        )

if not modo_gestao:
    if nivel_tecnico:
        base_ranking_nivel = df[
            (df["Data"].dt.month == mes_atual)
            & (df["Data"].dt.year == ano_atual)
            & (df["Nível"] == nivel_tecnico)
        ]

        ranking_nivel_diario = (
            base_ranking_nivel[base_ranking_nivel["Data"] == ultima_data_produtividade]
            .groupby("Técnico", as_index=False)["Realizado"]
            .sum()
            .sort_values(by="Realizado", ascending=False)
            .reset_index(drop=True)
        )

        ranking_nivel_mensal = (
            base_ranking_nivel.groupby("Técnico", as_index=False)["Realizado"]
            .sum()
            .sort_values(by="Realizado", ascending=False)
            .reset_index(drop=True)
        )
        meta_diaria_nivel = meta_esperada_nivel(nivel_tecnico)
        datas_validas_ranking = {
            pd.Timestamp(data_linha).date()
            for data_linha in base_ranking_nivel["Data"].dropna().unique()
            if pd.Timestamp(data_linha).date() <= ultima_data_produtividade.date()
            and pd.Timestamp(data_linha).weekday() < 5
            and not eh_feriado_federal(pd.Timestamp(data_linha).date())
        }
        meta_mensal_nivel = meta_diaria_nivel * len(datas_validas_ranking)

        st.divider()
        st.subheader(f"Ranking do Seu Nível - {nivel_tecnico}")
        aba_nivel_diario, aba_nivel_mensal = st.tabs(["Diário", "Mensal"])

        with aba_nivel_diario:
            st.plotly_chart(
                montar_grafico_ranking(
                    ranking_nivel_diario,
                    f"Ranking Diário - {nivel_tecnico}",
                    "Produtividade do dia",
                    meta_referencia=meta_diaria_nivel,
                ),
                use_container_width=True,
            )

        with aba_nivel_mensal:
            st.plotly_chart(
                montar_grafico_ranking(
                    ranking_nivel_mensal,
                    f"Ranking Mensal - {nivel_tecnico}",
                    "Produtividade do mês",
                    meta_referencia=meta_mensal_nivel,
                ),
                use_container_width=True,
            )

st.divider()
st.subheader("📊 Produtividade Mensal")

dados_produtividade = dados_tecnico[
    (dados_tecnico["Data"].dt.month == mes_atual)
    & (dados_tecnico["Data"].dt.year == ano_atual)
]

data_inicial_padrao = dados_produtividade["Data"].min()
data_final_padrao = dados_produtividade["Data"].max()

col_periodo_1, col_periodo_2 = st.columns(2)
with col_periodo_1:
    periodo_inicial_texto = st.text_input(
        "Período inicial",
        value=data_inicial_padrao.strftime("%d/%m/%Y"),
    )
with col_periodo_2:
    periodo_final_texto = st.text_input(
        "Período final",
        value=data_final_padrao.strftime("%d/%m/%Y"),
    )

try:
    periodo_inicial = datetime.strptime(periodo_inicial_texto, "%d/%m/%Y").date()
    periodo_final = datetime.strptime(periodo_final_texto, "%d/%m/%Y").date()
except ValueError:
    st.error("Informe o período no formato dd/mm/aaaa.")
    st.stop()

if periodo_inicial > periodo_final:
    st.error("O período inicial não pode ser maior que o período final.")
    st.stop()

dados_produtividade = dados_produtividade[
    (dados_produtividade["Data"].dt.date >= periodo_inicial)
    & (dados_produtividade["Data"].dt.date <= periodo_final)
]
dados_produtividade = dados_produtividade[
    (dados_produtividade["Data"].dt.weekday < 5)
    & ~dados_produtividade["Data"].dt.date.map(eh_feriado_federal)
]

produtividade = (
    dados_produtividade.groupby(["Data", "Data Formatada"])
    .agg({"Realizado": "sum", "Esperado": "sum"})
    .reset_index()
    .sort_values(by="Data")
)
produtividade["Data Curta"] = produtividade["Data"].dt.strftime("%d/%m")
produtividade["Desvio Diário"] = (
    produtividade["Realizado"] - produtividade["Esperado"]
)
produtividade["Situação Meta"] = produtividade["Desvio Diário"].apply(
    lambda desvio: (
        "Acima da meta"
        if desvio > 0
        else "Na meta"
        if desvio == 0
        else "Abaixo da meta"
    )
)

dias_acima_meta = int((produtividade["Desvio Diário"] > 0).sum())
dias_na_meta = int((produtividade["Desvio Diário"] == 0).sum())
dias_abaixo_meta = int((produtividade["Desvio Diário"] < 0).sum())

st.markdown(
    f"""
    <div class="produtividade-resumo">
        <div class="produtividade-resumo-item">
            <span class="produtividade-resumo-ponto" style="background:#16A34A;"></span>
            <strong>{dias_acima_meta}</strong> dia(s) acima da meta
        </div>
        <div class="produtividade-resumo-item">
            <span class="produtividade-resumo-ponto" style="background:#6B7280;"></span>
            <strong>{dias_na_meta}</strong> dia(s) na meta
        </div>
        <div class="produtividade-resumo-item">
            <span class="produtividade-resumo-ponto" style="background:#DC2626;"></span>
            <strong>{dias_abaixo_meta}</strong> dia(s) abaixo da meta
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

cores_realizado = produtividade["Desvio Diário"].apply(
    lambda desvio: "#16A34A" if desvio >= 0 else "#EF4444"
).tolist()
rotulos_desvio = produtividade["Desvio Diário"].apply(
    lambda desvio: f"+{int(desvio)}" if desvio > 0 else f"{int(desvio)}"
).tolist()
cores_desvio = produtividade["Desvio Diário"].apply(
    lambda desvio: "#15803D" if desvio >= 0 else "#B91C1C"
).tolist()

grafico_produtividade = go.Figure()
grafico_produtividade.add_trace(
    go.Bar(
        name="Realizado",
        x=produtividade["Data Curta"],
        y=produtividade["Realizado"],
        marker_color=cores_realizado,
        width=0.48,
        text=produtividade["Realizado"].round().astype(int),
        textposition="outside",
        textfont=dict(color="#374151", size=12),
        cliponaxis=False,
        customdata=produtividade[
            ["Esperado", "Desvio Diário", "Situação Meta"]
        ].to_numpy(),
        hovertemplate=(
            "<b>%{x}</b><br>"
            "Realizado: %{y}<br>"
            "Esperado: %{customdata[0]}<br>"
            "Desvio: %{customdata[1]:+}<br>"
            "%{customdata[2]}<extra></extra>"
        ),
    )
)
grafico_produtividade.add_trace(
    go.Scatter(
        name="Esperado",
        x=produtividade["Data Curta"],
        y=produtividade["Esperado"],
        mode="lines+markers",
        line=dict(color="#4B5563", width=2),
        marker=dict(color="#FFFFFF", line=dict(color="#4B5563", width=2), size=7),
        text=produtividade["Esperado"].round().astype(int),
        hovertemplate="<b>%{x}</b><br>Esperado: %{y}<extra></extra>",
    )
)

maior_valor_produtividade = max(
    produtividade["Realizado"].max() if not produtividade.empty else 0,
    produtividade["Esperado"].max() if not produtividade.empty else 0,
    1,
)
espaco_desvio = max(maior_valor_produtividade * 0.1, 2)
for indice, linha in produtividade.reset_index(drop=True).iterrows():
    grafico_produtividade.add_annotation(
        x=linha["Data Curta"],
        y=max(linha["Realizado"], linha["Esperado"]) + espaco_desvio,
        text=rotulos_desvio[indice],
        showarrow=False,
        font=dict(color=cores_desvio[indice], size=11),
    )

grafico_produtividade.update_layout(
    plot_bgcolor=COR_BRANCO,
    paper_bgcolor=COR_BRANCO,
    font_color="#4B5563",
    height=430,
    margin=dict(l=45, r=20, t=55, b=55),
    bargap=0.45,
    xaxis_title="Dia",
    yaxis_title="Quantidade",
    title=dict(
        text="Produtividade do Mês",
        x=0,
        font=dict(size=15, color="#111827"),
    ),
    legend=dict(
        orientation="h",
        yanchor="bottom",
        y=1.02,
        xanchor="right",
        x=1,
        title_text="",
    ),
    xaxis=dict(
        type="category",
        showgrid=False,
        linecolor="#E5E7EB",
        tickfont=dict(size=11, color="#4B5563"),
    ),
    yaxis=dict(
        showgrid=True,
        gridcolor="#F3F4F6",
        gridwidth=1,
        zeroline=False,
        rangemode="tozero",
        range=[0, maior_valor_produtividade + (espaco_desvio * 2.5)],
        tickfont=dict(size=11, color="#6B7280"),
    ),
    hovermode="x unified",
)

st.plotly_chart(grafico_produtividade, use_container_width=True)

st.divider()
st.subheader("SSC Atendido por Dia")

ssc_diario = (
    dados_produtividade.groupby(["Data", "Data Formatada"])["SSC"]
    .sum()
    .reset_index()
    .sort_values(by="Data")
)
ssc_diario["Data Curta"] = ssc_diario["Data"].dt.strftime("%d/%m")
ssc_total_periodo = int(round(ssc_diario["SSC"].sum()))
ssc_media_diaria = float(ssc_diario["SSC"].mean()) if not ssc_diario.empty else 0

if ssc_diario.empty:
    melhor_data_ssc = "-"
    melhor_valor_ssc = 0
    indice_melhor_ssc = None
else:
    indice_melhor_ssc = ssc_diario["SSC"].idxmax()
    melhor_linha_ssc = ssc_diario.loc[indice_melhor_ssc]
    melhor_data_ssc = melhor_linha_ssc["Data Curta"]
    melhor_valor_ssc = int(round(melhor_linha_ssc["SSC"]))

st.markdown(
    f"""
    <div class="produtividade-resumo">
        <div class="produtividade-resumo-item">
            <span class="produtividade-resumo-ponto" style="background:#16A34A;"></span>
            <span>SSC total <strong>{ssc_total_periodo}</strong></span>
        </div>
        <div class="produtividade-resumo-item">
            <span class="produtividade-resumo-ponto" style="background:#6B7280;"></span>
            <span>Média diária <strong>{ssc_media_diaria:.0f}</strong></span>
        </div>
        <div class="produtividade-resumo-item">
            <span class="produtividade-resumo-ponto" style="background:#F97316;"></span>
            <span>Melhor dia <strong>{melhor_data_ssc} · {melhor_valor_ssc}</strong></span>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

grafico_ssc = go.Figure()
grafico_ssc.add_trace(
    go.Scatter(
        name="SSC",
        x=ssc_diario["Data Curta"],
        y=ssc_diario["SSC"],
        mode="lines+markers+text",
        line=dict(color="#15803D", width=2.5),
        marker=dict(
            color="#FFFFFF",
            size=8,
            line=dict(color="#15803D", width=2),
        ),
        fill="tozeroy",
        fillcolor="rgba(22, 163, 74, 0.10)",
        text=ssc_diario["SSC"].round().astype(int),
        textposition="top center",
        textfont=dict(color="#374151", size=11),
        cliponaxis=False,
        customdata=ssc_diario[["Data Formatada"]].to_numpy(),
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "SSC atendido: %{y}<br>"
            f"Média do período: {ssc_media_diaria:.0f}<extra></extra>"
        ),
    )
)

if indice_melhor_ssc is not None:
    melhor_linha_ssc = ssc_diario.loc[indice_melhor_ssc]
    grafico_ssc.add_trace(
        go.Scatter(
            name="Melhor dia",
            x=[melhor_linha_ssc["Data Curta"]],
            y=[melhor_linha_ssc["SSC"]],
            mode="markers",
            marker=dict(
                color=COR_LARANJA,
                size=12,
                line=dict(color="#FFFFFF", width=2),
            ),
            hoverinfo="skip",
            showlegend=False,
        )
    )
    grafico_ssc.add_annotation(
        x=melhor_linha_ssc["Data Curta"],
        y=melhor_linha_ssc["SSC"],
        text="Melhor dia",
        showarrow=True,
        arrowhead=0,
        arrowcolor=COR_LARANJA,
        ax=0,
        ay=-42,
        font=dict(color="#C2410C", size=11),
        bgcolor="#FFFFFF",
        bordercolor="#FED7AA",
        borderpad=4,
    )

if ssc_media_diaria > 0:
    grafico_ssc.add_hline(
        y=ssc_media_diaria,
        line_color="#9CA3AF",
        line_dash="dash",
        line_width=1.2,
        annotation_text=f"Média: {ssc_media_diaria:.0f}",
        annotation_position="top left",
        annotation_font_color="#6B7280",
        annotation_font_size=11,
        annotation_bgcolor="#FFFFFF",
    )

maior_ssc = max(
    ssc_diario["SSC"].max() if not ssc_diario.empty else 0,
    ssc_media_diaria,
    1,
)
grafico_ssc.update_layout(
    plot_bgcolor=COR_BRANCO,
    paper_bgcolor=COR_BRANCO,
    font_color="#4B5563",
    showlegend=False,
    height=410,
    margin=dict(l=45, r=20, t=38, b=55),
    xaxis_title="Dia",
    yaxis_title="SSC Atendido",
    xaxis=dict(
        type="category",
        showgrid=False,
        linecolor="#E5E7EB",
        tickfont=dict(size=11, color="#4B5563"),
    ),
    yaxis=dict(
        showgrid=True,
        gridcolor="#F3F4F6",
        gridwidth=1,
        zeroline=False,
        rangemode="tozero",
        range=[0, maior_ssc * 1.25],
        tickfont=dict(size=11, color="#6B7280"),
    ),
    hovermode="x unified",
)

st.plotly_chart(grafico_ssc, use_container_width=True)

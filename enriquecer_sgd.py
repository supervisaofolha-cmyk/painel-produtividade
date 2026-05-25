import argparse
import concurrent.futures
import json
import html
import re
import threading
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


DEFAULT_SRC = Path(r"C:\Users\esther.queiroz\Downloads\clientes_vs_abandonos 052026.xlsx")
DEFAULT_OUT = Path(r"C:\Users\esther.queiroz\Downloads\clientes_vs_abandonos 052026_enriquecido.xlsx")
PROFILE = Path(r"C:\Users\esther.queiroz\AppData\Local\Temp\sgd_playwright_profile")
SEARCH_URL = "https://sgd.dominiosistemas.com.br/sgsc/faces/loc-cliente.html"
LOGIN_URL_PART = "sgd.dominiosistemas.com.br/login"
CDP_URL = "http://127.0.0.1:9222"
DEFAULT_BATCH_DIR = Path(
    r"C:\Users\esther.queiroz\Downloads\lotes_clientes_vs_abandonos_052026"
)
thread_local = threading.local()

NEW_HEADERS = [
    "Nome_Licenciado",
    "Codigo_DECA",
    "SGD_Status",
    "SGD_Observacao",
    "Consultado_Em",
]


def digits(value):
    return re.sub(r"\D+", "", str(value or ""))


def ensure_workbook(src_path, out_path):
    if out_path.exists():
        wb = load_workbook(out_path)
    else:
        wb = load_workbook(src_path)

    ws = wb["Resultado da consulta"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    for header in NEW_HEADERS:
        if header not in headers:
            ws.cell(1, ws.max_column + 1).value = header
            headers.append(header)

    wb.save(out_path)
    return wb, ws


def header_map(ws):
    return {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}


def select_pending_rows(ws, cols, start_row, limit):
    rows = []
    for row in range(start_row, ws.max_row + 1):
        phone = digits(ws.cell(row, 1).value)
        if not phone:
            continue
        if ws.cell(row, cols["SGD_Status"]).value:
            continue
        rows.append((row, phone))
        if limit and len(rows) >= limit:
            break
    return rows


def next_batch_number(batch_dir):
    existing = []
    for path in batch_dir.glob("clientes_vs_abandonos_052026_lote_*.xlsx"):
        match = re.search(r"_lote_(\d+)\.xlsx$", path.name)
        if match:
            existing.append(int(match.group(1)))
    return (max(existing) + 1) if existing else 1


def export_batch_workbook(source_ws, batch_rows, batch_path):
    batch_wb = Workbook()
    batch_ws = batch_wb.active
    batch_ws.title = source_ws.title

    for col in range(1, source_ws.max_column + 1):
        batch_ws.cell(1, col).value = source_ws.cell(1, col).value

    for out_row, src_row in enumerate(batch_rows, start=2):
        for col in range(1, source_ws.max_column + 1):
            batch_ws.cell(out_row, col).value = source_ws.cell(src_row, col).value

    batch_wb.save(batch_path)


def summarize_batch(ws, cols, batch_rows):
    counts = {}
    for src_row in batch_rows:
        status = ws.cell(src_row, cols["SGD_Status"]).value or "Sem status"
        counts[status] = counts.get(status, 0) + 1
    return counts


def build_http_config(context, page):
    return {
        "cookies": context.cookies(),
        "user_agent": page.evaluate("() => navigator.userAgent"),
    }


def make_requests_session(http_config):
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": http_config["user_agent"],
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
        }
    )
    for cookie in http_config["cookies"]:
        session.cookies.set(
            cookie["name"],
            cookie["value"],
            domain=cookie.get("domain"),
            path=cookie.get("path", "/"),
        )
    return session


def get_thread_session(http_config):
    session = getattr(thread_local, "session", None)
    marker = getattr(thread_local, "http_marker", None)
    current_marker = (http_config["user_agent"], len(http_config["cookies"]))
    if session is None or marker != current_marker:
        session = make_requests_session(http_config)
        thread_local.session = session
        thread_local.http_marker = current_marker
    return session


def normalize_text(text):
    return re.sub(r"\s+", " ", text or "").strip()


def extract_from_text(text, phone):
    clean = normalize_text(text)
    zero_occurrence = re.search(r"Encontrada\(s\)\s*0\s*ocorr", clean, re.IGNORECASE)
    if zero_occurrence:
        return {
            "status": "Nao encontrado",
            "nome": "",
            "deca": "",
            "obs": "Nenhum resultado na pesquisa por telefone",
        }

    if "Nenhum registro" in clean or "não encontrado" in clean.lower():
        return {
            "status": "Nao encontrado",
            "nome": "",
            "deca": "",
            "obs": "Nenhum resultado na pesquisa por telefone",
        }

    deca_match = re.search(r"(?:DECA|Deca|deca)\D{0,20}(\d{2,})", clean)
    code_match = re.search(r"(?:Código|Codigo|Cód\.?|Cod\.?)\D{0,20}(\d{2,})", clean)

    deca = deca_match.group(1) if deca_match else ""
    code = code_match.group(1) if code_match else ""

    # Prefer labels commonly used by the client details page.
    name = ""
    for pattern in [
        r"(?:Licenciado|Cliente|Nome|Razão Social|Razao Social)\s*:?\s*([A-Z0-9][^:\n\r]{3,120})",
        r"\b\d{2,}\s+([A-Z][A-Z0-9 .,&/-]{5,120})",
    ]:
        match = re.search(pattern, clean)
        if match:
            name = normalize_text(match.group(1))
            break

    if deca or code or name:
        return {
            "status": "Encontrado",
            "nome": name,
            "deca": deca or code,
            "obs": "",
        }

    if phone in clean:
        return {
            "status": "Encontrado",
            "nome": "",
            "deca": "",
            "obs": "Resultado encontrado, mas campos nao identificados automaticamente",
        }

    return {
        "status": "Nao encontrado",
        "nome": "",
        "deca": "",
        "obs": "Telefone nao apareceu no resultado da pesquisa",
    }


def extract_result_rows(page):
    return page.eval_on_selector_all(
        "table.tableSorter tbody tr",
        """rows => rows.map(row => {
            const cells = Array.from(row.querySelectorAll('td')).map(td =>
                td.innerText.replace(/\\s+/g, ' ').trim()
            );
            const details = row.querySelector('a[href*="d-cliente.html?clienteID="]');
            return {
                codigo: cells[0] || '',
                razao: cells[1] || '',
                representante: cells[2] || '',
                tipo: cells[6] || '',
                telefone: cells[9] || '',
                detailsHref: details ? details.getAttribute('href') : ''
            };
        }).filter(row => row.codigo || row.razao)""",
    )


def extract_view_state(content):
    match = re.search(
        r'name="javax\.faces\.ViewState"[^>]*value="([^"]*)"', content
    )
    if not match:
        match = re.search(
            r'id="[^"]*:javax\.faces\.ViewState:[^"]*"[^>]*value="([^"]*)"',
            content,
        )
    return html.unescape(match.group(1)) if match else ""


def strip_tags(value):
    value = re.sub(r"<br\s*/?>", " ", value, flags=re.IGNORECASE)
    value = re.sub(r"<[^>]+>", " ", value)
    return normalize_text(html.unescape(value))


def extract_result_rows_from_html(content):
    table_match = re.search(
        r'<table[^>]+class="[^"]*\btableSorter\b[^"]*"[^>]*>.*?<tbody>(.*?)</tbody>',
        content,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not table_match:
        return []

    rows = []
    for row_html in re.findall(
        r"<tr[^>]*>(.*?)</tr>", table_match.group(1), flags=re.IGNORECASE | re.DOTALL
    ):
        cells = re.findall(
            r"<td[^>]*>(.*?)</td>", row_html, flags=re.IGNORECASE | re.DOTALL
        )
        if len(cells) < 10:
            continue
        rows.append(
            {
                "codigo": strip_tags(cells[0]),
                "razao": strip_tags(cells[1]),
                "representante": strip_tags(cells[2]),
                "tipo": strip_tags(cells[6]),
                "telefone": strip_tags(cells[9]),
            }
        )
    return [row for row in rows if row["codigo"] or row["razao"]]


def result_from_row(row):
    obs = []
    if row.get("representante"):
        obs.append(f"Representante: {row['representante']}")
    if row.get("telefone"):
        obs.append(f"Telefone suporte: {row['telefone']}")
    if row.get("tipo"):
        obs.append(f"Tipo: {row['tipo']}")
    return {
        "status": "Encontrado",
        "nome": row.get("razao", ""),
        "deca": row.get("codigo", ""),
        "obs": "; ".join(obs),
    }


def result_from_result_rows(result_rows, phone, debug_dir, row, content):
    if len(result_rows) == 1:
        return result_from_row(result_rows[0])
    if len(result_rows) > 1:
        phone_matches = [
            item for item in result_rows if digits(item.get("telefone")) == phone
        ]
        if len(phone_matches) == 1:
            return result_from_row(phone_matches[0])

        candidates = " | ".join(
            f"{item.get('codigo')} - {item.get('razao')}" for item in result_rows
        )
        debug_path = debug_dir / f"row_{row}_{phone}_multiplo.html"
        debug_path.write_text(content, encoding="utf-8")
        return {
            "status": "Multiplo resultado",
            "nome": "",
            "deca": "",
            "obs": f"{len(result_rows)} resultados: {candidates}; HTML: {debug_path}",
        }
    return None


def save_result(ws, cols, row, result):
    ws.cell(row, cols["Nome_Licenciado"]).value = result["nome"]
    ws.cell(row, cols["Codigo_DECA"]).value = result["deca"]
    ws.cell(row, cols["SGD_Status"]).value = result["status"]
    ws.cell(row, cols["SGD_Observacao"]).value = result["obs"]
    ws.cell(row, cols["Consultado_Em"]).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def wait_for_login(page):
    page.goto(SEARCH_URL, wait_until="load")
    if LOGIN_URL_PART not in page.url:
        return

    print("Chrome aberto na tela de login do SGD. Faça login nessa janela.")
    print("Depois do login, o script continua automaticamente.")
    deadline = time.time() + 600
    while time.time() < deadline:
        if LOGIN_URL_PART not in page.url:
            page.goto(SEARCH_URL, wait_until="load")
            return
        time.sleep(2)
    raise RuntimeError("Login nao concluido em 10 minutos.")


def goto_search(page):
    try:
        if "sgd.dominiosistemas.com.br/sgsc/faces/loc-cliente.html" not in page.url:
            page.goto(SEARCH_URL, wait_until="domcontentloaded", timeout=60000)
        else:
            page.wait_for_load_state("domcontentloaded", timeout=30000)
    except PlaywrightTimeoutError:
        page.wait_for_timeout(2000)


def search_phone(page, phone, debug_dir, row):
    goto_search(page)
    page.locator('select[name="locForm:usuario"]').select_option("5")
    page.locator('input[name="locForm:palavraChave"]').fill(phone)

    try:
        with page.expect_navigation(wait_until="domcontentloaded", timeout=60000):
            page.locator('input[name="locForm:localizarBtn"]').click(no_wait_after=True)
    except PlaywrightTimeoutError:
        page.wait_for_load_state("domcontentloaded", timeout=30000)
    page.wait_for_timeout(1500)

    result_rows = extract_result_rows(page)
    parsed_result = result_from_result_rows(
        result_rows, phone, debug_dir, row, page.content()
    )
    if parsed_result:
        return parsed_result

    body_text = page.locator("body").inner_text(timeout=10000)
    result = extract_from_text(body_text, phone)
    if result["status"] != "Encontrado" or not result["deca"] or not result["nome"]:
        debug_path = debug_dir / f"row_{row}_{phone}.html"
        debug_path.write_text(page.content(), encoding="utf-8")
        if result["obs"]:
            result["obs"] = f"{result['obs']}; HTML: {debug_path}"
        else:
            result["obs"] = f"HTML: {debug_path}"
    return result


def search_phone_http(request_context, view_state, phone, debug_dir, row):
    if not view_state:
        response = request_context.get(SEARCH_URL, timeout=60000)
        content = response.text()
        view_state = extract_view_state(content)

    data = {
        "locForm": "locForm",
        "origemRequest": "",
        "telefoneRequest": "",
        "conversationID": "",
        "locForm:segmento": "0",
        "locForm:usuario": "5",
        "locForm:palavraChave": phone,
        "locForm:localizarBtn": "Localizar",
        "javax.faces.ViewState": view_state,
    }
    response = request_context.post(SEARCH_URL, form=data, timeout=60000)
    content = response.text()
    next_view_state = extract_view_state(content)

    if LOGIN_URL_PART in response.url or "Domínio Sistemas - Login" in content:
        raise RuntimeError("Sessao expirada ou nao autenticada")

    result_rows = extract_result_rows_from_html(content)
    parsed_result = result_from_result_rows(
        result_rows, phone, debug_dir, row, content
    )
    if parsed_result:
        return parsed_result, next_view_state

    result = extract_from_text(strip_tags(content), phone)
    if result["status"] != "Encontrado" or not result["deca"] or not result["nome"]:
        debug_path = debug_dir / f"row_{row}_{phone}.html"
        debug_path.write_text(content, encoding="utf-8")
        if result["obs"]:
            result["obs"] = f"{result['obs']}; HTML: {debug_path}"
        else:
            result["obs"] = f"HTML: {debug_path}"
    return result, next_view_state


def search_phone_requests(http_config, phone, debug_dir, row):
    session = get_thread_session(http_config)
    response = session.get(SEARCH_URL, timeout=60)
    content = response.text
    view_state = extract_view_state(content)
    data = {
        "locForm": "locForm",
        "origemRequest": "",
        "telefoneRequest": "",
        "conversationID": "",
        "locForm:segmento": "0",
        "locForm:usuario": "5",
        "locForm:palavraChave": phone,
        "locForm:localizarBtn": "Localizar",
        "javax.faces.ViewState": view_state,
    }
    response = session.post(SEARCH_URL, data=data, timeout=60)
    content = response.text

    if LOGIN_URL_PART in response.url or "DomÃ­nio Sistemas - Login" in content:
        raise RuntimeError("Sessao expirada ou nao autenticada")

    result_rows = extract_result_rows_from_html(content)
    parsed_result = result_from_result_rows(
        result_rows, phone, debug_dir, row, content
    )
    if parsed_result:
        return parsed_result

    result = extract_from_text(strip_tags(content), phone)
    if result["status"] != "Encontrado" or not result["deca"] or not result["nome"]:
        debug_path = debug_dir / f"row_{row}_{phone}.html"
        debug_path.write_text(content, encoding="utf-8")
        if result["obs"]:
            result["obs"] = f"{result['obs']}; HTML: {debug_path}"
        else:
            result["obs"] = f"HTML: {debug_path}"
    return result


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SRC)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--batch-size", type=int, default=0)
    parser.add_argument("--batch-output-dir", type=Path, default=DEFAULT_BATCH_DIR)
    parser.add_argument("--start-row", type=int, default=2)
    parser.add_argument("--save-every", type=int, default=10)
    parser.add_argument("--connect-cdp", action="store_true")
    parser.add_argument("--workers", type=int, default=4)
    args = parser.parse_args()

    source_path = args.source
    output_path = args.output

    wb, ws = ensure_workbook(source_path, output_path)
    cols = header_map(ws)
    debug_dir = output_path.with_suffix("")
    debug_dir.mkdir(exist_ok=True)
    batch_dir = args.batch_output_dir
    batch_dir.mkdir(parents=True, exist_ok=True)

    effective_limit = args.batch_size or args.limit
    rows = select_pending_rows(ws, cols, args.start_row, effective_limit)

    print(f"Linhas pendentes selecionadas: {len(rows)}")
    if not rows:
        return 0

    batch_rows = [row for row, _phone in rows]
    batch_number = next_batch_number(batch_dir)
    batch_path = batch_dir / f"clientes_vs_abandonos_052026_lote_{batch_number:04d}.xlsx"

    with sync_playwright() as p:
        browser = None
        if args.connect_cdp:
            browser = p.chromium.connect_over_cdp(CDP_URL)
            context = browser.contexts[0]
            page = context.pages[0] if context.pages else context.new_page()
        else:
            context = p.chromium.launch_persistent_context(
                str(PROFILE),
                channel="chrome",
                headless=False,
                viewport={"width": 1280, "height": 800},
            )
            page = context.pages[0] if context.pages else context.new_page()
        wait_for_login(page)
        view_state = ""
        http_config = build_http_config(context, page) if args.connect_cdp else None

        done = 0
        if args.connect_cdp and args.workers > 1:
            future_map = {}
            with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as executor:
                for row, phone in rows:
                    print(f"Consultando linha {row}: {phone}")
                    future = executor.submit(
                        search_phone_requests, http_config, phone, debug_dir, row
                    )
                    future_map[future] = (row, phone)

                for future in concurrent.futures.as_completed(future_map):
                    row, phone = future_map[future]
                    try:
                        result = future.result()
                    except requests.Timeout as exc:
                        result = {
                            "status": "Erro",
                            "nome": "",
                            "deca": "",
                            "obs": f"Timeout na consulta: {exc}",
                        }
                    except Exception as exc:
                        result = {
                            "status": "Erro",
                            "nome": "",
                            "deca": "",
                            "obs": f"Erro na consulta: {exc}",
                        }

                    save_result(ws, cols, row, result)
                    done += 1
                    print(f"  -> linha {row}: {result['status']} | {result['nome']} | {result['deca']}")
                    if done % args.save_every == 0:
                        wb.save(output_path)
                        print(f"Progresso salvo em {output_path}")
        else:
            for row, phone in rows:
                try:
                    print(f"Consultando linha {row}: {phone}")
                    if args.connect_cdp:
                        result, view_state = search_phone_http(
                            context.request, view_state, phone, debug_dir, row
                        )
                    else:
                        result = search_phone(page, phone, debug_dir, row)
                except PlaywrightTimeoutError as exc:
                    result = {
                        "status": "Erro",
                        "nome": "",
                        "deca": "",
                        "obs": f"Timeout na consulta: {exc}",
                    }
                except Exception as exc:
                    result = {
                        "status": "Erro",
                        "nome": "",
                        "deca": "",
                        "obs": f"Erro na consulta: {exc}",
                    }

                save_result(ws, cols, row, result)
                done += 1
                print(f"  -> {result['status']} | {result['nome']} | {result['deca']}")
                if done % args.save_every == 0:
                    wb.save(output_path)
                    print(f"Progresso salvo em {output_path}")

        wb.save(output_path)
        export_batch_workbook(ws, batch_rows, batch_path)
        if browser:
            browser.close()
        else:
            context.close()

    summary = {
        "batch_number": batch_number,
        "batch_size": len(batch_rows),
        "row_start": batch_rows[0],
        "row_end": batch_rows[-1],
        "main_workbook": str(output_path),
        "batch_workbook": str(batch_path),
        "status_counts": summarize_batch(ws, cols, batch_rows),
    }
    print(f"Finalizado. Arquivo: {output_path}")
    print("BATCH_SUMMARY " + json.dumps(summary, ensure_ascii=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
